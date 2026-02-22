"""
DART 재무제표 추출 웹 서버

FastAPI를 사용한 백엔드 서버
- 기업 검색 API
- 재무제표 추출 API (진행상태 SSE)
- 파일 다운로드 API
- 작업 취소 API
"""

import os
import asyncio
import uuid
import threading
import math
import time
from datetime import datetime
from dotenv import load_dotenv

# .env 파일에서 환경변수 로드
load_dotenv()
from typing import Optional, Dict, Any
from contextlib import asynccontextmanager

from fastapi import FastAPI, HTTPException, BackgroundTasks, Request, Response, Depends, Cookie
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# 데이터베이스 모듈
import database as db
import json

import dart_fss as dart
import pandas as pd


# ============================================================
# 전역 변수: 작업 상태 관리
# ============================================================
# 작업 ID별 상태 저장: {task_id: {"status": "...", "progress": 0, "message": "...", "file_path": "...",
#                               "cancelled": False, "last_accessed": timestamp, "created_at": timestamp}}
TASKS: Dict[str, Dict[str, Any]] = {}
TASKS_LOCK = threading.Lock()

# 메모리 정리 설정
TASK_IDLE_TIMEOUT = 120  # 진행 중 작업: 2분간 status 조회 없으면 취소
TASK_COMPLETED_TTL = 300  # 완료된 작업: 5분 후 전체 삭제 (파일은 유지)
PREVIEW_DATA_TTL = 300   # preview_data: AI분석 미시작 시 5분 후 정리 (VCM 생성에 60초+ 소요되므로 여유 확보)

# 기업 리스트 캐시 (한 번만 로드)
CORP_LIST = None
CORP_LIST_LOCK = threading.Lock()

# KSIC 업종코드 매핑 (한국표준산업분류)
KSIC_CODES: Dict[str, str] = {}
try:
    ksic_path = os.path.join(os.path.dirname(__file__), 'ksic_codes.json')
    if os.path.exists(ksic_path):
        with open(ksic_path, 'r', encoding='utf-8') as f:
            KSIC_CODES = json.load(f)
        print(f"[INIT] KSIC 업종코드 {len(KSIC_CODES)}개 로드 완료")
except Exception as e:
    print(f"[INIT] KSIC 업종코드 로드 실패: {e}")


# ============================================================
# Pydantic 모델
# ============================================================
class SearchRequest(BaseModel):
    """기업 검색 요청"""
    company_name: str
    market: Optional[str] = None  # 'Y': 코스피, 'K': 코스닥, 'N': 코넥스, 'E': 기타


class ExtractRequest(BaseModel):
    """재무제표 추출 요청"""
    corp_code: str  # DART 고유번호
    corp_name: str  # 회사명 (파일명용)
    start_year: int = 2020
    end_year: Optional[int] = None
    company_info: Optional[Dict[str, Any]] = None  # 기업개황정보


class RegisterRequest(BaseModel):
    """회원가입 요청"""
    email: str
    password: str


class LoginRequest(BaseModel):
    """로그인 요청"""
    email: str
    password: str


# ============================================================
# 인증 관련 함수
# ============================================================
async def get_current_user(request: Request, session_token: Optional[str] = Cookie(None)) -> Optional[Dict]:
    """현재 로그인한 사용자 조회 (선택적 - 없으면 None 반환)"""
    if not session_token:
        return None
    session = db.get_session(session_token)
    if not session:
        return None
    return session


async def require_auth(request: Request, session_token: Optional[str] = Cookie(None)) -> Dict:
    """인증 필수 - 로그인 안 되어 있으면 에러"""
    if not session_token:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다")
    session = db.get_session(session_token)
    if not session:
        raise HTTPException(status_code=401, detail="세션이 만료되었습니다")
    return session


async def require_admin(request: Request, session_token: Optional[str] = Cookie(None)) -> Dict:
    """관리자 권한 필수"""
    session = await require_auth(request, session_token)
    if session.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다")
    return session


# ============================================================
# 유틸리티 함수
# ============================================================
def get_corp_list():
    """기업 리스트를 싱글톤으로 로드"""
    global CORP_LIST
    with CORP_LIST_LOCK:
        if CORP_LIST is None:
            api_key = os.environ.get('DART_API_KEY')
            if not api_key:
                raise ValueError("DART_API_KEY 환경변수가 설정되지 않았습니다.")
            dart.set_api_key(api_key)
            CORP_LIST = dart.get_corp_list()
        return CORP_LIST


def cleanup_task(task_id: str, force_delete: bool = False):
    """작업 정리 및 메모리 회수

    Args:
        task_id: 작업 ID
        force_delete: True면 파일 포함 완전 삭제, False면 preview_data만 정리
    """
    with TASKS_LOCK:
        if task_id not in TASKS:
            return

        task = TASKS[task_id]

        # 취소된 작업: 파일 삭제 및 TASKS에서 제거
        if task.get('cancelled') and task.get('file_path'):
            try:
                if os.path.exists(task['file_path']):
                    os.remove(task['file_path'])
                    print(f"[CLEANUP] 취소된 작업 파일 삭제: {task_id}")
            except Exception:
                pass

        if task.get('cancelled') or force_delete:
            del TASKS[task_id]
            print(f"[CLEANUP] 작업 삭제: {task_id}")
        else:
            # 완료된 작업: preview_data만 정리 (파일 경로는 유지)
            if 'preview_data' in task:
                del task['preview_data']
                print(f"[CLEANUP] preview_data 정리: {task_id}")


def cleanup_preview_data(task_id: str):
    """완료된 작업의 preview_data만 정리 (다운로드는 유지)"""
    with TASKS_LOCK:
        if task_id in TASKS and 'preview_data' in TASKS[task_id]:
            del TASKS[task_id]['preview_data']
            print(f"[CLEANUP] preview_data 정리: {task_id}")


def background_cleanup_thread():
    """백그라운드에서 주기적으로 오래된 작업 정리"""
    while True:
        try:
            time.sleep(30)  # 30초마다 체크
            current_time = time.time()

            with TASKS_LOCK:
                tasks_to_cleanup = []

                for task_id, task in list(TASKS.items()):
                    last_accessed = task.get('last_accessed', task.get('created_at', current_time))
                    status = task.get('status', '')
                    idle_seconds = int(current_time - last_accessed)

                    # 진행 중 작업: TASK_IDLE_TIMEOUT 동안 조회 없으면 취소
                    if status == 'running':
                        if current_time - last_accessed > TASK_IDLE_TIMEOUT:
                            task['cancelled'] = True
                            task['status'] = 'cancelled'
                            task['message'] = '사용자 이탈로 작업 취소됨'
                            tasks_to_cleanup.append((task_id, True))  # 파일 포함 삭제
                            print(f"[CLEANUP] 유휴 작업 취소: {task_id} (마지막 조회: {idle_seconds}초 전)")

                    # 완료된 작업이지만 백그라운드 분석/리서치가 진행 중인 경우
                    elif status == 'completed':
                        # 브라우저 닫힘 감지 (heartbeat 없음)
                        if current_time - last_accessed > TASK_IDLE_TIMEOUT:
                            # 재무분석 AI 취소
                            if task.get('analysis_status') == 'running':
                                task['analysis_status'] = 'cancelled'
                                task['analysis_message'] = '사용자 이탈로 분석 취소됨'
                                print(f"[CLEANUP] 재무분석 AI 취소: {task_id} (마지막 조회: {idle_seconds}초 전)")

                            # 기업 리서치 취소
                            if task.get('super_research_status') == 'running':
                                task['super_research_status'] = 'cancelled'
                                task['super_research_message'] = '사용자 이탈로 리서치 취소됨'
                                print(f"[CLEANUP] 기업 리서치 취소: {task_id} (마지막 조회: {idle_seconds}초 전)")

                        # 메모리 절약: AI분석/리서치 시작 안 한 경우 preview_data 조기 정리
                        # ★ completed_at이 설정되지 않았으면 (데이터 준비 중) 정리하지 않음
                        completed_at = task.get('completed_at')
                        if (completed_at and
                            current_time - completed_at > PREVIEW_DATA_TTL and
                            'preview_data' in task and
                            task.get('analysis_status') != 'running' and
                            task.get('super_research_status') != 'running' and
                            'chatbot' not in task):
                            del task['preview_data']
                            print(f"[CLEANUP] preview_data 조기 정리: {task_id} (AI 미시작, {int(current_time - completed_at)}초 경과)")

                        # last_accessed가 갱신되고 있으면 (heartbeat) 삭제하지 않음
                        if current_time - last_accessed > TASK_COMPLETED_TTL:
                            tasks_to_cleanup.append((task_id, True))  # 전체 삭제 (파일은 유지)
                            print(f"[CLEANUP] 만료된 작업 삭제: {task_id} (마지막 접근: {idle_seconds}초 전)")

                # Lock 밖에서 cleanup 수행 (deadlock 방지)
            for task_id, force in tasks_to_cleanup:
                if force:
                    with TASKS_LOCK:
                        if task_id in TASKS:
                            # 파일은 삭제하지 않음 (output 폴더에 유지)
                            del TASKS[task_id]

        except Exception as e:
            print(f"[CLEANUP] 에러: {e}")


# 백그라운드 cleanup 스레드 시작
_cleanup_thread = threading.Thread(target=background_cleanup_thread, daemon=True)
_cleanup_thread.start()


# ============================================================
# FastAPI 앱 설정
# ============================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    """앱 시작/종료 시 실행"""
    # 시작 시: output 폴더 생성
    os.makedirs("output", exist_ok=True)

    # 데이터베이스 초기화
    db.init_db()
    print("[SERVER] 데이터베이스 초기화 완료")

    yield
    # 종료 시: 모든 작업 정리
    with TASKS_LOCK:
        for task_id in list(TASKS.keys()):
            TASKS[task_id]['cancelled'] = True
    for task_id in list(TASKS.keys()):
        cleanup_task(task_id, force_delete=True)


app = FastAPI(title="DART 재무제표 추출기", lifespan=lifespan)


# ============================================================
# 인증 API 엔드포인트
# ============================================================
@app.post("/api/auth/register")
async def register(request: RegisterRequest):
    """회원가입 API"""
    # 이메일 형식 검증
    if '@' not in request.email or '.' not in request.email:
        raise HTTPException(status_code=400, detail="올바른 이메일 형식이 아닙니다")

    # 비밀번호 길이 검증
    if len(request.password) < 6:
        raise HTTPException(status_code=400, detail="비밀번호는 6자 이상이어야 합니다")

    # 사용자 생성
    user_id = db.create_user(request.email, request.password)
    if not user_id:
        raise HTTPException(status_code=400, detail="이미 등록된 이메일입니다")

    print(f"[AUTH] 회원가입 완료: {request.email}")
    return {"success": True, "message": "회원가입이 완료되었습니다"}


@app.post("/api/auth/login")
async def login(request: LoginRequest, req: Request, response: Response):
    """로그인 API"""
    user = db.authenticate_user(request.email, request.password)
    if not user:
        raise HTTPException(status_code=401, detail="이메일 또는 비밀번호가 올바르지 않습니다")

    # 세션 생성
    ip_address = req.client.host if req.client else None
    user_agent = req.headers.get('user-agent')
    token = db.create_session(user['id'], ip_address, user_agent)

    # 쿠키 설정 (브라우저 세션 - 브라우저 닫으면 만료)
    response.set_cookie(
        key="session_token",
        value=token,
        httponly=True,
        samesite="lax",
        secure=False,  # HTTPS 사용 시 True로 변경
        max_age=None   # None = 브라우저 세션 쿠키 (브라우저 닫으면 삭제)
    )

    print(f"[AUTH] 로그인: {request.email}")
    return {
        "success": True,
        "user": {
            "id": user['id'],
            "email": user['email'],
            "role": user['role'],
            "tier": user['tier']
        }
    }


@app.post("/api/auth/logout")
async def logout(response: Response, session_token: Optional[str] = Cookie(None)):
    """로그아웃 API"""
    if session_token:
        db.delete_session(session_token)

    # 쿠키 삭제
    response.delete_cookie(key="session_token")

    print("[AUTH] 로그아웃")
    return {"success": True, "message": "로그아웃되었습니다"}


class ChangePasswordRequest(BaseModel):
    """비밀번호 변경 요청"""
    current_password: str
    new_password: str


@app.post("/api/auth/change-password")
async def change_password(request: ChangePasswordRequest, user: Dict = Depends(require_auth)):
    """비밀번호 변경 API"""
    import bcrypt

    # 현재 비밀번호 확인
    user_data = db.get_user_by_id(user['user_id'])
    if not user_data:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    # 현재 비밀번호 검증
    if not bcrypt.checkpw(request.current_password.encode('utf-8'), user_data['password_hash'].encode('utf-8')):
        raise HTTPException(status_code=400, detail="현재 비밀번호가 일치하지 않습니다")

    # 새 비밀번호 해시 생성
    new_password_hash = bcrypt.hashpw(request.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    # 비밀번호 업데이트
    db.update_user_password(user['user_id'], new_password_hash)

    print(f"[AUTH] 비밀번호 변경: user_id={user['user_id']}")
    return {"success": True, "message": "비밀번호가 변경되었습니다"}


@app.get("/api/auth/me")
async def get_me(user: Optional[Dict] = Depends(get_current_user)):
    """현재 로그인한 사용자 정보 조회"""
    if not user:
        return {"logged_in": False, "user": None}

    # 사용량 통계 조회
    stats = db.get_user_stats(user['user_id'])

    return {
        "logged_in": True,
        "user": {
            "id": user['user_id'],
            "email": user['email'],
            "role": user['role'],
            "tier": user['tier'],
            "search_count": user.get('search_count', 5),
            "search_limit": user['search_limit'],
            "search_used": user['search_used'],
            "extract_limit": user['extract_limit'],
            "extract_used": user['extract_used'],
            "ai_limit": user['ai_limit'],
            "ai_used": user['ai_used'],
            "subscription_start": user.get('subscription_start'),
            "subscription_end": user.get('expires_at'),
            "expires_at": user['expires_at'],
            "created_at": user.get('created_at')
        },
        "stats": stats
    }


@app.get("/api/admin/users")
async def get_users(admin: Dict = Depends(require_admin)):
    """모든 사용자 조회 (관리자 전용)"""
    users = db.get_all_users()
    return {"users": users}


@app.get("/api/admin/analytics")
async def get_admin_analytics(admin: Dict = Depends(require_admin)):
    """관리자용 종합 분석 데이터 조회"""
    analytics = db.get_admin_analytics()
    return {"analytics": analytics}


class UpdateUserRequest(BaseModel):
    """회원 정보 수정 요청"""
    tier: Optional[str] = None
    subscription_start: Optional[str] = None  # 유료 시작일
    expires_at: Optional[str] = None  # 유료 만료일
    search_count: Optional[int] = None  # 검색횟수
    search_limit: Optional[int] = None
    extract_limit: Optional[int] = None
    ai_limit: Optional[int] = None


@app.put("/api/admin/users/{user_id}")
async def update_user(user_id: int, request: UpdateUserRequest, admin: Dict = Depends(require_admin)):
    """회원 정보 수정 (관리자 전용)"""
    user = db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    with db.get_db() as conn:
        cursor = conn.cursor()

        updates = []
        params = []

        if request.tier is not None:
            updates.append("tier = ?")
            params.append(request.tier)

            # 티어에 따른 기본 한도 및 검색횟수 설정
            from datetime import datetime, timedelta
            today = datetime.now().strftime('%Y-%m-%d')

            tier_config = {
                'free': {'search': 10, 'extract': 5, 'ai': 3, 'search_count': 5, 'duration_months': 0},
                'basic': {'search': 100, 'extract': 50, 'ai': 20, 'search_count': 300, 'duration_months': 1},
                'pro': {'search': 9999, 'extract': 9999, 'ai': 100, 'search_count': 4000, 'duration_months': 12}
            }
            if request.tier in tier_config:
                config = tier_config[request.tier]
                updates.extend(["search_limit = ?", "extract_limit = ?", "ai_limit = ?", "search_count = ?"])
                params.extend([config['search'], config['extract'], config['ai'], config['search_count']])

                # 유료 등급인 경우 시작일/만료일 자동 설정
                if request.tier in ['basic', 'pro']:
                    updates.append("subscription_start = ?")
                    params.append(today)
                    # 만료일 계산
                    expiry = datetime.now() + timedelta(days=config['duration_months'] * 30)
                    updates.append("expires_at = ?")
                    params.append(expiry.strftime('%Y-%m-%d'))
                else:
                    # Free로 변경 시 날짜 초기화
                    updates.extend(["subscription_start = ?", "expires_at = ?"])
                    params.extend([None, None])

        if request.subscription_start is not None:
            updates.append("subscription_start = ?")
            params.append(request.subscription_start if request.subscription_start else None)

        if request.expires_at is not None:
            updates.append("expires_at = ?")
            params.append(request.expires_at if request.expires_at else None)

        if request.search_limit is not None:
            updates.append("search_limit = ?")
            params.append(request.search_limit)

        if request.extract_limit is not None:
            updates.append("extract_limit = ?")
            params.append(request.extract_limit)

        if request.ai_limit is not None:
            updates.append("ai_limit = ?")
            params.append(request.ai_limit)

        if request.search_count is not None:
            updates.append("search_count = ?")
            params.append(request.search_count)

        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(user_id)
            query = f"UPDATE users SET {', '.join(updates)} WHERE id = ?"
            cursor.execute(query, params)

    print(f"[ADMIN] 회원 정보 수정: user_id={user_id}, updates={request}")
    return {"success": True, "message": "회원 정보가 수정되었습니다"}


@app.get("/api/admin/users/{user_id}/history")
async def get_user_history(user_id: int, admin: Dict = Depends(require_admin)):
    """회원 활동 이력 조회 (관리자 전용)"""
    result = db.get_user_activity_history(user_id)
    if not result:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    return result


@app.post("/api/admin/users/{user_id}/reset-usage")
async def reset_user_usage(user_id: int, admin: Dict = Depends(require_admin)):
    """회원 사용량 초기화 (관리자 전용)"""
    user = db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    db.reset_user_usage(user_id)
    print(f"[ADMIN] 사용량 초기화: user_id={user_id}")
    return {"success": True, "message": "사용량이 초기화되었습니다"}


@app.delete("/api/admin/users/{user_id}")
async def delete_user(user_id: int, admin: Dict = Depends(require_admin)):
    """회원 삭제 (관리자 전용)"""
    user = db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    # 관리자 자신은 삭제 불가
    if user_id == admin['user_id']:
        raise HTTPException(status_code=400, detail="자신의 계정은 삭제할 수 없습니다")

    with db.get_db() as conn:
        cursor = conn.cursor()
        # 관련 데이터 삭제
        cursor.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM login_history WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM search_history WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM extraction_history WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM llm_usage WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))

    print(f"[ADMIN] 회원 삭제: user_id={user_id}")
    return {"success": True, "message": "회원이 삭제되었습니다"}


# ============================================================
# API 엔드포인트
# ============================================================
@app.get("/", response_class=HTMLResponse)
async def index():
    """메인 페이지"""
    html_path = os.path.join(os.path.dirname(__file__), "index.html")
    with open(html_path, "r", encoding="utf-8") as f:
        return HTMLResponse(
            content=f.read(),
            headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"}
        )


@app.get("/admin", response_class=HTMLResponse)
async def admin_page():
    """관리자 페이지"""
    html_path = os.path.join(os.path.dirname(__file__), "admin.html")
    with open(html_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())


@app.post("/api/search")
async def search_company(request: SearchRequest):
    """기업 검색 API"""
    import traceback
    
    try:
        corp_list = get_corp_list()

        # market이 None이거나 'None' 문자열이면 전체 시장 검색 ('YKNE')
        # Y: 코스피, K: 코스닥, N: 코넥스, E: 기타(비상장)
        market_filter = request.market
        if market_filter is None or market_filter == 'None' or market_filter == '':
            market_filter = 'YKNE'  # 전체 시장 (상장+비상장 모두)

        # 회사명으로 검색 (market=None이면 상장/비상장 모두 검색)
        companies = corp_list.find_by_corp_name(
            request.company_name,
            exactly=False,
            market=market_filter  # None이면 전체 검색 (기타법인 포함)
        )

        # None 체크 및 리스트로 변환
        if companies is None:
            companies = []
        elif not isinstance(companies, list):
            companies = [companies]

        # 결과 포맷팅
        results = []
        for i, corp in enumerate(companies[:50]):  # 최대 50개
            if corp is None:
                continue
            results.append({
                "corp_code": corp.corp_code,
                "corp_name": corp.corp_name,
                "stock_code": corp.stock_code or "",
                "market": _get_market_name(corp)
            })
        
        return {"success": True, "data": results, "count": len(results)}
    
    except ValueError as e:
        print(f"[ERROR] ValueError: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"[ERROR] Exception: {e}")
        print(f"[ERROR] Traceback:\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"검색 실패: {str(e)}")


def _get_market_name(corp) -> str:
    """시장 구분명 반환"""
    if hasattr(corp, 'stock_code') and corp.stock_code:
        # 상장사 - 시장 구분 확인
        try:
            if hasattr(corp, 'corp_cls'):
                cls = corp.corp_cls
                if cls == 'Y':
                    return "코스피"
                elif cls == 'K':
                    return "코스닥"
                elif cls == 'N':
                    return "코넥스"
        except:
            pass
        return "상장"
    return "비상장"


@app.get("/api/company-info/{corp_code}")
async def get_company_info(corp_code: str, user: Optional[Dict] = Depends(get_current_user)):
    """
    기업개황정보 조회 API

    DART API를 통해 기업의 상세 개황정보를 조회합니다.
    - 회사명, 영문명, 대표자
    - 법인번호, 사업자번호
    - 주소, 홈페이지, 전화번호
    - 업종, 설립일, 결산월 등
    """
    import traceback

    try:
        print(f"[COMPANY_INFO] 조회 요청: corp_code={corp_code}")

        corp_list = get_corp_list()
        corp = corp_list.find_by_corp_code(corp_code)

        if not corp:
            raise HTTPException(status_code=404, detail="기업을 찾을 수 없습니다.")

        # 상세 정보 로드 (DART API 호출)
        corp.load()

        # 응답 데이터 구성
        info = {
            "corp_code": corp.corp_code or "",
            "corp_name": corp.corp_name or "",
            "corp_name_eng": getattr(corp, 'corp_name_eng', "") or "",
            "stock_code": corp.stock_code or "",
            "stock_name": getattr(corp, 'stock_name', "") or "",
            "ceo_nm": getattr(corp, 'ceo_nm', "") or "",
            "corp_cls": corp.corp_cls or "",
            "market_name": _get_market_name(corp),
            "jurir_no": getattr(corp, 'jurir_no', "") or "",
            "bizr_no": getattr(corp, 'bizr_no', "") or "",
            "adres": getattr(corp, 'adres', "") or "",
            "hm_url": getattr(corp, 'hm_url', "") or "",
            "ir_url": getattr(corp, 'ir_url', "") or "",
            "phn_no": getattr(corp, 'phn_no', "") or "",
            "fax_no": getattr(corp, 'fax_no', "") or "",
            "induty_code": getattr(corp, 'induty_code', "") or "",
            "est_dt": getattr(corp, 'est_dt', "") or "",
            "acc_mt": getattr(corp, 'acc_mt', "") or "",
        }

        # 업종명 추가 (KSIC 코드 → 업종명 변환)
        induty_code = info.get("induty_code", "")
        if induty_code and KSIC_CODES:
            info["induty_name"] = KSIC_CODES.get(induty_code, "")
        else:
            info["induty_name"] = ""

        # 날짜 포맷팅
        if info["est_dt"] and len(info["est_dt"]) == 8:
            info["est_dt_formatted"] = f"{info['est_dt'][:4]}-{info['est_dt'][4:6]}-{info['est_dt'][6:]}"
        else:
            info["est_dt_formatted"] = info["est_dt"]

        # 결산월 포맷팅
        if info["acc_mt"]:
            info["acc_mt_formatted"] = f"{info['acc_mt']}월"
        else:
            info["acc_mt_formatted"] = ""

        print(f"[COMPANY_INFO] 조회 완료: {info['corp_name']}")

        # 검색 기록 저장 (로그인한 사용자만)
        if user:
            try:
                db.log_search(
                    user_id=user['user_id'],
                    corp_code=corp_code,
                    corp_name=info['corp_name'],
                    market=info.get('market_name', '')
                )
                print(f"[COMPANY_INFO] 검색 기록 저장: user_id={user['user_id']}, corp={info['corp_name']}")
            except Exception as e:
                print(f"[COMPANY_INFO] 검색 기록 저장 실패: {e}")

        return {"success": True, "data": info}

    except HTTPException:
        raise
    except Exception as e:
        print(f"[COMPANY_INFO ERROR] {e}")
        print(f"[COMPANY_INFO ERROR] Traceback:\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"기업개황정보 조회 실패: {str(e)}")


@app.post("/api/extract")
async def start_extraction(
    request: ExtractRequest,
    background_tasks: BackgroundTasks,
    user: Optional[Dict] = Depends(get_current_user)
):
    """재무제표 추출 시작 API"""
    # 검색횟수 확인 및 차감 (관리자는 무제한)
    if user and user.get('role') != 'admin':
        user_search_count = user.get('search_count', 0)
        if user_search_count is not None and user_search_count <= 0:
            raise HTTPException(status_code=403, detail="검색횟수가 부족합니다. 유료 결제를 진행해 주세요.")
        # 검색횟수 차감
        if not db.use_search(user['user_id']):
            raise HTTPException(status_code=403, detail="검색횟수 차감에 실패했습니다.")

    # 작업 ID 생성
    task_id = str(uuid.uuid4())

    # 작업 상태 초기화 (사용자 정보 포함)
    current_time = time.time()
    TASKS[task_id] = {
        "status": "pending",
        "progress": 0,
        "message": "작업 대기 중...",
        "file_path": None,
        "cancelled": False,
        "corp_name": request.corp_name,
        "corp_code": request.corp_code,
        "company_info": request.company_info,  # 기업개황정보 저장
        "user_id": user['user_id'] if user else None,  # 사용자 ID 저장
        "start_year": request.start_year,
        "end_year": request.end_year,
        "created_at": current_time,  # 작업 생성 시간
        "last_accessed": current_time  # 마지막 접근 시간 (유휴 감지용)
    }

    # 백그라운드에서 추출 작업 실행
    background_tasks.add_task(
        extract_financial_data,
        task_id,
        request.corp_code,
        request.corp_name,
        request.start_year,
        request.end_year,
        request.company_info  # 기업개황정보 전달
    )

    return {"success": True, "task_id": task_id}


async def extract_financial_data(
    task_id: str,
    corp_code: str,
    corp_name: str,
    start_year: int,
    end_year: Optional[int],
    company_info: Optional[Dict[str, Any]] = None
):
    """재무제표 추출 백그라운드 작업"""
    import traceback
    
    try:
        print(f"[EXTRACT] 추출 시작: task_id={task_id}, corp_code={corp_code}, corp_name={corp_name}")
        
        task = TASKS.get(task_id)
        if not task:
            print(f"[EXTRACT] task를 찾을 수 없음: {task_id}")
            return
        
        # 취소 확인
        if task['cancelled']:
            print(f"[EXTRACT] 작업 취소됨: {task_id}")
            cleanup_task(task_id)
            return
        
        # 진행상태 업데이트: 시작
        task['status'] = 'running'
        task['progress'] = 10
        task['message'] = '재무제표 데이터 요청 중...'
        print(f"[EXTRACT] 상태 업데이트: running, 10%")
        
        # 종료 연도 설정
        if end_year is None:
            end_year = datetime.now().year

        start_date = f"{start_year}0101"
        # end_year의 사업보고서는 다음 해 초(3~4월)에 제출되므로
        # 해당 연도 사업보고서를 포함하려면 검색 범위를 +1년 확장
        search_end_year = end_year + 1
        end_date = f"{search_end_year}1231"
        print(f"[EXTRACT] 기간: {start_date} ~ {end_date} (요청: ~FY{end_year}, 검색: ~{search_end_year})")
        
        # 취소 확인
        if task['cancelled']:
            cleanup_task(task_id)
            return
        
        # 진행상태 업데이트: 재무제표 추출
        task['progress'] = 30
        task['message'] = '재무상태표 추출 중...'
        print(f"[EXTRACT] DART API 호출 시작...")
        
        # 재무제표 추출 (동기 함수를 별도 스레드에서 실행)
        loop = asyncio.get_event_loop()
        
        # 진행률 콜백 함수
        def progress_callback(progress: int, message: str):
            task['progress'] = progress
            task['message'] = message
        
        # 회사 객체에서 직접 extract_fs() 호출 방식으로 변경
        fs_data = None
        try:
            fs_data = await loop.run_in_executor(
                None,
                lambda: extract_fs_from_corp(corp_code, start_date, end_date, progress_callback)
            )
            print(f"[EXTRACT] 재무제표 추출 성공")
        except Exception as e:
            print(f"[EXTRACT] 추출 실패: {e}")
            raise RuntimeError(f"재무제표를 찾을 수 없습니다. {str(e)}")
        
        print(f"[EXTRACT] DART API 호출 완료, fs_data 타입: {type(fs_data)}")

        # 취소 확인
        if task['cancelled']:
            cleanup_task(task_id)
            return

        # 사업보고서에서 현재 대표자/주소 추출 시도
        current_address = None
        current_ceo = None
        try:
            task['progress'] = 65
            task['message'] = '정기보고서에서 최신 정보 확인 중...'
            from dart_company_info import DartCompanyInfo
            company_info_client = DartCompanyInfo()
            report_info = await loop.run_in_executor(
                None,
                lambda: company_info_client.get_company_info_from_report(corp_code)
            )
            # 정기보고서가 없는 경우, 기업개황정보를 최신으로 사용
            if report_info.get('no_report'):
                print(f"[EXTRACT] 정기보고서 없음 - 기업개황정보를 최신으로 사용")
                if company_info and company_info.get('ceo_nm'):
                    current_ceo = company_info['ceo_nm']
                    print(f"[EXTRACT] 기업개황 대표자 사용: {current_ceo}")
                    task['current_ceo'] = current_ceo
                if company_info and company_info.get('adres'):
                    current_address = company_info['adres']
                    print(f"[EXTRACT] 기업개황 주소 사용: {current_address}")
                    task['current_address'] = current_address
                task['no_report'] = True
            else:
                if report_info.get('address'):
                    current_address = report_info['address']
                    print(f"[EXTRACT] 사업보고서 현재주소: {current_address}")
                    task['current_address'] = current_address
                else:
                    # 사업보고서에 주소가 없으면 기업개황정보 주소를 최신으로 사용
                    if company_info and company_info.get('adres'):
                        current_address = company_info['adres']
                        print(f"[EXTRACT] 사업보고서 주소 없음 - 기업개황 주소 사용: {current_address}")
                        task['current_address'] = current_address
                    else:
                        print(f"[EXTRACT] 사업보고서에서 주소 추출 실패 또는 없음")
                if report_info.get('ceo'):
                    current_ceo = report_info['ceo']
                    print(f"[EXTRACT] 사업보고서 현재 대표자: {current_ceo}")
                    task['current_ceo'] = current_ceo
                else:
                    # 사업보고서에 대표자가 없으면 기업개황정보 대표자를 최신으로 사용
                    if company_info and company_info.get('ceo_nm'):
                        current_ceo = company_info['ceo_nm']
                        print(f"[EXTRACT] 사업보고서 대표자 없음 - 기업개황 대표자 사용: {current_ceo}")
                        task['current_ceo'] = current_ceo
                    else:
                        print(f"[EXTRACT] 사업보고서에서 대표자 추출 실패 또는 없음")
        except Exception as info_err:
            print(f"[EXTRACT] 사업보고서 정보 추출 오류: {info_err}")

        task['progress'] = 90
        task['message'] = '데이터 정리 중...'

        # 최신 정보로 company_info 업데이트
        if company_info:
            if current_ceo:
                company_info['ceo_nm'] = current_ceo
                print(f"[EXTRACT] 대표자 업데이트: {current_ceo}")
            if current_address:
                company_info['adres'] = current_address
                print(f"[EXTRACT] 주소 업데이트: {current_address}")

        # 엑셀은 AI 분석 완료 후 생성 - fs_data와 company_info를 task에 저장
        task['fs_data'] = fs_data
        task['company_info'] = company_info
        task['corp_name'] = corp_name

        # 취소 확인
        if task['cancelled']:
            cleanup_task(task_id)
            return

        # 완료 (엑셀 파일은 아직 없음 - AI 분석 후 생성)
        task['status'] = 'completed'
        task['progress'] = 100
        task['message'] = '추출 완료! AI 분석을 진행해주세요.'
        task['file_path'] = None  # 아직 엑셀 없음
        task['filename'] = None
        # ★ completed_at은 데이터 준비 완료 후 설정 (아래 VCM 생성 완료 후)
        # task['completed_at']은 line 1279 부근에서 설정됨

        # 사용량 로깅 (로그인한 사용자인 경우)
        if task.get('user_id'):
            try:
                db.log_extraction(
                    user_id=task['user_id'],
                    corp_code=corp_code,
                    corp_name=corp_name,
                    start_year=task.get('start_year', start_year),
                    end_year=task.get('end_year') or end_year,
                    file_path=filepath
                )
                print(f"[EXTRACT] 사용량 기록 완료: user_id={task['user_id']}")
            except Exception as log_err:
                print(f"[EXTRACT] 사용량 기록 실패: {log_err}")

        # VCM 포맷용 전체 데이터 저장 (XBRL 컬럼 정규화 적용)
        task['preview_data'] = {}
        print(f"[EXTRACT] fs_data 키: {list(fs_data.keys())}")
        
        # IS와 CIS 병합: 최대한 많은 데이터 수집
        is_df = fs_data.get('is')
        cis_df = fs_data.get('cis')
        is_empty = is_df is None or (hasattr(is_df, 'empty') and is_df.empty)
        cis_valid = cis_df is not None and hasattr(cis_df, 'empty') and not cis_df.empty
        print(f"[EXTRACT] is_empty={is_empty}, cis_valid={cis_valid}")

        # IS와 CIS 병합 로직
        if is_empty and cis_valid:
            # IS가 완전히 비어있으면 CIS로 시작
            fs_data['is'] = fs_data['cis']
            print(f"[EXTRACT] IS 데이터 없음, CIS를 IS로 사용")
        elif not is_empty and cis_valid:
            # IS와 CIS 모두 있으면 병합 (중복 제거, 계정과목 기반)
            try:
                import pandas as pd
                # 계정과목 기반 비교 (인덱스 기반은 XBRL RangeIndex에서 오작동)
                acc_col_name = '계정과목'
                if acc_col_name in is_df.columns and acc_col_name in cis_df.columns:
                    is_acc_set = set(is_df[acc_col_name].astype(str).tolist())
                    cis_unique_rows = cis_df[~cis_df[acc_col_name].astype(str).isin(is_acc_set)]
                    # OCI(기타포괄손익) 항목은 IS에 병합하지 않음
                    oci_keywords = ['기타포괄', '확정급여제도의 재측정', '보험수리', '해외사업환산', '파생상품평가', '재분류조정', '법인세효과']
                    if not cis_unique_rows.empty:
                        non_oci_mask = ~cis_unique_rows[acc_col_name].astype(str).apply(
                            lambda x: any(kw in x for kw in oci_keywords)
                        )
                        cis_is_items = cis_unique_rows[non_oci_mask]
                        if not cis_is_items.empty:
                            fs_data['is'] = pd.concat([is_df, cis_is_items], axis=0, ignore_index=True)
                            print(f"[EXTRACT] IS와 CIS 병합: CIS에서 {len(cis_is_items)}개 계정과목 추가 (OCI {len(cis_unique_rows)-len(cis_is_items)}개 제외)")
                        else:
                            print(f"[EXTRACT] IS와 CIS 병합: 추가할 IS 항목 없음 (OCI만 {len(cis_unique_rows)}개)")
                    else:
                        print(f"[EXTRACT] IS와 CIS 병합: CIS 고유 계정과목 없음")
                else:
                    # 계정과목 컬럼이 없으면 기존 인덱스 방식 fallback
                    is_accounts = set(is_df.index) if hasattr(is_df, 'index') else set()
                    cis_accounts = set(cis_df.index) if hasattr(cis_df, 'index') else set()
                    unique_cis = cis_accounts - is_accounts
                    if unique_cis:
                        cis_unique_df = cis_df.loc[list(unique_cis)]
                        fs_data['is'] = pd.concat([is_df, cis_unique_df], axis=0)
                        print(f"[EXTRACT] IS와 CIS 병합 (인덱스): CIS에서 {len(unique_cis)}개 추가")
                    else:
                        print(f"[EXTRACT] IS와 CIS 병합: 추가할 CIS 계정과목 없음")
            except Exception as e:
                print(f"[EXTRACT] IS/CIS 병합 실패: {e}, IS만 사용")
                # 병합 실패 시 IS만 사용
        
        for key in ['bs', 'is', 'cis', 'cf']:
            df = fs_data.get(key)
            print(f"[EXTRACT] fs_data[{key}]: 존재={df is not None}, empty={df.empty if df is not None else 'N/A'}, 행수={len(df) if df is not None and not df.empty else 0}")
            if df is not None and not df.empty:
                # XBRL 형식인 경우 컬럼 정규화 후 JSON 변환
                if key == 'cis':
                    print(f"[EXTRACT] CIS 원본 컬럼: {list(df.columns)[:10]}")
                df_normalized = normalize_xbrl_columns(df)
                if key == 'cis':
                    print(f"[EXTRACT] CIS 정규화 후 컬럼: {list(df_normalized.columns)}")
                task['preview_data'][key] = safe_dataframe_to_json(df_normalized)
                print(f"[EXTRACT] preview_data[{key}] 생성: {len(task['preview_data'][key])}개 행")

                # CIS 디버깅: 급여 관련 계정과목 확인
                if key == 'cis':
                    print(f"[EXTRACT] CIS 계정과목 샘플:")
                    for i, row in enumerate(task['preview_data']['cis'][:10]):
                        account = row.get('계정과목', 'N/A')
                        print(f"  [{i}] {account}")
                    print(f"[EXTRACT] CIS 급여 관련 계정과목:")
                    for row in task['preview_data']['cis']:
                        account = row.get('계정과목', '') or ''
                        if '급여' in account or '수수료' in account or '임차' in account:
                            fy2024 = row.get('FY2024', 'N/A')
                            print(f"  - {account}: FY2024={fy2024}")
        
        # 주석 테이블들도 preview_data에 추가 (프론트엔드에서 세부항목 표시용)
        # fs_data에서 주석 데이터 사용 (Excel은 AI 분석 후 생성됨)
        print(f"[EXTRACT] 주석 데이터 처리...")
        import sys
        sys.stdout.flush()
        notes = fs_data.get('notes', {'is_notes': [], 'bs_notes': [], 'cf_notes': []})
        if notes:
            print(f"[EXTRACT] fs_data 주석 로드 완료: IS={len(notes.get('is_notes', []))}개, BS={len(notes.get('bs_notes', []))}개, CF={len(notes.get('cf_notes', []))}개")
        else:
            print(f"[EXTRACT] 주석 데이터 없음")
            notes = {'is_notes': [], 'bs_notes': [], 'cf_notes': []}
        
        if notes and isinstance(notes, dict):
            # 손익계산서 주석들의 데이터를 is/cis DataFrame에 병합
            is_notes = notes.get('is_notes', [])
            print(f"[EXTRACT] is_notes 수: {len(is_notes)}")
            if is_notes:
                # fs_data의 is 또는 cis DataFrame에 병합
                target_key = 'cis' if fs_data.get('cis') is not None else 'is'
                if fs_data.get(target_key) is not None:
                    print(f"[EXTRACT] {target_key}에 주석 병합 시도...")
                    import pandas as pd
                    base_df = fs_data[target_key]

                    for note in is_notes:
                        try:
                            note_df = note['df']
                            if note_df is not None and not note_df.empty:
                                print(f"[EXTRACT] IS 주석 병합: shape={note_df.shape}")
                                # 계정과목 컬럼이 있는지 확인
                                if '계정과목' in note_df.columns:
                                    # 기존 계정과목 목록
                                    existing_accounts = set(base_df['계정과목'].values) if '계정과목' in base_df.columns else set()
                                    # 새로운 계정과목만 추가
                                    new_rows = note_df[~note_df['계정과목'].isin(existing_accounts)]
                                    if len(new_rows) > 0:
                                        base_df = pd.concat([base_df, new_rows], ignore_index=True)
                                        print(f"[EXTRACT] IS 주석 {len(new_rows)}개 행 추가")
                                        for acc in new_rows['계정과목'].head(5):
                                            print(f"[EXTRACT]   - {acc}")
                        except Exception as e:
                            print(f"[EXTRACT] IS 주석 병합 실패: {e}")

                    fs_data[target_key] = base_df
                    print(f"[EXTRACT] {target_key} 병합 완료: {len(base_df)}개 행")

                # preview_data도 업데이트
                if not task['preview_data'].get('is'):
                    task['preview_data']['is'] = []
                merged_is_data = list(task['preview_data']['is'])
                for note in is_notes:
                    try:
                        note_df = note['df']
                        if note_df is not None and not note_df.empty:
                            note_json = safe_dataframe_to_json(note_df)
                            existing_accounts = {row.get('계정과목', '') for row in merged_is_data}
                            for row in note_json:
                                account = row.get('계정과목', '')
                                if account and account not in existing_accounts:
                                    merged_is_data.append(row)
                                    existing_accounts.add(account)
                    except Exception as e:
                        pass
                task['preview_data']['is'] = merged_is_data
            
            # 재무상태표 주석들도 병합
            bs_notes = notes.get('bs_notes', [])
            if bs_notes and fs_data.get('bs') is not None:
                print(f"[EXTRACT] bs에 주석 병합 시도...")
                import pandas as pd
                base_df = fs_data['bs']

                for note in bs_notes:
                    try:
                        note_df = note['df']
                        if note_df is not None and not note_df.empty:
                            note_df_normalized = normalize_xbrl_columns(note_df)
                            if '계정과목' in note_df_normalized.columns:
                                existing_accounts = set(base_df['계정과목'].values) if '계정과목' in base_df.columns else set()
                                new_rows = note_df_normalized[~note_df_normalized['계정과목'].isin(existing_accounts)]
                                if len(new_rows) > 0:
                                    base_df = pd.concat([base_df, new_rows], ignore_index=True)
                                    print(f"[EXTRACT] BS 주석 {len(new_rows)}개 행 추가")
                    except Exception as e:
                        print(f"[EXTRACT] BS 주석 병합 실패: {e}")

                fs_data['bs'] = base_df
                print(f"[EXTRACT] BS 병합 완료: {len(base_df)}개 행")

                # preview_data도 업데이트
                if task['preview_data'].get('bs'):
                    merged_bs_data = list(task['preview_data']['bs'])
                    for note in bs_notes:
                        try:
                            note_df = note['df']
                            if note_df is not None and not note_df.empty:
                                note_df_normalized = normalize_xbrl_columns(note_df)
                                note_json = safe_dataframe_to_json(note_df_normalized)
                                existing_accounts = {row.get('계정과목', '') for row in merged_bs_data}
                                for row in note_json:
                                    account = row.get('계정과목', '')
                                    if account and account not in existing_accounts:
                                        merged_bs_data.append(row)
                                        existing_accounts.add(account)
                        except Exception as e:
                            pass
                    task['preview_data']['bs'] = merged_bs_data

        # ★ 주석 데이터를 별도로 preview_data에 저장 (프론트엔드 주석 탭용)
        notes_for_preview = []
        seen_note_signatures = set()  # 중복 제거용
        if notes and isinstance(notes, dict):
            import re as _re_notes
            for note_type, type_label in [('is_notes', '손익계산서'), ('bs_notes', '재무상태표'), ('cf_notes', '현금흐름표')]:
                for note in notes.get(note_type, []):
                    try:
                        note_df = note.get('df')
                        if note_df is not None and not note_df.empty:
                            # 컬럼 정규화
                            note_df_clean = normalize_xbrl_columns(note_df)
                            # 계정과목 + FY컬럼만 유지
                            _keep_cols = [c for c in note_df_clean.columns if c == '계정과목' or (isinstance(c, str) and _re_notes.match(r'^FY\d{4}', c))]
                            # 계정과목 컬럼 필수 + FY 컬럼 최소 1개
                            fy_cols = [c for c in _keep_cols if c != '계정과목']
                            if '계정과목' not in _keep_cols or len(fy_cols) == 0:
                                continue  # 계정과목 또는 FY 컬럼 없는 테이블은 스킵
                            note_df_clean = note_df_clean[_keep_cols]
                            # 최소 2행 이상 (의미 있는 데이터)
                            if len(note_df_clean) < 2:
                                continue
                            note_json = safe_dataframe_to_json(note_df_clean)
                            if not note_json:
                                continue
                            # 중복 제거: (제목, 첫 행 계정과목, FY컬럼 목록, 행 수) 기준
                            note_title = note.get('name', note.get('title', '주석'))
                            first_acc = note_json[0].get('계정과목', '') if note_json else ''
                            sig = (note_title, first_acc, tuple(sorted(fy_cols)), len(note_json))
                            if sig in seen_note_signatures:
                                continue
                            seen_note_signatures.add(sig)
                            is_consolidated = note.get('consolidated', False)
                            notes_for_preview.append({
                                'title': note_title,
                                'type': type_label,
                                'consolidated': is_consolidated,
                                'source': f"{'연결' if is_consolidated else '별도'}재무제표 주석 — {type_label} 관련",
                                'data': note_json
                            })
                    except Exception as e:
                        print(f"[EXTRACT] 주석 preview 변환 실패: {e}")
            if notes_for_preview:
                print(f"[EXTRACT] 주석 preview_data 생성: {len(notes_for_preview)}개 테이블 (중복/불량 제거 후)")
        task['preview_data']['notes'] = notes_for_preview

        # ★ 주석 병합 후 컬럼 정제 (주석 테이블이 본문에 concat되면서 불필요 컬럼 유입 방지)
        import re as _re2
        _meta_cols2 = {'개념ID', '계정과목', '계정과목(영문)', '분류1', '분류2', '분류3', '분류4'}
        for key in ['bs', 'is', 'cis', 'cf']:
            df = fs_data.get(key)
            if df is not None and not df.empty:
                orig_cols = len(df.columns)
                valid_cols = [c for c in df.columns if c in _meta_cols2 or (isinstance(c, str) and _re2.match(r'^FY\d{4}', c))]
                if valid_cols and len(valid_cols) < orig_cols:
                    fs_data[key] = df[valid_cols]
                    print(f"[EXTRACT] 주석 병합 후 {key} 컬럼 정제: {orig_cols}열 → {len(valid_cols)}열")
                # 중복 컬럼 제거
                if fs_data[key] is not None and fs_data[key].columns.duplicated().any():
                    fs_data[key] = fs_data[key].loc[:, ~fs_data[key].columns.duplicated()]
                    print(f"[EXTRACT] {key} 중복 컬럼 제거")

        # ★ Excel 파일이 있으면 모든 데이터를 다시 읽어서 preview_data 업데이트
        # 참고: 현재 엑셀은 AI 분석 후 생성되므로 추출 단계에서는 보통 스킵됨
        excel_filepath = task.get('file_path')
        if excel_filepath and os.path.exists(excel_filepath):
            try:
                print(f"[EXTRACT] Excel에서 preview_data 재생성...")
                xl = pd.ExcelFile(excel_filepath)

                # BS, IS, CIS, CF 시트 매핑
                sheet_mapping = {
                    'bs': '재무상태표',
                    'is': '손익계산서',
                    'cis': '포괄손익계산서',
                    'cf': '현금흐름표'
                }

                for key, sheet_name in sheet_mapping.items():
                    if sheet_name in xl.sheet_names:
                        try:
                            df = pd.read_excel(excel_filepath, sheet_name=sheet_name)
                            if df is not None and not df.empty:
                                task['preview_data'][key] = safe_dataframe_to_json(df)
                                print(f"[EXTRACT] preview_data[{key}] Excel에서 재생성: {len(task['preview_data'][key])}개 행")
                        except Exception as e:
                            print(f"[EXTRACT] {key} Excel 읽기 실패: {e}")

                # VCM 포맷 데이터도 preview_data에 추가
                if 'Frontdata' in xl.sheet_names:
                    vcm_df = pd.read_excel(excel_filepath, sheet_name='Frontdata', engine='openpyxl')
                    if vcm_df is not None and not vcm_df.empty:
                        task['preview_data']['vcm'] = safe_dataframe_to_json(vcm_df)
                        print(f"[EXTRACT] preview_data['vcm'] 생성: {len(task['preview_data']['vcm'])}개 행")

                # Financials 시트에서 데이터 읽기
                if 'Financials' in xl.sheet_names:
                    try:
                        fin_df = pd.read_excel(excel_filepath, sheet_name='Financials', header=1, engine='openpyxl')
                        if fin_df is not None and not fin_df.empty:
                            # Financials 시트는 좌우 병렬 구조 (왼쪽: BS, 오른쪽: IS)
                            # API용으로 상하 병렬 구조로 변환
                            cols = fin_df.columns.tolist()

                            # .1 접미사가 있는 컬럼이 있으면 좌우 병렬 구조
                            has_side_by_side = any('.1' in str(c) for c in cols)

                            if has_side_by_side:
                                # 왼쪽(BS) 컬럼 추출 (.1 없는 것들, Unnamed 제외)
                                bs_cols = [c for c in cols if '.1' not in str(c) and 'Unnamed' not in str(c)]
                                # 오른쪽(IS) 컬럼 추출 (.1 있는 것들)
                                is_cols = [c for c in cols if '.1' in str(c)]

                                if bs_cols and is_cols:
                                    # BS 데이터 추출
                                    bs_df = fin_df[bs_cols].copy()
                                    # 첫 번째 컬럼명을 '항목'으로
                                    bs_df = bs_df.rename(columns={bs_df.columns[0]: '항목'})
                                    # 빈 행 제거
                                    bs_df = bs_df[bs_df['항목'].notna() & (bs_df['항목'] != '')]

                                    # IS 데이터 추출
                                    is_df = fin_df[is_cols].copy()
                                    # 컬럼명에서 .1 제거
                                    is_df.columns = [str(c).replace('.1', '') for c in is_df.columns]
                                    # 첫 번째 컬럼명을 '항목'으로
                                    is_df = is_df.rename(columns={is_df.columns[0]: '항목'})
                                    # 빈 행 제거
                                    is_df = is_df[is_df['항목'].notna() & (is_df['항목'] != '')]

                                    # BS + IS 합치기 (상하 병렬)
                                    fin_df = pd.concat([bs_df, is_df], ignore_index=True)
                                    print(f"[EXTRACT] Financials 좌우→상하 변환: BS {len(bs_df)}행 + IS {len(is_df)}행 = {len(fin_df)}행")
                            else:
                                # 단일 구조면 첫 번째 컬럼명만 변경
                                if len(fin_df.columns) > 0 and fin_df.columns[0] == '(단위: 백만원)':
                                    fin_df = fin_df.rename(columns={fin_df.columns[0]: '항목'})

                            task['preview_data']['vcm_display'] = safe_dataframe_to_json(fin_df)
                            print(f"[EXTRACT] preview_data['vcm_display'] 생성: {len(task['preview_data']['vcm_display'])}개 행")
                    except Exception as fin_err:
                        print(f"[EXTRACT] Financials 파싱 실패: {fin_err}")
            except Exception as e:
                print(f"[EXTRACT] Excel preview_data 재생성 실패: {e}")
        else:
            print(f"[EXTRACT] Excel 파일 없음 - VCM 데이터 직접 생성 시도...")
            # Excel 파일이 없을 때 fs_data에서 직접 VCM 데이터 생성
            try:
                vcm_result = create_vcm_format(fs_data, None)
                if isinstance(vcm_result, tuple):
                    vcm_df, display_df = vcm_result
                else:
                    vcm_df = vcm_result
                    display_df = None

                if vcm_df is not None and not vcm_df.empty:
                    task['preview_data']['vcm'] = safe_dataframe_to_json(vcm_df)
                    print(f"[EXTRACT] preview_data['vcm'] 생성: {len(task['preview_data']['vcm'])}개 행")

                if display_df is not None and not display_df.empty:
                    task['preview_data']['vcm_display'] = safe_dataframe_to_json(display_df)
                    print(f"[EXTRACT] preview_data['vcm_display'] 생성: {len(task['preview_data']['vcm_display'])}개 행")
            except Exception as vcm_err:
                print(f"[EXTRACT] VCM 데이터 생성 실패: {vcm_err}")
                import traceback
                print(f"[EXTRACT] VCM 오류 상세:\n{traceback.format_exc()}")

        # ★ 사용자 요청 년도 범위에 맞게 FY 컬럼 필터링
        req_start = task.get('start_year')
        req_end = task.get('end_year')
        if req_start and req_end:
            def filter_fy_columns(rows, sy, ey):
                """preview_data의 행 리스트에서 요청 범위 밖의 FY 컬럼 제거"""
                if not rows:
                    return rows
                filtered = []
                for row in rows:
                    new_row = {}
                    for k, v in row.items():
                        if k.startswith('FY'):
                            try:
                                yr = int(k[2:])
                                if sy <= yr <= ey:
                                    new_row[k] = v
                            except:
                                new_row[k] = v
                        else:
                            new_row[k] = v
                    filtered.append(new_row)
                return filtered

            for key in ['vcm', 'vcm_display', 'bs', 'is', 'cis', 'cf']:
                if key in task.get('preview_data', {}):
                    data = task['preview_data'][key]
                    if isinstance(data, list) and data:
                        # bs/is/cis/cf는 FY 컬럼이 딕셔너리 안에 있음
                        task['preview_data'][key] = filter_fy_columns(data, req_start, req_end)

            print(f"[EXTRACT] FY 컬럼 필터링 적용: FY{req_start}~FY{req_end}")

        # ★ 모든 데이터 준비 완료 후 completed_at 설정 (TTL 카운트다운 시작점)
        task['completed_at'] = time.time()
        print(f"[EXTRACT] 추출 완료: {corp_name} (completed_at 설정)")

    except Exception as e:
        print(f"[EXTRACT ERROR] 추출 실패: {e}")
        print(f"[EXTRACT ERROR] Traceback:\n{traceback.format_exc()}")
        if task_id in TASKS:
            TASKS[task_id]['status'] = 'error'
            TASKS[task_id]['message'] = f'오류: {str(e)}'
            TASKS[task_id]['progress'] = 0


def extract_fs_from_corp(corp_code: str, start_date: str, end_date: str, progress_callback=None):
    """
    회사 객체에서 재무제표 추출
    
    1차: 사업보고서에서 추출 시도 (상장사)
    2차: 실패 시 감사보고서에서 추출 (비상장사)
    
    Args:
        progress_callback: 진행률 콜백 함수 (progress: int, message: str)
    """
    from dart_fss.filings import search as search_filings
    
    def update_progress(progress: int, message: str):
        if progress_callback:
            progress_callback(progress, message)
    
    print(f"[FS] 회사 검색: corp_code={corp_code}")
    update_progress(15, '회사 정보 조회 중...')
    
    corp_list = get_corp_list()
    corp = corp_list.find_by_corp_code(corp_code)
    
    if not corp:
        raise RuntimeError(f"회사를 찾을 수 없습니다: {corp_code}")
    
    print(f"[FS] 회사 찾음: {corp.corp_name}")
    print(f"[FS] 재무제표 추출 시작: {start_date} ~ {end_date}")

    # datetime import (함수 내에서 사용)
    import datetime

    fs_data = {'bs': None, 'is': None, 'cis': None, 'cf': None, 'notes': None}

    # 현재 연도 확인
    current_year = datetime.datetime.now().year
    
    # 1차: 사업보고서에서 추출 시도 (XBRL 데이터)
    xbrl_data = None
    has_current_year_annual = False  # 당해년도 사업보고서 존재 여부
    annual_report = None  # 사업보고서 객체 저장

    try:
        print(f"[FS] 사업보고서에서 추출 시도...")
        update_progress(20, '사업보고서 확인 중...')

        # 1차: 연결재무제표 시도
        fs = None
        try:
            fs = corp.extract_fs(bgn_de=start_date, end_de=end_date, separate=False)
        except Exception as consolidated_err:
            err_str = str(consolidated_err)
            if 'consolidated' in err_str.lower() or 'NotFoundConsolidated' in type(consolidated_err).__name__:
                # 연결재무제표 없음 → 별도재무제표로 재시도
                print(f"[FS] 연결재무제표 없음, 별도재무제표로 재시도...")
                update_progress(22, '별도재무제표 확인 중...')
                fs = corp.extract_fs(bgn_de=start_date, end_de=end_date, separate=True)
            else:
                raise  # 다른 오류는 그대로 전달

        if fs is not None:
            for key in ['bs', 'is', 'cis', 'cf']:
                try:
                    fs_data[key] = fs[key]
                except:
                    pass

            # dart_fss가 주석 테이블을 본문에 병합하여 불필요한 컬럼이 추가되는 경우 정제
            import re as _re
            _meta_cols = {'개념ID', '계정과목', '계정과목(영문)', '분류1', '분류2', '분류3', '분류4'}
            for key in ['bs', 'is', 'cis', 'cf']:
                df = fs_data.get(key)
                if df is not None and not df.empty:
                    orig_cols = len(df.columns)
                    # 메타 컬럼 + FY로 시작하는 연도 컬럼만 유지
                    valid_cols = [c for c in df.columns if c in _meta_cols or (isinstance(c, str) and _re.match(r'^FY\d{4}', c))]
                    if valid_cols and len(valid_cols) < orig_cols:
                        fs_data[key] = df[valid_cols]
                        print(f"[FS] {key} 컬럼 정제: {orig_cols}열 → {len(valid_cols)}열 (불필요 {orig_cols - len(valid_cols)}개 제거)")
                    # 중복 컬럼명 제거
                    if fs_data[key] is not None and fs_data[key].columns.duplicated().any():
                        fs_data[key] = fs_data[key].loc[:, ~fs_data[key].columns.duplicated()]
                        print(f"[FS] {key} 중복 컬럼 제거")

            has_data = any(fs_data[key] is not None for key in fs_data)
            # 사업보고서 추출 직후 상태 로깅
            for key in ['bs', 'is', 'cis', 'cf']:
                df = fs_data.get(key)
                print(f"[FS] 사업보고서 fs_data[{key}]: {df is not None and not df.empty if df is not None else False}")
            if has_data:
                print(f"[FS] 사업보고서에서 추출 성공")
                xbrl_data = fs_data.copy()

                # fs 객체에서 주석 테이블 추출 시도 (비용의 성격별 분류 등)
                try:
                    notes_tables = {'is_notes': [], 'bs_notes': [], 'cf_notes': []}
                    # fs.tables 또는 fs.info 등에서 추가 테이블 검색
                    tables_to_check = []
                    if hasattr(fs, 'tables'):
                        tables_to_check = fs.tables if fs.tables else []
                    elif hasattr(fs, '_tables'):
                        tables_to_check = fs._tables if fs._tables else []

                    print(f"[FS] fs 객체 테이블 수: {len(tables_to_check)}")

                    for table in tables_to_check:
                        try:
                            if hasattr(table, 'to_DataFrame'):
                                temp_df = table.to_DataFrame()
                            elif hasattr(table, 'dataframe'):
                                temp_df = table.dataframe
                            else:
                                continue

                            if temp_df is None or temp_df.empty:
                                continue

                            col_str = str(temp_df.columns[0]) if len(temp_df.columns) > 0 else ""
                            is_consolidated = '연결' in col_str or 'Consolidated' in col_str

                            # 비용/수익 관련 주석 테이블 찾기
                            expense_keywords = ['비용', '수익', '매출', '판매비', '관리비', '원가', '급여', '감가상각', '성격별']
                            if any(kw in col_str for kw in expense_keywords):
                                if not any(kw in col_str for kw in ['재무상태표', '손익계산서', '현금흐름표', '포괄손익']):
                                    notes_tables['is_notes'].append({
                                        'name': col_str[:80],
                                        'df': temp_df,
                                        'consolidated': is_consolidated
                                    })
                                    print(f"[FS] 주석 테이블 발견: {col_str[:60]}... shape={temp_df.shape}")
                        except Exception as e:
                            pass

                    if any(notes_tables[key] for key in notes_tables):
                        fs_data['notes'] = notes_tables
                        print(f"[FS] 주석 테이블 추출 완료: IS={len(notes_tables['is_notes'])}개")
                except Exception as e:
                    print(f"[FS] fs 객체 주석 추출 실패: {e}")

                # 사업보고서 객체 찾기 (HTML 주석 추출용)
                annual_list = []
                try:
                    from dart_fss.filings import search as search_filings
                    annual_filings = search_filings(
                        corp_code=corp_code,
                        bgn_de=start_date,
                        end_de=end_date,
                        pblntf_ty='A',
                        page_count=100,
                        last_reprt_at='Y'
                    )
                    if annual_filings:
                        annual_list = [f for f in annual_filings if '사업보고서' in str(f) and '반기' not in str(f) and '분기' not in str(f)]
                        if annual_list:
                            annual_report = annual_list[0]
                            print(f"[FS] 사업보고서 객체 찾음: {annual_report.report_nm} (총 {len(annual_list)}개)")
                except Exception as e:
                    print(f"[FS] 사업보고서 객체 검색 실패: {e}")

                # 사업보고서를 못 찾은 경우 감사보고서에서 주석 추출 시도 (비상장사용)
                if annual_report is None:
                    try:
                        print(f"[FS] 사업보고서 없음, 감사보고서에서 주석 추출 시도...")
                        audit_filings = search_filings(
                            corp_code=corp_code,
                            bgn_de=start_date,
                            end_de=end_date,
                            pblntf_ty=None
                        )
                        if audit_filings:
                            for af in audit_filings:
                                if '감사보고서' in str(af.report_nm) and '연결' in str(af.report_nm):
                                    annual_report = af
                                    print(f"[FS] 연결감사보고서 발견: {af.report_nm}")
                                    break
                            if annual_report is None:
                                for af in audit_filings:
                                    if '감사보고서' in str(af.report_nm):
                                        annual_report = af
                                        print(f"[FS] 감사보고서 발견: {af.report_nm}")
                                        break
                    except Exception as e:
                        print(f"[FS] 감사보고서 검색 실패: {e}")

                # 당해년도 사업보고서가 있는지 확인 (컬럼에서 FY{current_year} 존재 여부)
                for key in ['bs', 'is']:
                    df = fs_data.get(key)
                    if df is not None and not df.empty:
                        for col in df.columns:
                            col_str = str(col[0]) if isinstance(col, tuple) else str(col)
                            if str(current_year) in col_str:
                                has_current_year_annual = True
                                print(f"[FS] 당해년도({current_year}) 사업보고서 데이터 존재")
                                break
                    if has_current_year_annual:
                        break

                # 사업보고서 HTML 주석 페이지 추출 (모든 사업보고서 순회)
                # annual_list 전체를 순회하여 각 연도의 주석 테이블을 추출
                reports_to_process = []
                if annual_report and hasattr(annual_report, 'pages'):
                    reports_to_process = [annual_report]
                # annual_list의 나머지 보고서도 추가 (첫번째는 이미 포함)
                for ar in annual_list[1:]:
                    if hasattr(ar, 'pages'):
                        reports_to_process.append(ar)

                if reports_to_process:
                    print(f"[FS] 사업보고서 HTML 주석 추출: {len(reports_to_process)}개 보고서 처리")
                    import re
                    all_html_fy_cols = set()  # 이미 추출된 FY 컬럼 추적

                    for rpt_idx, rpt in enumerate(reports_to_process):
                        try:
                            year_match = re.search(r'\((\d{4})', rpt.report_nm)
                            rpt_year = int(year_match.group(1)) if year_match else current_year - rpt_idx
                            print(f"[FS] 사업보고서 [{rpt_idx}] 처리: {rpt.report_nm} → 보고서연도={rpt_year}")

                            html_notes = extract_fs_from_pages(rpt, rpt_year)
                            if html_notes and html_notes.get('notes'):
                                # 모든 주석 테이블을 추가 (VCM 스코어링에서 우선순위 결정)
                                for note_type in ['is_notes', 'bs_notes', 'cf_notes']:
                                    new_notes = html_notes['notes'].get(note_type, [])
                                    valid_notes = []
                                    for note in new_notes:
                                        note_df = note.get('df')
                                        if note_df is not None and not note_df.empty:
                                            fy_cols_in_note = [c for c in note_df.columns if str(c).startswith('FY')]
                                            if fy_cols_in_note:
                                                valid_notes.append(note)
                                                all_html_fy_cols.update(fy_cols_in_note)

                                    if valid_notes:
                                        if not fs_data.get('notes'):
                                            fs_data['notes'] = {'bs_notes': [], 'is_notes': [], 'cf_notes': [], 'other_notes': []}
                                        fs_data['notes'][note_type].extend(valid_notes)
                                        print(f"[FS] 보고서[{rpt_idx}] {note_type} 주석 {len(valid_notes)}개 추가")

                                print(f"[FS] 보고서[{rpt_idx}] HTML 주석 병합 완료")

                            # DART API 속도 제한 방지
                            import time
                            time.sleep(0.5)

                        except Exception as e:
                            print(f"[FS] 보고서[{rpt_idx}] HTML 주석 추출 실패: {e}")
                            import traceback
                            print(f"[FS] 상세: {traceback.format_exc()}")

                    print(f"[FS] 전체 HTML 주석 병합 완료: IS={len(fs_data.get('notes', {}).get('is_notes', []))}개, 추출된 FY={sorted(all_html_fy_cols)}")

                    # ★ 핵심: 주석 데이터를 실제 DataFrame에 병합 (Excel 저장 전에 반영)
                    import pandas as pd

                    # 손익계산서 주석을 cis 또는 is DataFrame에 병합
                    if fs_data.get('notes'):
                        is_notes = fs_data['notes'].get('is_notes', [])
                        if is_notes:
                            # cis 우선, 없으면 is 사용
                            target_key = 'cis' if fs_data.get('cis') is not None else 'is'
                            if fs_data.get(target_key) is not None:
                                print(f"[FS] {target_key}에 HTML 주석 병합 시도...")
                                base_df = fs_data[target_key]

                                # ★ 핵심 수정: base_df를 먼저 정규화하여 컬럼 구조 통일
                                base_df_normalized = normalize_xbrl_columns(base_df)

                                for note in is_notes:
                                    try:
                                        note_df = note['df']
                                        if note_df is not None and not note_df.empty:
                                            print(f"[FS] IS 주석 병합: shape={note_df.shape}")
                                            # 계정과목 컬럼이 있는지 확인
                                            if '계정과목' in note_df.columns:
                                                # 기존 계정과목 목록
                                                existing_accounts = set(base_df_normalized['계정과목'].values) if '계정과목' in base_df_normalized.columns else set()
                                                # 새로운 계정과목만 추가
                                                new_rows = note_df[~note_df['계정과목'].isin(existing_accounts)].copy()
                                                if len(new_rows) > 0:
                                                    # ★ 주석 데이터는 천원 단위이므로 원 단위로 변환 (*1000)
                                                    def convert_to_won(val):
                                                        """천원 단위를 원 단위로 변환"""
                                                        if pd.isna(val):
                                                            return val
                                                        try:
                                                            val_str = str(val).replace(',', '').replace('(', '').replace(')', '').strip()
                                                            if val_str == '' or val_str == '-':
                                                                return None
                                                            return float(val_str) * 1000
                                                        except:
                                                            return val

                                                    for col in new_rows.columns:
                                                        if col.startswith('FY'):
                                                            new_rows[col] = new_rows[col].apply(convert_to_won)
                                                    print(f"[FS] IS 주석 단위 변환 완료 (천원→원)")
                                                    # 정규화된 DataFrame에 병합
                                                    base_df_normalized = pd.concat([base_df_normalized, new_rows], ignore_index=True)
                                                    print(f"[FS] IS 주석 {len(new_rows)}개 행 추가")
                                                    for acc in new_rows['계정과목'].head(5):
                                                        print(f"[FS]   - {acc}")
                                    except Exception as e:
                                        print(f"[FS] IS 주석 병합 실패: {e}")

                                fs_data[target_key] = base_df_normalized
                                print(f"[FS] {target_key} 병합 완료: {len(base_df_normalized)}개 행")

                        # 재무상태표 주석을 bs DataFrame에 병합
                        bs_notes = fs_data['notes'].get('bs_notes', [])
                        if bs_notes and fs_data.get('bs') is not None:
                            print(f"[FS] bs에 HTML 주석 병합 시도...")
                            base_df = fs_data['bs']

                            # ★ 핵심 수정: base_df를 먼저 정규화하여 컬럼 구조 통일
                            base_df_normalized = normalize_xbrl_columns(base_df)

                            for note in bs_notes:
                                try:
                                    note_df = note['df']
                                    if note_df is not None and not note_df.empty:
                                        note_df_normalized = normalize_xbrl_columns(note_df)
                                        if '계정과목' in note_df_normalized.columns:
                                            existing_accounts = set(base_df_normalized['계정과목'].values) if '계정과목' in base_df_normalized.columns else set()
                                            new_rows = note_df_normalized[~note_df_normalized['계정과목'].isin(existing_accounts)]
                                            if len(new_rows) > 0:
                                                base_df_normalized = pd.concat([base_df_normalized, new_rows], ignore_index=True)
                                                print(f"[FS] BS 주석 {len(new_rows)}개 행 추가")
                                except Exception as e:
                                    print(f"[FS] BS 주석 병합 실패: {e}")

                            fs_data['bs'] = base_df_normalized
                            print(f"[FS] BS 병합 완료: {len(base_df_normalized)}개 행")

                # 당해년도 사업보고서가 있으면 바로 반환
                if has_current_year_annual:
                    update_progress(80, '데이터 처리 중...')
                    return fs_data

                # 당해년도 사업보고서가 없으면 분기/반기 데이터 추가 조회
                print(f"[FS] 당해년도({current_year}) 사업보고서 없음, 분기/반기 보고서 추가 검색...")
    except Exception as e:
        print(f"[FS] 사업보고서 추출 실패: {e}")
    
    # 2차: 당해년도 분기/반기 보고서 추가 검색 (상장사: XBRL에 병합 / 비상장사: 감사보고서+분기)
    # 비상장사 여부 확인 (stock_code가 없으면 비상장사)
    is_listed = hasattr(corp, 'stock_code') and corp.stock_code

    if is_listed:
        print(f"[FS] 상장사: 당해년도 분기/반기 보고서만 추가 검색...")
    else:
        print(f"[FS] 비상장사: 감사보고서에서 추출 시도 (주석 데이터 포함)...")
    update_progress(25, '공시보고서 검색 중...')
    
    # 연도별 데이터 저장용 딕셔너리 (키: 연도 또는 "2025 3Q" 형식)
    yearly_data = {'bs': {}, 'is': {}, 'cis': {}, 'cf': {}}

    # 연도별 주석 데이터 저장용 (각 감사보고서의 주석을 병합하기 위해)
    all_notes = {'is_notes': [], 'bs_notes': [], 'cf_notes': []}
    
    try:
        # 1. 감사보고서 검색 (비상장사는 항상, 상장사는 XBRL이 없을 때만)
        # 비상장사는 주석 데이터가 감사보고서에만 있으므로 반드시 검색해야 함
        audit_filings = []
        if not is_listed or not xbrl_data:
            filings = search_filings(
                corp_code=corp_code,
                bgn_de=start_date,
                end_de=end_date,
                pblntf_ty='F',  # F: 감사보고서
            )
            audit_filings = list(filings) if filings else []
            print(f"[FS] 검색된 감사보고서 수: {len(audit_filings)}")
        
        # 2. 당해년도 분기/반기 보고서 검색 (가장 최신 것만 사용)
        latest_periodic = None
        periodic_filings = None
        try:
            # 분기 보고서 검색 기간: end_year 다음 연도부터 오늘까지 검색
            # 예: 2024년까지 요청 시 2025년 1월 ~ 오늘까지 검색하여 2025 3Q 등을 찾음
            today_str = datetime.datetime.now().strftime('%Y%m%d')
            end_year = int(end_date[:4])
            next_year = end_year + 1

            # 다음 연도가 현재 연도를 초과하면 검색하지 않음
            if next_year > current_year:
                print(f"[FS] 분기/반기 보고서 검색 건너뜀: 다음 연도({next_year})가 현재 연도({current_year})를 초과")
            else:
                periodic_filings = search_filings(
                    corp_code=corp_code,
                    bgn_de=f"{next_year}0101",
                    end_de=today_str,  # 항상 오늘까지 검색
                    pblntf_ty='A',  # A: 정기공시 (사업보고서, 반기, 분기)
                )
            if periodic_filings:
                # 분기/반기 보고서만 필터링하고 가장 최신 것 선택
                periodic_list = [f for f in periodic_filings if '반기' in str(f) or '분기' in str(f)]
                if periodic_list:
                    # 가장 최신 보고서 선택 (리스트의 첫 번째가 최신)
                    latest_periodic = periodic_list[0]
                    print(f"[FS] 당해년도 최신 보고서: {latest_periodic}")
        except Exception as e:
            print(f"[FS] 분기/반기 보고서 검색 실패: {e}")
        
        # 감사보고서 + 당해년도 최신 분기/반기 보고서 합치기
        filings = audit_filings
        if latest_periodic:
            filings = filings + [latest_periodic]
        
        total_filings = len(filings) if filings else 0
        print(f"[FS] 처리할 총 보고서 수: {total_filings}")
        
        if filings and total_filings > 0:
            import re
            for idx, filing in enumerate(filings):
                # 진행률 계산: 30% ~ 75% 구간에서 감사보고서 처리
                progress = 30 + int((idx / total_filings) * 45)
                update_progress(progress, f'감사보고서 처리 중... ({idx + 1}/{total_filings})')
                try:
                    # Report 객체 정보 출력
                    report_name = str(filing)
                    print(f"[FS] 보고서 처리: {report_name}")
                    
                    # 보고서에서 연도/기간 추출
                    # 감사보고서: "감사보고서 (2024.12)" -> 2024
                    # 반기보고서: "반기보고서 (2025.06)" -> "2025 반기"
                    # 분기보고서: "분기보고서 (2025.09)" -> "2025 3Q"
                    report_period = None
                    period_match = re.search(r'\((\d{4})\.(\d{2})\)', report_name)
                    if period_match:
                        year = int(period_match.group(1))
                        month = int(period_match.group(2))
                        
                        if '감사보고서' in report_name or month == 12:
                            # 연간 결산
                            report_period = year
                        elif '반기' in report_name:
                            # 반기보고서: "2025 반기"
                            report_period = f"{year} 반기"
                        elif '분기' in report_name:
                            # 분기보고서: 월에 따라 1Q, 2Q, 3Q 결정
                            quarter = (month - 1) // 3 + 1
                            if quarter == 4:
                                quarter = 3  # 4Q는 연간결산이므로 3Q로
                            report_period = f"{year} {quarter}Q"
                        else:
                            # 기타: 월 표시
                            report_period = f"{year} {month}월"
                        print(f"[FS] 보고서 기간: {report_period}")
                    
                    # Report 객체에서 XBRL 데이터 추출 시도
                    xbrl = filing.xbrl
                    if xbrl is not None:
                        print(f"[FS] XBRL 데이터 발견")
                        extracted = extract_fs_from_xbrl(xbrl)
                        print(f"[FS] extracted 결과: {[(k, v is not None) for k, v in extracted.items()]}")
                        print(f"[FS] report_period={report_period}, current_year={current_year}")
                        if extracted:
                            for key in ['bs', 'is', 'cis', 'cf']:
                                if extracted.get(key) is not None:
                                    print(f"[FS] {key} 데이터 있음, 분기 체크: {report_period}, {str(current_year) in str(report_period) if report_period else False}")
                                    # 분기 데이터인 경우 yearly_data에 저장 (상장사 XBRL 병합용)
                                    if report_period and str(current_year) in str(report_period):
                                        yearly_data[key][report_period] = extracted[key]
                                        print(f"[FS] 분기 XBRL 데이터 저장: {key} -> {report_period}")
                                    elif fs_data[key] is None:
                                        fs_data[key] = extracted[key]
                                    else:
                                        fs_data[key] = pd.concat([fs_data[key], extracted[key]], ignore_index=True).drop_duplicates()
                            # 주석 테이블도 병합 (덮어쓰기가 아닌 병합)
                            if extracted.get('notes'):
                                for note_type in ['is_notes', 'bs_notes', 'cf_notes']:
                                    if extracted['notes'].get(note_type):
                                        all_notes[note_type].extend(extracted['notes'][note_type])
                                print(f"[FS] 주석 데이터 병합: IS={len(extracted['notes'].get('is_notes', []))}개 (총 {len(all_notes['is_notes'])}개)")
                    else:
                        print(f"[FS] XBRL 없음, 페이지에서 추출 시도...")
                        # 페이지에서 재무제표 테이블 추출 시도
                        extracted = extract_fs_from_pages(filing, report_period)
                        if extracted:
                            for key in ['bs', 'is', 'cis', 'cf']:
                                if extracted.get(key) is not None:
                                    df = extracted[key]
                                    if report_period and len(df.columns) >= 2:
                                        # 기간별로 데이터 저장
                                        yearly_data[key][report_period] = df
                            # 주석 테이블도 병합 (XBRL 없는 감사보고서에서도)
                            if extracted.get('notes'):
                                for note_type in ['is_notes', 'bs_notes', 'cf_notes']:
                                    if extracted['notes'].get(note_type):
                                        all_notes[note_type].extend(extracted['notes'][note_type])
                                print(f"[FS] 페이지 주석 데이터 병합: IS={len(extracted['notes'].get('is_notes', []))}개 (총 {len(all_notes['is_notes'])}개)")
                except Exception as e:
                    print(f"[FS] 감사보고서 처리 실패: {e}")
                    import traceback
                    print(f"[FS] 상세 오류: {traceback.format_exc()}")
                    continue
    except Exception as e:
        print(f"[FS] 감사보고서 검색 실패: {e}")

    # 모든 감사보고서에서 수집한 주석 데이터를 fs_data에 저장
    if any(all_notes[key] for key in all_notes):
        fs_data['notes'] = all_notes
        print(f"[FS] 전체 주석 데이터: IS={len(all_notes['is_notes'])}개, BS={len(all_notes['bs_notes'])}개, CF={len(all_notes['cf_notes'])}개")

    # 연도별 데이터를 합쳐서 정규화된 형태로 변환
    update_progress(80, '데이터 정규화 중...')
    
    # XBRL 데이터가 있는 상장사: 분기 데이터를 XBRL에 병합
    if xbrl_data and any(yearly_data[key] for key in yearly_data):
        print(f"[FS] 상장사: XBRL 데이터에 분기 데이터 병합...")
        print(f"[FS] yearly_data 키: {[(k, list(v.keys())) for k, v in yearly_data.items() if v]}")
        
        for key in ['bs', 'is', 'cis', 'cf']:
            # xbrl_data[key]가 None이고 yearly_data에 데이터가 있으면 분기 데이터를 기본으로 사용
            if yearly_data[key] and xbrl_data.get(key) is None:
                first_period = list(yearly_data[key].keys())[0]
                xbrl_data[key] = yearly_data[key][first_period]
                print(f"[FS] {key}: XBRL 데이터 없음, 분기 데이터({first_period})를 기본으로 사용")
            
            if yearly_data[key] and xbrl_data.get(key) is not None:
                xbrl_df = xbrl_data[key]
                print(f"[FS] {key}: XBRL 컬럼 = {list(xbrl_df.columns)[:5]}...")
                
                # 분기 XBRL DataFrame에서 당해년도 컬럼 추출하여 XBRL에 추가
                for period_key, quarterly_df in yearly_data[key].items():
                    # period_key 예: "2025 3Q" -> new_col_name: "FY2025 3Q"
                    new_col_name = f"FY{period_key}"
                    print(f"[FS] {key}: 분기 데이터 {period_key} 처리")
                    print(f"[FS] {key}: 분기 컬럼 전체 = {list(quarterly_df.columns)}")
                    
                    # 분기 DataFrame에서 연결재무제표의 당해년도 컬럼 찾기
                    # 우선순위: 1) 당해년도 2) 연결재무제표 3) 누적기간(01월 시작)
                    target_col = None
                    target_col_cumulative = None  # 누적기간 컬럼
                    quarterly_label_col = None
                    for col in quarterly_df.columns:
                        # 튜플 컬럼인 경우 두 번째 요소가 label_ko인지 확인
                        if isinstance(col, tuple):
                            if len(col) >= 2 and 'label_ko' in str(col[1]):
                                quarterly_label_col = col
                            # 첫 번째 요소에 당해년도가 있는지 확인
                            col_str = str(col[0])
                            if str(current_year) in col_str:
                                is_consolidated = len(col) > 1 and '연결' in str(col[1])
                                # 누적기간 확인 (20250101-20250930 형식에서 01월 시작)
                                is_cumulative = '-' in col_str and col_str[4:6] == '01'
                                # 재무상태표는 시점 데이터 (20250930)
                                is_point_data = '-' not in col_str and len(col_str) == 8
                                
                                if is_consolidated:
                                    if is_cumulative:
                                        target_col_cumulative = col
                                    elif is_point_data:
                                        # 재무상태표용 시점 데이터
                                        if target_col is None:
                                            target_col = col
                                    elif target_col is None:
                                        target_col = col
                        else:
                            col_str = str(col)
                            if 'label_ko' in col_str:
                                quarterly_label_col = col
                            if str(current_year) in col_str:
                                target_col = col
                    
                    # 누적기간 컬럼이 있으면 우선 사용 (손익계산서, 현금흐름표용)
                    if target_col_cumulative is not None:
                        target_col = target_col_cumulative
                    
                    print(f"[FS] {key}: quarterly_label_col = {quarterly_label_col}")
                    print(f"[FS] {key}: target_col = {target_col}")
                    
                    if target_col is not None and quarterly_label_col is not None:
                        print(f"[FS] 당해년도 컬럼 발견: {target_col}")
                        
                        # XBRL DataFrame의 label_ko 컬럼 찾기
                        xbrl_label_col = None
                        for col in xbrl_df.columns:
                            col_str = str(col) if not isinstance(col, tuple) else str(col)
                            if 'label_ko' in col_str:
                                xbrl_label_col = col
                                break
                        
                        if xbrl_label_col is not None:
                            # 계정과목 기준으로 분기 값 매핑
                            # 분기 데이터의 계정과목-값 딕셔너리 생성
                            quarterly_values = dict(zip(
                                quarterly_df[quarterly_label_col].astype(str),
                                quarterly_df[target_col]
                            ))
                            
                            # XBRL DataFrame에 새 컬럼 추가 (계정과목 매칭)
                            if new_col_name not in xbrl_df.columns:
                                xbrl_df[new_col_name] = xbrl_df[xbrl_label_col].apply(
                                    lambda x: quarterly_values.get(str(x), None)
                                )
                                print(f"[FS] 새 컬럼 추가 완료: {new_col_name}, 매칭된 값: {xbrl_df[new_col_name].notna().sum()}개")
                            else:
                                print(f"[FS] 컬럼 이미 존재: {new_col_name}")
                        else:
                            print(f"[FS] XBRL label_ko 컬럼을 찾을 수 없음")
                    else:
                        print(f"[FS] 분기 데이터에서 필요한 컬럼을 찾을 수 없음")
                
                xbrl_data[key] = xbrl_df
        
        # notes 데이터 유지 (분기 보고서에서 추출한 주석)
        # ★ 중요: HTML 주석으로 병합된 DataFrame도 유지해야 함!
        notes_backup = fs_data.get('notes')

        # HTML 주석으로 확장된 DataFrame이 있으면 xbrl_data 대신 사용
        import pandas as pd
        for key in ['bs', 'is', 'cis', 'cf']:
            fs_df = fs_data.get(key)
            xbrl_df = xbrl_data.get(key)

            # fs_data에 데이터가 있고, xbrl_data보다 행이 더 많으면 (HTML 주석 병합됨)
            if fs_df is not None and xbrl_df is not None:
                fs_row_count = len(fs_df) if hasattr(fs_df, '__len__') else 0
                xbrl_row_count = len(xbrl_df) if hasattr(xbrl_df, '__len__') else 0
                print(f"[FS] {key}: fs_row_count={fs_row_count}, xbrl_row_count={xbrl_row_count}")

                if fs_row_count >= xbrl_row_count:
                    print(f"[FS] {key}: HTML 주석 병합 데이터 유지 ({fs_row_count}행 > {xbrl_row_count}행)")
                    # xbrl_df를 정규화하여 fs_df와 컬럼 구조 통일
                    xbrl_df_normalized = normalize_xbrl_columns(xbrl_df)
                    fs_df_normalized = normalize_xbrl_columns(fs_df)  # fs_df도 정규화 (이미 정규화되어있으면 그대로)

                    # xbrl_df의 새 컬럼들을 fs_df에 추가
                    if '계정과목' in fs_df_normalized.columns and '계정과목' in xbrl_df_normalized.columns:
                        for col in xbrl_df_normalized.columns:
                            # 계정과목 컬럼 자체는 스킵
                            if col == '계정과목':
                                continue
                            # 이미 있는 컬럼은 스킵 (중복 방지)
                            if col in fs_df_normalized.columns:
                                continue

                            # 계정과목 기준 매칭
                            try:
                                # xbrl_df_normalized에서 해당 컬럼이 중복되어 있으면 첫 번째만 사용
                                if isinstance(xbrl_df_normalized[col], pd.DataFrame):
                                    # 중복 컬럼인 경우 첫 번째만 사용
                                    xbrl_col_data = xbrl_df_normalized[col].iloc[:, 0]
                                else:
                                    xbrl_col_data = xbrl_df_normalized[col]

                                col_values = dict(zip(xbrl_df_normalized['계정과목'], xbrl_col_data))
                                fs_df_normalized[col] = fs_df_normalized['계정과목'].map(col_values)
                                print(f"[FS] {key}: 분기 컬럼 추가 - {col}")
                            except Exception as e:
                                print(f"[FS] {key}: 컬럼 {col} 추가 실패 - {e}")
                    xbrl_data[key] = fs_df_normalized

        fs_data = xbrl_data
        if notes_backup:
            fs_data['notes'] = notes_backup
            print(f"[FS] 주석 데이터 유지: IS={len(notes_backup.get('is_notes', []))}개, BS={len(notes_backup.get('bs_notes', []))}개")
        print(f"[FS] 상장사: 병합 완료")
        return fs_data
    
    # 비상장사: 감사보고서 데이터 정규화
    for key in ['bs', 'is', 'cis', 'cf']:
        if yearly_data[key]:
            fs_data[key] = normalize_yearly_data(yearly_data[key], key)
    
    # 최소한 하나의 재무제표라도 있는지 확인
    has_data = any(fs_data[key] is not None for key in fs_data)
    if not has_data:
        raise RuntimeError("사업보고서와 감사보고서 모두에서 재무제표를 찾을 수 없습니다.")
    
    print(f"[FS] 감사보고서에서 추출 완료")
    return fs_data


def extract_fs_from_xbrl(xbrl):
    """XBRL 데이터에서 재무제표 추출 (모든 주석 테이블 포함)"""
    fs_data = {'bs': None, 'is': None, 'cis': None, 'cf': None}
    # 주석 테이블들 (재무제표별로 분류)
    notes_tables = {
        'bs_notes': [],   # 재무상태표 관련 주석
        'is_notes': [],   # 손익계산서 관련 주석
        'cf_notes': [],   # 현금흐름표 관련 주석
        'other_notes': [] # 기타 주석
    }
    
    try:
        print(f"[XBRL Extract] xbrl 타입: {type(xbrl)}")
        
        # 방법 0: xbrl.tables에서 모든 테이블을 가져와 컬럼명으로 분류 (분기보고서용)
        if hasattr(xbrl, 'tables') and xbrl.tables:
            print(f"[XBRL Extract] tables 속성 발견, 테이블 수: {len(xbrl.tables)}")
            
            for table in xbrl.tables:
                try:
                    if hasattr(table, 'to_DataFrame'):
                        temp_df = table.to_DataFrame()
                    elif hasattr(table, 'dataframe'):
                        temp_df = table.dataframe
                    else:
                        continue
                    
                    if temp_df is None or temp_df.empty:
                        continue
                    
                    # 컬럼명으로 재무제표 타입 식별
                    col_str = str(temp_df.columns[0]) if len(temp_df.columns) > 0 else ""
                    
                    # 연결재무제표 우선 (D21xxxx는 별도, D210000은 연결)
                    is_consolidated = '연결' in col_str or 'Consolidated' in col_str
                    
                    if '재무상태표' in col_str or 'financial position' in col_str.lower():
                        if fs_data['bs'] is None or is_consolidated:
                            fs_data['bs'] = temp_df
                            print(f"[XBRL Extract] bs 테이블 발견 (tables): {temp_df.shape}, 연결={is_consolidated}")
                    elif '포괄손익' in col_str or 'comprehensive income' in col_str.lower():
                        # 포괄손익계산서를 먼저 체크 (손익계산서보다 먼저)
                        if fs_data['cis'] is None or is_consolidated:
                            fs_data['cis'] = temp_df
                            print(f"[XBRL Extract] cis 테이블 발견 (tables): {temp_df.shape}, 연결={is_consolidated}")
                    elif ('손익계산서' in col_str or 'income statement' in col_str.lower()) and '포괄' not in col_str:
                        # 손익계산서 (포괄손익 제외)
                        if fs_data['is'] is None or is_consolidated:
                            fs_data['is'] = temp_df
                            print(f"[XBRL Extract] is 테이블 발견 (tables): {temp_df.shape}, 연결={is_consolidated}")
                    elif '현금흐름' in col_str or 'cash flow' in col_str.lower():
                        if fs_data['cf'] is None or is_consolidated:
                            fs_data['cf'] = temp_df
                            print(f"[XBRL Extract] cf 테이블 발견 (tables): {temp_df.shape}, 연결={is_consolidated}")
                    
                    # 주석 테이블 분류 (기본 재무제표가 아닌 모든 테이블)
                    # 손익계산서 관련 주석
                    is_note_table = False
                    if any(kw in col_str for kw in ['수익', '매출', '판매비', '관리비', '비용', '원가', '세분화', '영업']):
                        if not any(kw in col_str for kw in ['재무상태표', '손익계산서', '현금흐름표', '포괄손익']):
                            notes_tables['is_notes'].append({'name': col_str[:80], 'df': temp_df, 'consolidated': is_consolidated})
                            print(f"[XBRL Extract] 주석(IS관련) 발견: {col_str[:60]}... shape={temp_df.shape}")
                            is_note_table = True
                    # 재무상태표 관련 주석
                    if any(kw in col_str for kw in ['자산', '부채', '자본', '투자', '재고', '채권', '채무']):
                        if not any(kw in col_str for kw in ['재무상태표', '손익계산서', '현금흐름표', '포괄손익']):
                            if not is_note_table:
                                notes_tables['bs_notes'].append({'name': col_str[:80], 'df': temp_df, 'consolidated': is_consolidated})
                                print(f"[XBRL Extract] 주석(BS관련) 발견: {col_str[:60]}... shape={temp_df.shape}")
                                is_note_table = True
                    # 현금흐름표 관련 주석
                    if any(kw in col_str for kw in ['현금', '금융', '이자', '배당']):
                        if not any(kw in col_str for kw in ['재무상태표', '손익계산서', '현금흐름표', '포괄손익']):
                            if not is_note_table:
                                notes_tables['cf_notes'].append({'name': col_str[:80], 'df': temp_df, 'consolidated': is_consolidated})
                                print(f"[XBRL Extract] 주석(CF관련) 발견: {col_str[:60]}... shape={temp_df.shape}")
                            
                except Exception as e:
                    print(f"[XBRL Extract] tables 처리 중 오류: {e}")
        
        # 방법 1: get_financial_statement 메서드 사용 (tables에서 못 찾은 경우)
        if hasattr(xbrl, 'get_financial_statement'):
            for fs_type in ['bs', 'is', 'cis', 'cf']:
                if fs_data[fs_type] is not None:
                    continue  # 이미 tables에서 찾은 경우 스킵
                try:
                    result = xbrl.get_financial_statement(fs_type)
                    df = None
                    
                    # 결과가 리스트인 경우 (분기보고서 등)
                    if isinstance(result, list) and len(result) > 0:
                        print(f"[XBRL Extract] {fs_type}: 리스트 반환됨, 길이={len(result)}")
                        # 리스트에서 올바른 재무제표 찾기
                        for item in result:
                            if hasattr(item, 'to_DataFrame'):
                                temp_df = item.to_DataFrame()
                            elif hasattr(item, 'dataframe'):
                                temp_df = item.dataframe
                            elif isinstance(item, pd.DataFrame):
                                temp_df = item
                            else:
                                continue
                            
                            if temp_df is not None and not temp_df.empty:
                                # 컬럼명으로 재무제표 타입 확인
                                col_str = str(temp_df.columns[0]) if len(temp_df.columns) > 0 else ""
                                is_match = False
                                
                                if fs_type == 'bs' and ('재무상태표' in col_str or 'financial position' in col_str.lower()):
                                    is_match = True
                                elif fs_type == 'is' and ('손익계산서' in col_str or 'income statement' in col_str.lower()):
                                    is_match = True
                                elif fs_type == 'cis' and ('포괄손익' in col_str or 'comprehensive income' in col_str.lower()):
                                    is_match = True
                                elif fs_type == 'cf' and ('현금흐름' in col_str or 'cash flow' in col_str.lower()):
                                    is_match = True
                                
                                if is_match:
                                    df = temp_df
                                    print(f"[XBRL Extract] {fs_type}: 매칭된 테이블 발견")
                                    break
                        
                        if df is not None:
                            fs_data[fs_type] = df
                            print(f"[XBRL Extract] {fs_type} 추출 성공 (리스트): {df.shape}")
                    
                    elif isinstance(result, pd.DataFrame) and not result.empty:
                        fs_data[fs_type] = result
                        print(f"[XBRL Extract] {fs_type} 추출 성공: {result.shape}")
                        
                except Exception as e:
                    print(f"[XBRL Extract] {fs_type} get_financial_statement 실패: {e}")
        
        # 방법 2: 직접 속성 접근 (분기보고서용)
        if all(v is None for v in fs_data.values()):
            print(f"[XBRL Extract] get_financial_statement 실패, 직접 속성 접근 시도...")
            fs_mapping = {'bs': 'balance_sheet', 'is': 'income_statement', 'cis': 'comprehensive_income', 'cf': 'cash_flow'}
            for fs_type, attr_name in fs_mapping.items():
                if hasattr(xbrl, attr_name):
                    try:
                        df = getattr(xbrl, attr_name)
                        if df is not None and not df.empty:
                            fs_data[fs_type] = df
                            print(f"[XBRL Extract] {fs_type} ({attr_name}) 추출 성공: {df.shape}")
                    except Exception as e:
                        print(f"[XBRL Extract] {fs_type} ({attr_name}) 실패: {e}")
        
        # 방법 3: 인덱싱 접근
        if all(v is None for v in fs_data.values()):
            print(f"[XBRL Extract] 직접 속성 실패, 인덱싱 시도...")
            for fs_type in ['bs', 'is', 'cis', 'cf']:
                try:
                    df = xbrl[fs_type]
                    if df is not None and not df.empty:
                        fs_data[fs_type] = df
                        print(f"[XBRL Extract] {fs_type} 인덱싱 추출 성공: {df.shape}")
                except Exception as e:
                    print(f"[XBRL Extract] {fs_type} 인덱싱 실패: {e}")
                    
    except Exception as e:
        print(f"[XBRL Extract] 추출 실패: {e}")
        import traceback
        print(f"[XBRL Extract] 상세: {traceback.format_exc()}")
    
    # 주석 테이블들을 fs_data에 추가
    fs_data['notes'] = notes_tables
    print(f"[XBRL Extract] 주석 테이블 수: BS={len(notes_tables['bs_notes'])}, IS={len(notes_tables['is_notes'])}, CF={len(notes_tables['cf_notes'])}")
    return fs_data


def extract_fs_from_pages(report, report_year=None):
    """
    Report 객체의 페이지에서 재무제표 테이블 추출
    감사보고서 페이지 구조를 파싱하여 재무제표를 DataFrame으로 변환

    Args:
        report: Report 객체
        report_year: 보고서 연도 (예: 2024)
    """
    import re
    from bs4 import BeautifulSoup

    fs_data = {'bs': None, 'is': None, 'cis': None, 'cf': None}
    notes_tables = {'bs_notes': [], 'is_notes': [], 'cf_notes': [], 'other_notes': []}

    # 재무제표 관련 페이지 찾기
    fs_keywords = {
        'bs': ['재무상태표', '대차대조표', 'Statement of Financial Position', 'Balance Sheet'],
        'is': ['손익계산서', 'Income Statement', 'Statement of Income'],
        'cis': ['포괄손익계산서', 'Comprehensive Income'],
        'cf': ['현금흐름표', 'Cash Flow', 'Statement of Cash Flows']
    }

    try:
        pages = report.pages
        print(f"[PAGES] 총 페이지 수: {len(pages)}")

        for page in pages:
            page_title = page.title if hasattr(page, 'title') else str(page)
            # 공백 제거한 제목으로 매칭 (예: "재 무 상 태 표" -> "재무상태표")
            page_title_normalized = re.sub(r'\s+', '', page_title)
            print(f"[PAGES] 페이지: {page_title} -> {page_title_normalized}")

            # 어떤 재무제표 유형인지 확인
            for fs_type, keywords in fs_keywords.items():
                if fs_data[fs_type] is not None:
                    continue  # 이미 추출됨

                for keyword in keywords:
                    # 키워드도 공백 제거 후 비교
                    keyword_normalized = re.sub(r'\s+', '', keyword)
                    if keyword_normalized in page_title_normalized:
                        print(f"[PAGES] {fs_type} 페이지 발견: {page_title}")
                        try:
                            # 페이지 HTML 가져오기
                            html = page.html if hasattr(page, 'html') else None
                            print(f"[PAGES] HTML 길이: {len(html) if html else 0}")
                            if html:
                                df = parse_fs_table_from_html(html, fs_type)
                                if df is not None and not df.empty:
                                    fs_data[fs_type] = df
                                    print(f"[PAGES] {fs_type} 추출 성공, 행 수: {len(df)}")
                                else:
                                    print(f"[PAGES] {fs_type} 테이블 파싱 결과 없음")
                        except Exception as e:
                            print(f"[PAGES] {fs_type} 추출 실패: {e}")
                            import traceback
                            print(f"[PAGES] 상세: {traceback.format_exc()}")
                        break

            # 주석 페이지 찾기 (재무제표 주석, 주석사항, 개별 비용 주석 페이지 등)
            is_note_page = any(kw in page_title_normalized for kw in ['주석', 'Notes', '주석사항'])
            # 개별 비용 관련 주석 페이지도 포함 (예: "24. 판매비 및 관리비", "28. 비용의 성격별 분류")
            is_expense_page = any(kw in page_title_normalized for kw in [
                '판매비', '관리비', '판관비', '비용의성격별분류', '비용의성격별',
                '매출원가', '매출액과매출원가'
            ])
            if is_note_page or is_expense_page:
                page_type_label = '비용주석' if is_expense_page else '주석'
                print(f"[PAGES] {page_type_label} 페이지 발견: {page_title}")
                try:
                    html = page.html if hasattr(page, 'html') else None
                    if html:
                        # 비용의 성격별 분류 테이블 찾기
                        from io import StringIO
                        soup = BeautifulSoup(html, 'html.parser')
                        tables = pd.read_html(StringIO(str(soup)))

                        # ★ 분리형 테이블 감지 및 병합 (당기/전기가 별도 테이블로 분리된 경우)
                        # 패턴: [헤더(당기)] [데이터] [헤더(전기)] [데이터]
                        if is_expense_page and len(tables) >= 2:
                            period_data = {}  # {'FY2024': df, 'FY2023': df, ...}
                            current_period = None
                            for tidx, tbl in enumerate(tables):
                                tbl_str = tbl.to_string()
                                # 헤더 테이블 감지 (당기/전기/전전기 + 단위)
                                if len(tbl) <= 2:
                                    for val in tbl.values.flat:
                                        val_str = re.sub(r'\s', '', str(val))
                                        if '당기' in val_str or '금기' in val_str:
                                            year = report_year if report_year else 2024
                                            current_period = f'FY{year}'
                                        elif '전전기' in val_str:
                                            year = (report_year - 2) if report_year else 2022
                                            current_period = f'FY{year}'
                                        elif '전기' in val_str:
                                            year = (report_year - 1) if report_year else 2023
                                            current_period = f'FY{year}'
                                    continue

                                # 데이터 테이블 (현재 period가 설정된 상태)
                                if current_period and len(tbl) > 2 and len(tbl.columns) == 2:
                                    # 비용 키워드 확인
                                    expense_keywords_check = ['급여', '복리후생', '감가상각', '수수료', '퇴직급여', '종업원', '원재료']
                                    if any(kw in tbl_str for kw in expense_keywords_check):
                                        # 계정과목 + 금액 형태
                                        clean_df = tbl.copy()
                                        clean_df.columns = ['계정과목', current_period]
                                        # 계정과목 정제: "급여, 판관비" → "급여"
                                        clean_df['계정과목'] = clean_df['계정과목'].apply(
                                            lambda x: re.sub(r',\s*(판관비|판매비와관리비|비용의\s*성격별\s*분류)$', '', str(x)).strip()
                                        )
                                        period_data[current_period] = clean_df
                                        print(f"[PAGES] 분리형 테이블 발견: {current_period}, rows={len(clean_df)}")
                                        current_period = None  # 다음 헤더 대기

                            # 분리형 테이블 병합
                            if len(period_data) >= 1:
                                # 모든 period 데이터를 하나의 DataFrame으로 병합
                                merged_df = None
                                for fy, df in sorted(period_data.items()):
                                    if merged_df is None:
                                        merged_df = df
                                    else:
                                        merged_df = merged_df.merge(df, on='계정과목', how='outer')

                                if merged_df is not None and len(merged_df) > 2:
                                    notes_tables['is_notes'].append({
                                        'name': page_title,
                                        'df': merged_df,
                                        'consolidated': True
                                    })
                                    print(f"[PAGES] 분리형 IS 주석 테이블 추가: {merged_df.shape}, 컬럼={list(merged_df.columns)}")
                                    continue  # 이 페이지의 추가 처리 불필요

                        for idx, table in enumerate(tables):
                            # 급여, 복리후생비, 감가상각비 등이 포함된 테이블 찾기
                            table_str = table.to_string()
                            expense_keywords = ['급여', '복리후생', '감가상각', '판매수수료', '광고선전비', '지급수수료']

                            if any(kw in table_str for kw in expense_keywords):
                                # 비용 관련 주석 테이블
                                print(f"[PAGES] 비용 세부 테이블 발견: shape={table.shape}")

                                # 테이블 정제
                                if len(table) > 2:  # 최소 3행 이상
                                    # 첫 행 확인 (단위: 천원 등)
                                    cleaned_table = table.copy()

                                    # 첫 번째 행이 "단위" 정보면 제거
                                    if any('단위' in str(val) for val in cleaned_table.iloc[0].values):
                                        cleaned_table = cleaned_table.iloc[1:].reset_index(drop=True)

                                    # 컬럼명 정제: "구 분" -> "계정과목", "당기" -> "FY{year}", "전기" -> "FY{year-1}"
                                    if len(cleaned_table) > 0:
                                        new_columns = []
                                        # 먼저 컬럼명에서 패턴 확인
                                        cols_have_patterns = any('구' in str(c) or '당' in str(c) or '전' in str(c) for c in cleaned_table.columns)

                                        for i, col in enumerate(cleaned_table.columns):
                                            # 컬럼명에 패턴이 있으면 컬럼명 사용, 없으면 첫 행 값 사용
                                            if cols_have_patterns:
                                                col_str = str(col).strip()
                                            else:
                                                col_str = str(cleaned_table.iloc[0, i]).strip() if len(cleaned_table) > 0 else str(col)
                                            # 공백 제거 후 비교
                                            col_normalized = col_str.replace(' ', '')

                                            if '구분' in col_normalized:
                                                new_columns.append('계정과목')
                                            elif '당기' in col_normalized or '금기' in col_normalized:
                                                # 당기 = 보고서 연도
                                                year = report_year if report_year else 2023
                                                new_columns.append(f'FY{year}')
                                            elif '전기' in col_normalized and '전전기' not in col_normalized:
                                                # 전기 = 보고서 연도 - 1
                                                year = (report_year - 1) if report_year else 2022
                                                new_columns.append(f'FY{year}')
                                            elif '전전기' in col_normalized:
                                                # 전전기 = 보고서 연도 - 2
                                                year = (report_year - 2) if report_year else 2021
                                                new_columns.append(f'FY{year}')
                                            else:
                                                # 계정과목 컬럼이 첫 번째일 가능성 높음
                                                if i == 0:
                                                    new_columns.append('계정과목')
                                                else:
                                                    new_columns.append(col_str)

                                        # 컬럼명에 패턴이 없으면 첫 행 제거 (첫 행이 헤더였음)
                                        if not cols_have_patterns:
                                            cleaned_table = cleaned_table.iloc[1:].reset_index(drop=True)
                                        cleaned_table.columns = new_columns[:len(cleaned_table.columns)]

                                        notes_tables['is_notes'].append({
                                            'name': '비용의 성격별 분류',
                                            'df': cleaned_table,
                                            'consolidated': True
                                        })
                                        print(f"[PAGES] IS 주석 테이블 추가: {cleaned_table.shape}, 컬럼={list(cleaned_table.columns)}")
                except Exception as e:
                    print(f"[PAGES] 주석 페이지 처리 실패: {e}")
                    import traceback
                    print(f"[PAGES] 상세: {traceback.format_exc()}")

    except Exception as e:
        print(f"[PAGES] 페이지 처리 실패: {e}")

    # 주석이 있으면 fs_data에 추가
    if any(notes_tables[key] for key in notes_tables):
        fs_data['notes'] = notes_tables
        print(f"[PAGES] 주석 테이블 수: IS={len(notes_tables['is_notes'])}, BS={len(notes_tables['bs_notes'])}")

    return fs_data


def safe_dataframe_to_json(df):
    """
    DataFrame을 JSON 직렬화 가능한 형태로 안전하게 변환
    NaN, inf, -inf 값을 None으로 변환
    """
    records = df.to_dict('records')
    clean_records = []
    
    for record in records:
        clean_record = {}
        for k, v in record.items():
            # 키 이름도 문자열로 변환
            key_str = str(k) if k is not None else ''
            
            # 값 처리
            if v is None:
                clean_record[key_str] = None
            elif isinstance(v, bool):
                clean_record[key_str] = v
            elif isinstance(v, int):
                clean_record[key_str] = v
            elif isinstance(v, str):
                # 문자열이 숫자인 경우 float로 변환 시도
                v_stripped = v.strip()
                if v_stripped == '' or v_stripped == '-' or v_stripped.lower() == 'nan':
                    clean_record[key_str] = None
                else:
                    try:
                        # 괄호로 표시된 음수 처리: (1,000,000) -> -1000000
                        if v_stripped.startswith('(') and v_stripped.endswith(')'):
                            num_str = v_stripped[1:-1].replace(',', '')
                            clean_record[key_str] = -float(num_str)
                        else:
                            clean_record[key_str] = float(v_stripped.replace(',', ''))
                    except ValueError:
                        clean_record[key_str] = v  # 변환 실패 시 원본 유지
            else:
                # float, numpy 타입 등 처리
                try:
                    float_val = float(v)
                    if math.isnan(float_val) or math.isinf(float_val):
                        clean_record[key_str] = None
                    else:
                        clean_record[key_str] = float_val
                except (TypeError, ValueError):
                    # 변환 불가능한 경우 문자열로
                    clean_record[key_str] = str(v) if v is not None else None
        clean_records.append(clean_record)
    
    return clean_records


def normalize_yearly_data(yearly_data: dict, fs_type: str):
    """
    연도별로 수집된 재무제표 데이터를 정규화된 형태로 변환
    
    Args:
        yearly_data: {연도: DataFrame} 형태의 딕셔너리
        fs_type: 재무제표 유형 (bs, is, cf 등)
    
    Returns:
        계정과목을 행으로, 연도(FY20XX)를 열로 하는 DataFrame
    """
    import re
    
    if not yearly_data:
        return None
    
    # 기간을 오름차순으로 정렬 (과거 -> 현재)
    # 정수(연도)와 문자열("2025 반기", "2025 3Q")을 모두 처리
    def sort_key(period):
        if isinstance(period, int):
            return (period, 0, 0)  # 연간 결산
        else:
            parts = str(period).split()
            year = int(parts[0])
            suffix = parts[1] if len(parts) > 1 else ''
            
            if '반기' in suffix:
                return (year, 1, 6)  # 반기는 6월
            elif 'Q' in suffix:
                q = int(suffix.replace('Q', ''))
                return (year, 1, q * 3)  # 1Q=3월, 2Q=6월, 3Q=9월
            elif '월' in suffix:
                month = int(suffix.replace('월', ''))
                return (year, 1, month)
            else:
                return (year, 0, 0)
    
    sorted_periods = sorted(yearly_data.keys(), key=sort_key)
    print(f"[NORMALIZE] {fs_type}: 기간 목록 = {sorted_periods}")
    
    # 각 기간의 데이터에서 계정과목과 당기금액 추출
    all_accounts = {}  # {계정과목: {기간: 금액}}
    
    for period in sorted_periods:
        df = yearly_data[period]
        if df is None or df.empty:
            continue
        
        # 첫 번째 열을 계정과목으로 가정
        account_col = df.columns[0]
        
        # 당기 금액 열 찾기
        # HTML 테이블에서 당기 열이 두 개의 하위 열(소계/금액)로 나뉘어 있음
        # 일부 항목은 첫 번째 열에, 일부(법인세 등)는 두 번째 열(.1)에 값이 있음
        # 따라서 후보 열들을 모두 저장해두고 각 행에서 값이 있는 열 사용
        period_str = str(period)
        candidate_cols = []
        
        for col in df.columns[1:]:
            col_str = str(col).replace(' ', '')
            if '당' in col_str or period_str in col_str:
                candidate_cols.append(col)
        
        if not candidate_cols and len(df.columns) >= 2:
            candidate_cols = [df.columns[1]]  # 두 번째 열을 기본값으로
        
        if not candidate_cols:
            continue
        
        # 로그용 기본 열 (첫 번째 후보)
        value_col = candidate_cols[0]
        
        if value_col is None:
            continue
        
        print(f"[NORMALIZE] {period}: 계정과목열={account_col}, 금액열={value_col}")
        
        # 디버깅: 해당 기간 DataFrame의 첫 5행 출력
        if fs_type == 'bs':
            print(f"[NORMALIZE] {period} 샘플데이터: {df.head(3).to_dict()}")
        
        # 데이터 추출
        for idx, row in df.iterrows():
            account_raw = str(row[account_col]).strip() if pd.notna(row[account_col]) else ''
            if not account_raw or account_raw in ['nan', 'NaN', '']:
                continue
            
            # 공백, 특수문자 정리
            account_raw = account_raw.replace('\n', ' ').replace('\r', '').strip()
            
            # 계정과목 정규화: 주석번호 제거, 공백 제거
            # 예: "현금및현금성자산(주석3,19)" -> "현금및현금성자산"
            # 예: "자 산 총 계" -> "자산총계"
            account = re.sub(r'\(주석[0-9,]+\)', '', account_raw)  # 주석번호 제거
            account = re.sub(r'\s+', '', account)  # 공백 제거
            account = re.sub(r'\([0-9]+\)', '', account)  # (1), (2) 등 제거
            
            # 후보 열들 중에서 값이 있는 것을 찾아서 사용
            value = None
            for col in candidate_cols:
                val = row[col]
                if pd.notna(val) and str(val).strip() not in ['', '-', 'nan', 'NaN']:
                    value = val
                    break
            
            if account not in all_accounts:
                all_accounts[account] = {}
            
            # 이미 값이 있으면 덮어쓰지 않음 (최신 데이터 우선 유지)
            if period not in all_accounts[account] or pd.isna(all_accounts[account][period]):
                all_accounts[account][period] = value
    
    if not all_accounts:
        return None
    
    # DataFrame 생성: 계정과목을 행으로, 기간을 열로
    # 컬럼 이름: 정수는 "FY2024", 문자열은 "FY2025 8월"
    def make_col_name(period):
        if isinstance(period, int):
            return f'FY{period}'
        else:
            return f'FY{period}'  # "FY2025 8월" 형식
    
    result_data = []
    for account, period_values in all_accounts.items():
        row = {'계정과목': account}
        for period in sorted_periods:
            col_name = make_col_name(period)
            row[col_name] = period_values.get(period, None)
        result_data.append(row)
    
    result_df = pd.DataFrame(result_data)
    
    # 열 순서 정렬: 계정과목, FY20XX (오름차순)
    cols = ['계정과목'] + [make_col_name(p) for p in sorted_periods]
    result_df = result_df[cols]
    
    print(f"[NORMALIZE] {fs_type}: 결과 shape = {result_df.shape}")
    
    # 디버깅: 계정과목 목록 출력 (처음 30개)
    if fs_type == 'is':
        accounts = result_df['계정과목'].tolist()[:30]
        print(f"[NORMALIZE] {fs_type} 계정과목 샘플: {accounts}")
    
    return result_df


def parse_fs_table_from_html(html, fs_type):
    """HTML에서 재무제표 테이블을 파싱하여 DataFrame으로 변환"""
    from bs4 import BeautifulSoup
    
    try:
        soup = BeautifulSoup(str(html), 'html.parser')
        tables = soup.find_all('table')
        print(f"[PARSE] {fs_type}: 테이블 수 = {len(tables)}")
        
        # 가장 큰 테이블 찾기 (행 수 기준)
        best_df = None
        best_rows = 0
        
        for idx, table in enumerate(tables):
            try:
                dfs = pd.read_html(str(table))
                if dfs and len(dfs) > 0:
                    df = dfs[0]
                    print(f"[PARSE] 테이블 {idx}: 행={len(df)}, 열={len(df.columns)}")
                    if len(df) > best_rows:
                        best_rows = len(df)
                        best_df = df
            except Exception as e:
                print(f"[PARSE] 테이블 {idx} 파싱 오류: {e}")
                continue
        
        if best_df is not None and best_rows > 3:
            print(f"[PARSE] 최적 테이블 선택: 행={best_rows}")
            return best_df
            
    except Exception as e:
        print(f"[PARSE] HTML 파싱 실패: {e}")
    
    return None


def create_vcm_format(fs_data, excel_filepath=None):
    """VCM 전용 포맷 DataFrame 생성 - 감사보고서(FY컬럼)와 사업보고서(XBRL) 모두 지원

    Args:
        fs_data: 재무제표 데이터 딕셔너리
        excel_filepath: 엑셀 파일 경로 (주석 시트 읽기용)
    """
    import re

    # 엑셀 파일에서 이미 정규화된 데이터 읽기 (재무상태표, 손익계산서 탭에서)
    bs_df = None
    is_df = None

    if excel_filepath and os.path.exists(excel_filepath):
        try:
            bs_df = pd.read_excel(excel_filepath, sheet_name='재무상태표', engine='openpyxl')
            print(f"[VCM] 엑셀에서 재무상태표 로드: {len(bs_df)}행, 컬럼: {list(bs_df.columns)}")

            # IS(손익계산서) 우선 로드 - 매출, 매출원가, 영업이익 등 기본 손익계산서 항목 포함
            # 포괄손익계산서(CIS)는 당기순이익부터 시작하여 기본 IS 항목이 없을 수 있음
            try:
                is_df = pd.read_excel(excel_filepath, sheet_name='손익계산서', engine='openpyxl')
                print(f"[VCM] 엑셀에서 손익계산서 로드: {len(is_df)}행")

                # 손익계산서에 매출 관련 항목이 있는지 확인
                has_revenue = False
                if '계정과목' in is_df.columns:
                    accounts_str = is_df['계정과목'].astype(str).str.cat(sep=' ')
                    has_revenue = '매출' in accounts_str or '영업수익' in accounts_str

                if not has_revenue:
                    # 매출 항목이 없으면 포괄손익계산서 시도
                    try:
                        cis_df = pd.read_excel(excel_filepath, sheet_name='포괄손익계산서', engine='openpyxl')
                        cis_accounts_str = cis_df['계정과목'].astype(str).str.cat(sep=' ') if '계정과목' in cis_df.columns else ''
                        if '매출' in cis_accounts_str or '영업수익' in cis_accounts_str:
                            is_df = cis_df
                            print(f"[VCM] 포괄손익계산서에서 매출 항목 발견, 전환: {len(is_df)}행")
                    except:
                        pass
            except:
                # 손익계산서 시트가 없으면 포괄손익계산서 로드
                is_df = pd.read_excel(excel_filepath, sheet_name='포괄손익계산서', engine='openpyxl')
                print(f"[VCM] 엑셀에서 포괄손익계산서 로드 (손익계산서 없음): {len(is_df)}행")
        except Exception as e:
            print(f"[VCM] 엑셀 파일 읽기 실패, 원본 데이터 사용: {e}")
            bs_df = fs_data.get('bs')
            # IS(손익계산서) 우선, 없으면 CIS(포괄손익계산서)
            is_df = fs_data.get('is')
            if is_df is None or (isinstance(is_df, pd.DataFrame) and is_df.empty):
                is_df = fs_data.get('cis')
    else:
        # 엑셀 파일 없으면 원본 사용
        bs_df = fs_data.get('bs')
        # IS(손익계산서) 우선, 없으면 CIS(포괄손익계산서)
        is_df = fs_data.get('is')
        if is_df is None or (isinstance(is_df, pd.DataFrame) and is_df.empty):
            is_df = fs_data.get('cis')

    if bs_df is None or is_df is None:
        print(f"[VCM] 필수 데이터 누락: bs={bs_df is not None}, is={is_df is not None}")
        return None

    # Excel 파일에서 주석 시트 읽기 (HTML에서 추출한 비용 항목 포함)
    notes_dfs = []
    if excel_filepath and os.path.exists(excel_filepath):
        try:
            xl = pd.ExcelFile(excel_filepath, engine='openpyxl')
            for sheet in xl.sheet_names:
                # 손익 관련 주석 시트 찾기
                if any(kw in sheet for kw in ['손익주석', '포괄손익주석', 'IS주석', '비용']):
                    note_df = pd.read_excel(excel_filepath, sheet_name=sheet, engine='openpyxl')
                    if not note_df.empty and '계정과목' in note_df.columns:
                        notes_dfs.append(note_df)
                        print(f"[VCM] 주석 시트 로드: {sheet}, {len(note_df)}개 행")
        except Exception as e:
            print(f"[VCM] 주석 시트 로드 실패: {e}")

    # fs_data에서 XBRL 주석 데이터 추가 (비용의 성격별 분류 등)
    if fs_data and 'notes' in fs_data and fs_data['notes'] is not None:
        is_notes = fs_data['notes'].get('is_notes', [])
        for note in is_notes:
            if isinstance(note, dict) and 'df' in note:
                note_df = note['df']
                if note_df is not None and not note_df.empty:
                    # XBRL 튜플 컬럼을 FY 형식으로 정규화 (search_notes에서 FY 컬럼 매칭 필요)
                    has_tuple_cols = any(isinstance(c, tuple) for c in note_df.columns)
                    if has_tuple_cols:
                        note_df = normalize_xbrl_columns(note_df)
                    # 계정과목 컬럼이 없으면 첫 번째 컬럼을 계정과목으로 설정
                    if '계정과목' not in note_df.columns:
                        first_col = note_df.columns[0]
                        note_df = note_df.rename(columns={first_col: '계정과목'})
                    notes_dfs.append(note_df)
                    all_cols = list(note_df.columns)
                    fy_cols_found = [c for c in all_cols if str(c).startswith('FY')]
                    print(f"[VCM] XBRL 주석 추가: {note.get('name', 'unknown')}, {len(note_df)}개 행, 전체컬럼: {all_cols[:8]}, FY컬럼: {fy_cols_found}")

    # 주석 테이블 우선순위 정렬: 종합 비용 테이블 먼저 검색하도록
    # 급여, 감가상각비, 지급수수료 등 여러 키워드가 모두 있는 테이블을 우선
    def score_expense_table(df):
        """비용 테이블 우선순위 점수 계산 (높을수록 종합 테이블)"""
        if '계정과목' not in df.columns:
            return 0
        try:
            col = df['계정과목']
            if isinstance(col, pd.DataFrame):
                col = col.iloc[:, 0]
            acc_str = col.astype(str).str.cat(sep=' ')
            score = 0
            # 종합 비용 테이블 특징 키워드
            priority_keywords = ['직원급여', '퇴직급여', '감가상각비', '지급수수료', '광고선전비', '수도광열비']
            for kw in priority_keywords:
                if kw in acc_str:
                    score += 10
            # 테이블 크기도 고려 (종합 테이블은 보통 10행 이상)
            if len(df) >= 10:
                score += 5
            return score
        except:
            return 0

    def safe_get_accounts(df, limit=None):
        """'계정과목' 컬럼을 안전하게 추출 (중복 컬럼 대응)"""
        col = df['계정과목']
        if isinstance(col, pd.DataFrame):
            col = col.iloc[:, 0]  # 중복 시 첫 번째 컬럼 사용
        result = col.astype(str).tolist()
        return result[:limit] if limit else result

    # 주석 테이블 중복 컬럼 정제
    for i, df in enumerate(notes_dfs):
        if df.columns.duplicated().any():
            notes_dfs[i] = df.loc[:, ~df.columns.duplicated()]

    # 정렬 전 디버그 로깅
    print(f"[VCM] === 정렬 전 디버그 시작: {len(notes_dfs)}개 테이블 ===")
    for i, df in enumerate(notes_dfs):
        score = score_expense_table(df)
        if '계정과목' in df.columns:
            accounts = safe_get_accounts(df, 5)
            print(f"[VCM] 주석 테이블 [{i}]: {len(df)}행, score={score}, accounts={accounts}...")
        else:
            print(f"[VCM] 주석 테이블 [{i}]: {len(df)}행, score={score}, 계정과목 컬럼 없음")

    notes_dfs.sort(key=score_expense_table, reverse=True)
    if notes_dfs:
        print(f"[VCM] 주석 테이블 정렬 완료: {len(notes_dfs)}개, 최우선 테이블 행수: {len(notes_dfs[0])}")
        # 최우선 테이블 상세 정보
        if '계정과목' in notes_dfs[0].columns:
            accounts = safe_get_accounts(notes_dfs[0])
            print(f"[VCM] 최우선 테이블 계정과목: {accounts}")

    # 데이터 형식 감지: XBRL vs 감사보고서
    is_xbrl = False
    fy_col_map = {}  # {원본컬럼: 표시용문자열}
    account_col = '계정과목'  # 기본값: 감사보고서

    # 1) FY 컬럼 확인 (감사보고서 형식 또는 추가된 분기 컬럼)
    for c in is_df.columns:
        col_name = c[0] if isinstance(c, tuple) else c
        if isinstance(col_name, str) and col_name.startswith('FY'):
            fy_col_map[c] = col_name

    # 2) XBRL 형식 날짜 컬럼도 확인 (FY 컬럼과 병합)
    # XBRL 형식: ('20240101-20241231', ('연결재무제표',)) 같은 날짜 튜플 컬럼 찾기
    for c in is_df.columns:
        if isinstance(c, tuple) and len(c) >= 2:
            date_part = c[0]
            if isinstance(date_part, str) and '-' in date_part and len(date_part) == 17:
                # '20240101-20241231' 형식 - 연결재무제표만 사용
                if '연결' in str(c[1]):
                    year = date_part[:4]
                    fy_col_map[c] = f'FY{year}'
                    is_xbrl = True

    # 3) XBRL 데이터 감지 (정규화된 경우 또는 튜플 컬럼인 경우)
    # 정규화된 경우: '분류3', '개념ID', '계정과목(영문)' 컬럼
    # 튜플 컬럼인 경우: 'label_ko', 'concept_id', 'class2' 등 포함
    xbrl_indicator_cols = ['분류3', '개념ID', '계정과목(영문)']
    bs_cols_str = [str(c) for c in bs_df.columns]

    # 정규화된 XBRL 체크
    if any(col in bs_cols_str for col in xbrl_indicator_cols):
        is_xbrl = True
        print(f"[VCM] 정규화된 XBRL 데이터 감지 (분류 컬럼 존재)")
    # 튜플 컬럼 XBRL 체크 (label_ko, concept_id 등이 컬럼명에 포함)
    elif any('label_ko' in col_str or 'concept_id' in col_str or 'class2' in col_str for col_str in bs_cols_str):
        is_xbrl = True
        print(f"[VCM] 튜플 컬럼 XBRL 데이터 감지")
    
    # XBRL에서 계정과목 컬럼 찾기 (label_ko가 포함된 튜플 컬럼)
    bs_account_col = '계정과목'  # BS용 기본값
    bs_fy_col_map = {}  # BS용 연도 컬럼 매핑

    if is_xbrl:
        # IS/CIS용 계정과목 컬럼
        for c in is_df.columns:
            if isinstance(c, tuple):
                for part in c:
                    if isinstance(part, str) and 'label_ko' in part:
                        account_col = c
                        break
        # BS용 계정과목 컬럼 (별도로 찾기)
        for c in bs_df.columns:
            if isinstance(c, tuple):
                for part in c:
                    if isinstance(part, str) and 'label_ko' in part:
                        bs_account_col = c
                        print(f"[VCM] BS 계정과목 컬럼: {bs_account_col}")
                        break
        # BS용 연도 컬럼 찾기 (XBRL 형식)
        for c in bs_df.columns:
            if isinstance(c, tuple) and len(c) >= 2:
                date_part = c[0]
                # BS는 특정 시점 (20241231 형식) 또는 기간 (20240101-20241231 형식) 모두 가능
                if isinstance(date_part, str):
                    if '-' in date_part and len(date_part) == 17:
                        # '20240101-20241231' 형식 - 연결재무제표만 사용
                        if '연결' in str(c[1]):
                            year = date_part.split('-')[1][:4]  # 종료 연도 사용
                            bs_fy_col_map[c] = f'FY{year}'
                    elif len(date_part) == 8 and date_part.isdigit():
                        # '20241231' 형식 (BS 특정 시점)
                        if '연결' in str(c[1]):
                            year = date_part[:4]
                            bs_fy_col_map[c] = f'FY{year}'
        print(f"[VCM] BS 연도 컬럼: {list(bs_fy_col_map.values())}")
    else:
        bs_account_col = account_col  # 감사보고서 형식이면 동일
        bs_fy_col_map = fy_col_map  # 감사보고서 형식이면 동일

    if not fy_col_map:
        print(f"[VCM] 연도 컬럼을 찾을 수 없음")
        return None
    
    print(f"[VCM] 데이터 형식: {'XBRL' if is_xbrl else '감사보고서'}, 연도: {list(fy_col_map.values())}, 계정과목컬럼: {account_col}")
    
    # 정렬된 원본 컬럼 목록
    fy_cols = sorted(fy_col_map.keys(), key=lambda c: fy_col_map[c])

    # ★ 2컬럼 주석 테이블에 FY 연도 컬럼 부여
    # XBRL 주석 테이블은 ['계정과목', <값>] 형태로 FY 컬럼이 없음
    # IS의 판관비 값과 비교하여 어떤 연도의 테이블인지 매핑
    def _assign_fy_to_notes():
        """2컬럼 주석 테이블에 FY 연도 할당"""
        # IS에서 판관비 값 추출 (연도별)
        sga_by_year = {}  # {year_str: sga_value}
        for fy_col in fy_cols:
            yr = fy_col_map[fy_col]
            for _, row in is_df.iterrows():
                acc = str(row.get(account_col, ''))
                acc_norm = re.sub(r'\s', '', acc)
                if '판매비와관리비' in acc_norm or '판매비및관리비' in acc_norm or '판관비' in acc_norm:
                    val = row.get(fy_col)
                    if isinstance(val, pd.Series):
                        val = val.iloc[0]
                    if pd.notna(val):
                        try:
                            sga_by_year[yr] = abs(float(str(val).replace(',', '')))
                        except:
                            pass
                    break

        if not sga_by_year:
            return

        # 판관비 합계가 있는 주석 테이블 찾기 → FY 연도 매핑
        # 2컬럼 테이블뿐만 아니라 3컬럼(계정과목+당기+전기) 테이블도 처리
        for i, note_df in enumerate(notes_dfs):
            if '계정과목' not in note_df.columns:
                continue

            non_acc_cols = [c for c in note_df.columns if c != '계정과목']
            if not non_acc_cols:
                continue

            # 이미 FY 컬럼만 있으면 스킵
            fy_data_cols = [c for c in non_acc_cols if str(c).startswith('FY')]
            non_fy_cols = [c for c in non_acc_cols if not str(c).startswith('FY')]

            if not non_fy_cols:
                continue  # 모든 데이터 컬럼이 이미 FY → 스킵

            # 2컬럼(계정과목+1데이터) 또는 3컬럼 이하만 처리
            if len(non_fy_cols) > 2:
                continue

            # 각 non-FY 데이터 컬럼에 대해 FY 매핑 시도
            for data_col in non_fy_cols:
                # 판관비 합계 행 찾기
                sga_total = None
                for _, row in note_df.iterrows():
                    acc = re.sub(r'\s', '', str(row.get('계정과목', '')))
                    if '판매비와관리비합계' in acc or '판관비합계' in acc or acc == '판매비와관리비':
                        try:
                            v = str(row[data_col]).replace(',', '').replace('(', '-').replace(')', '')
                            sga_total = abs(float(v)) * 1000  # 주석은 천원 → 원
                        except:
                            pass
                        break

                if sga_total is None:
                    # 비용의 성격별 분류: '성격별비용합계' 행 찾기
                    for _, row in note_df.iterrows():
                        acc = re.sub(r'\s', '', str(row.get('계정과목', '')))
                        if '성격별비용합계' in acc or '비용합계' in acc:
                            try:
                                v = str(row[data_col]).replace(',', '').replace('(', '-').replace(')', '')
                                sga_total = abs(float(v)) * 1000
                            except:
                                pass
                            break

                if sga_total is None:
                    print(f"[VCM] 주석 테이블 [{i}] col={data_col} FY매핑 스킵: 판관비합계 없음, rows={len(note_df)}")
                    continue

                if not sga_by_year:
                    print(f"[VCM] 주석 테이블 [{i}] col={data_col} FY매핑 스킵: 남은 IS연도 없음")
                    break

                # IS 판관비 값과 매칭 (20% 허용오차)
                best_year = None
                best_diff = float('inf')
                for yr, is_val in sga_by_year.items():
                    diff = abs(sga_total - is_val) / max(is_val, 1)
                    if diff < 0.2 and diff < best_diff:
                        best_diff = diff
                        best_year = yr

                if best_year:
                    notes_dfs[i] = note_df.rename(columns={data_col: best_year})
                    note_df = notes_dfs[i]  # 다음 컬럼 처리를 위해 갱신
                    print(f"[VCM] 주석 테이블 [{i}] FY매핑: {data_col} → {best_year} (판관비 {sga_total/1e6:.0f}M vs IS {sga_by_year[best_year]/1e6:.0f}M, diff={best_diff:.1%})")
                    # 매핑된 연도는 제거 (다른 테이블에 중복 매핑 방지)
                    del sga_by_year[best_year]
                else:
                    print(f"[VCM] 주석 테이블 [{i}] FY매핑 실패: col={data_col}, note_sga={sga_total/1e6:.0f}M, 남은 IS: {{{', '.join(f'{yr}={v/1e6:.0f}M' for yr, v in sga_by_year.items())}}}")

    _assign_fy_to_notes()

    # 계정과목 정규화 함수
    def normalize(s):
        if not s: return ''
        s = re.sub(r'\s', '', str(s))
        # 로마숫자 접두사 제거 (전각: Ⅰ,Ⅱ,Ⅲ / 반각: I,II,III,IV,V)
        s = re.sub(r'^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.', '', s)
        s = re.sub(r'^[IVX]+\.', '', s)  # 반각 로마숫자 제거
        # 주석 참조 제거: (주10), (주1,2), (주석1,2) 등
        s = re.sub(r'\(주[석\d,\s]*\)', '', s)
        s = re.sub(r'\(손실\)', '', s)
        return s

    # 주석 참조 제거 함수 (매칭용) - 값은 유지하고 계정명만 정규화
    def strip_note_ref(s):
        """계정과목명에서 주석 참조 제거 (예: '유형자산(주2,4,8)' → '유형자산')"""
        if not s: return ''
        s = re.sub(r'\(주[\d,\s]+\)', '', str(s))
        s = re.sub(r'\(주석[\d,\s]+\)', '', s)
        return s.strip()

    # ========== 섹션 기반 재무상태표 파싱 ==========
    def parse_bs_sections(df, year_col, acc_col):
        """재무상태표를 섹션별로 파싱하여 모든 항목 추출

        Args:
            df: 재무상태표 DataFrame
            year_col: 연도 컬럼명
            acc_col: 계정과목 컬럼명 (BS용)

        Returns:
            dict: {
                '유동자산': [{'name': '현금및현금성자산', 'value': 1000000}, ...],
                '비유동자산': [...],
                '유동부채': [...],
                '비유동부채': [...],
                '자본': [...],
                '총계': {'자산총계': val, '부채총계': val, '자본총계': val}
            }
        """
        sections = {
            '유동자산': [],
            '비유동자산': [],
            '유동부채': [],
            '비유동부채': [],
            '자본': [],
            '총계': {}
        }

        # 섹션 마커 패턴 (로마숫자/숫자 prefix 허용)
        section_patterns = {
            '유동자산': r'^(Ⅰ|I|1)?\.?\s*유동자산$',
            '비유동자산': r'^(Ⅱ|II|2)?\.?\s*비유동자산$',
            '유동부채': r'^(Ⅰ|I|1)?\.?\s*유동부채$',
            '비유동부채': r'^(Ⅱ|II|2)?\.?\s*비유동부채$',
            '자본': r'^(Ⅰ|I|1)?\.?\s*자본$',  # '자본금' 제외 (정확히 '자본'만 매칭)
        }

        # 종료 마커 (다음 섹션 시작 또는 총계)
        end_markers = {
            '유동자산': ['비유동자산', '자산총계'],
            '비유동자산': ['자산총계', '부채', '유동부채'],
            '유동부채': ['비유동부채', '부채총계'],
            '비유동부채': ['부채총계', '자본'],
            '자본': ['자본총계', '부채와자본총계', '부채및자본총계', '자본과부채총계'],
        }

        # 총계 마커
        total_markers = ['자산총계', '부채총계', '자본총계', '부채와자본총계', '부채및자본총계', '자본과부채총계']

        current_section = None

        for idx, row in df.iterrows():
            raw_acc = str(row.get(acc_col, '')).strip()
            if not raw_acc or raw_acc == 'nan':
                continue

            # 정규화된 계정명 (주석 참조 제거)
            acc_normalized = strip_note_ref(raw_acc)
            acc_clean = normalize(acc_normalized)

            # 값 추출
            val = row.get(year_col)
            val_num = None
            if pd.notna(val):
                try:
                    val_num = float(str(val).replace(',', ''))
                except:
                    pass

            # 섹션 시작 감지
            section_found = False
            for section_name, pattern in section_patterns.items():
                if re.match(pattern, acc_clean, re.IGNORECASE):
                    current_section = section_name
                    section_found = True
                    print(f"[섹션파싱] {section_name} 시작: {raw_acc}")
                    break

            if section_found:
                continue

            # 총계 항목 처리
            is_total = False
            for marker in total_markers:
                if marker in acc_clean or acc_clean == normalize(marker):
                    sections['총계'][acc_clean] = val_num
                    is_total = True
                    # 자산총계, 부채총계 도달 시 해당 섹션 종료
                    if '자산총계' in acc_clean:
                        current_section = None
                    elif '부채총계' in acc_clean:
                        current_section = None
                    break

            if is_total:
                continue

            # 현재 섹션에 항목 추가
            if current_section:
                # 종료 마커 체크
                should_end = False
                for end_marker in end_markers.get(current_section, []):
                    if normalize(end_marker) in acc_clean:
                        should_end = True
                        break

                if should_end:
                    current_section = None
                elif val_num is not None and val_num != 0:
                    # 하위 섹션 헤더가 아닌 실제 항목만 추가
                    # (자산, 부채, 투자자산 같은 중간 헤더 제외)
                    if not re.match(r'^(자산|부채|자본|투자자산|당좌자산)$', acc_clean):
                        sections[current_section].append({
                            'name': acc_normalized,  # 원본 계정명 유지 (주석 참조만 제거)
                            'value': val_num
                        })

        return sections

    def parse_bs_sections_xbrl(df, year_col, acc_col):
        """XBRL 형식 재무상태표를 '분류3' 컬럼 기반으로 파싱

        XBRL 데이터는 '분류3' 컬럼에 섹션 정보가 있음:
        - '유동자산', '비유동자산', '유동부채', '비유동부채', '자본'
        """
        sections = {
            '유동자산': [],
            '비유동자산': [],
            '유동부채': [],
            '비유동부채': [],
            '자본': [],
            '총계': {}
        }

        # 분류3 컬럼이 없으면 빈 결과 반환
        if '분류3' not in df.columns:
            print(f"[VCM XBRL] 분류3 컬럼 없음, 빈 섹션 반환")
            return sections

        # 섹션 매핑
        section_mapping = {
            '유동자산': '유동자산',
            '비유동자산': '비유동자산',
            '유동부채': '유동부채',
            '비유동부채': '비유동부채',
            '자본': '자본',
        }

        # 총계 마커
        total_markers = ['자산총계', '부채총계', '자본총계', '부채와자본총계', '부채및자본총계']

        for idx, row in df.iterrows():
            raw_acc = str(row.get(acc_col, '')).strip()
            if not raw_acc or raw_acc == 'nan':
                continue

            section_name = str(row.get('분류3', '')).strip()
            acc_clean = normalize(raw_acc)

            # 값 추출
            val = row.get(year_col)
            val_num = None
            if pd.notna(val):
                try:
                    val_num = float(str(val).replace(',', ''))
                except:
                    pass

            # 총계 항목 처리
            is_total = False
            for marker in total_markers:
                if marker in acc_clean or acc_clean == normalize(marker):
                    sections['총계'][acc_clean] = val_num
                    is_total = True
                    break

            if is_total:
                continue

            # 섹션 헤더 자체는 스킵 (예: 계정='유동자산', 분류3='유동자산')
            if raw_acc == section_name:
                continue

            # 섹션에 항목 추가
            if section_name in section_mapping and val_num is not None and val_num != 0:
                target_section = section_mapping[section_name]
                # 중간 헤더 제외
                if not re.match(r'^(자산|부채|자본|투자자산|당좌자산)$', acc_clean):
                    sections[target_section].append({
                        'name': strip_note_ref(raw_acc),
                        'value': val_num
                    })

        return sections

    # 섹션 데이터에서 값 찾기 (parse_bs_sections 결과 사용)
    def find_in_section(section_items, keywords, excludes=[]):
        """섹션 내 항목에서 키워드로 값 찾기"""
        for item in section_items:
            name = normalize(item['name'])
            excluded = any(normalize(ex) in name for ex in excludes)
            if excluded:
                continue
            for kw in keywords:
                if normalize(kw) in name:
                    return item['value']
        return None

    def sum_in_section(section_items, keywords_list):
        """섹션 내 여러 항목의 합계"""
        total = 0
        for item in section_items:
            name = normalize(item['name'])
            for keywords in keywords_list:
                if isinstance(keywords, str):
                    keywords = [keywords]
                for kw in keywords:
                    if normalize(kw) in name:
                        total += item['value'] or 0
                        break
        return total

    def get_unmatched_items(section_items, matched_keywords_list):
        """매칭되지 않은 항목들 반환 (기타 항목용)"""
        unmatched = []
        for item in section_items:
            name = normalize(item['name'])
            is_matched = False
            for keywords in matched_keywords_list:
                if isinstance(keywords, str):
                    keywords = [keywords]
                for kw in keywords:
                    if normalize(kw) in name:
                        is_matched = True
                        break
                if is_matched:
                    break
            if not is_matched:
                unmatched.append(item)
        return unmatched

    # 값 찾기 함수 (is_df + notes_dfs에서 검색) - 원 단위 그대로 반환
    # 판관비 세부항목은 주석 테이블 먼저 검색 (is_df에 리스 관련 중복값이 있을 수 있음)
    expense_keywords_for_notes_priority = ['급여', '퇴직급여', '감가상각비', '지급수수료', '광고선전비',
                                           '수도광열비', '임차료', '복리후생비', '무형자산상각비', '대손상각비']

    def find_val(df, keywords, year, excludes=[], notes_first=False):
        """
        값을 찾는 함수
        notes_first=True면 주석 테이블을 먼저 검색
        """
        # 판관비 세부항목 키워드가 포함되어 있으면 주석 먼저 검색
        search_notes_first = notes_first or any(
            any(normalize(exp_kw) in normalize(kw) for exp_kw in expense_keywords_for_notes_priority)
            for kw in keywords
        )

        def search_is_df():
            """주 DataFrame(is_df)에서 검색"""
            for _, row in df.iterrows():
                acc = normalize(str(row.get(account_col, '')))
                excluded = any(normalize(ex) in acc for ex in excludes)
                if excluded: continue
                for kw in keywords:
                    if normalize(kw) in acc:
                        val = row.get(year)
                        # Series인 경우 첫 번째 값 사용 (중복 컬럼 대응)
                        if isinstance(val, pd.Series):
                            val = val.iloc[0] if len(val) > 0 else None
                        if pd.notna(val):
                            try:
                                val_str = str(val).replace(',', '').strip()
                                # 괄호로 표시된 음수 처리: (1234) → -1234
                                if val_str.startswith('(') and val_str.endswith(')'):
                                    val_str = '-' + val_str[1:-1]
                                return float(val_str)
                            except:
                                pass
            return None

        def search_notes():
            """주석 시트에서 검색 (주석 테이블은 천원 단위이므로 1000배 변환)"""
            for note_idx, note_df in enumerate(notes_dfs):
                for _, row in note_df.iterrows():
                    acc = normalize(str(row.get('계정과목', '')))
                    excluded = any(normalize(ex) in acc for ex in excludes)
                    if excluded: continue
                    for kw in keywords:
                        if normalize(kw) in acc:
                            year_str = fy_col_map.get(year)
                            if year_str is None:
                                if isinstance(year, str) and year.startswith('FY'):
                                    year_str = year
                            if year_str and year_str in note_df.columns:
                                val = row.get(year_str)
                                # Series인 경우 첫 번째 값 사용 (중복 컬럼 대응)
                                if isinstance(val, pd.Series):
                                    val = val.iloc[0] if len(val) > 0 else None
                                if pd.notna(val):
                                    try:
                                        val_str = str(val).replace(',', '').strip()
                                        # 괄호로 표시된 음수 처리: (1234) → -1234
                                        if val_str.startswith('(') and val_str.endswith(')'):
                                            val_str = '-' + val_str[1:-1]
                                        result = float(val_str)
                                        # 주석 테이블은 천원 단위이므로 원 단위로 변환
                                        result = result * 1000
                                        return result
                                    except:
                                        pass
            return None

        # 검색 순서 결정
        if search_notes_first:
            result = search_notes()
            if result is not None:
                return result
            return search_is_df()
        else:
            result = search_is_df()
            if result is not None:
                return result
            return search_notes()

    # VCM 항목 정의
    vcm_items = []
    
    # ========== 매출/매출원가 하위 항목 동적 추출 ==========
    # 손익계산서에서 매출 및 매출원가 하위 항목을 자동으로 추출
    revenue_sub_items = []  # 매출 하위 항목 [(항목명, 계정과목명), ...]
    cogs_sub_items = []     # 매출원가 하위 항목 [(항목명, 계정과목명), ...]

    in_revenue_section = False
    in_cogs_section = False

    for idx, row in is_df.iterrows():
        acc = normalize(str(row.get(account_col, '')))

        # 매출액 섹션 시작 (헤더만 인식: '매출액'으로 정확히 일치)
        if acc == '매출액':
            in_revenue_section = True
            in_cogs_section = False
            continue

        # 매출원가 섹션 시작 (헤더만 인식)
        if acc == '매출원가':
            in_revenue_section = False
            in_cogs_section = True
            continue

        # 매출총이익 섹션 시작 (매출원가 섹션 종료)
        if '매출총이익' in acc or '총이익' in acc:
            in_revenue_section = False
            in_cogs_section = False
            continue

        # 매출 하위 항목 수집
        if in_revenue_section and acc:
            # 원본 계정과목명 가져오기 (정규화되지 않은)
            original_acc = str(row.get(account_col, '')).strip()
            # Ⅰ, Ⅱ 등 제거
            clean_acc = re.sub(r'^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.', '', original_acc).strip()
            clean_acc = re.sub(r'\(주석[0-9,]+\)', '', clean_acc).strip()

            if clean_acc and clean_acc not in ['매출액', '매출', '수익']:
                revenue_sub_items.append((clean_acc, acc))

        # 매출원가 하위 항목 수집
        if in_cogs_section and acc:
            original_acc = str(row.get(account_col, '')).strip()
            clean_acc = re.sub(r'^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.', '', original_acc).strip()
            clean_acc = re.sub(r'\(주석[0-9,]+\)', '', clean_acc).strip()

            if clean_acc and clean_acc not in ['매출원가', '원가']:
                cogs_sub_items.append((clean_acc, acc))

    print(f"[VCM] 매출 하위 항목 추출: {[item[0] for item in revenue_sub_items]}")
    print(f"[VCM] 매출원가 하위 항목 추출: {[item[0] for item in cogs_sub_items]}")

    # ========== 영업외수익/영업외비용 하위 항목 동적 추출 ==========
    # 매출/매출원가와 동일한 방식: 영업이익 이후 ~ 법인세비용차감전 사이의 모든 항목을 자동 분류
    non_op_income_items = []   # 영업외수익 하위 항목 [(항목명, 계정과목명), ...]
    non_op_expense_items = []  # 영업외비용 하위 항목 [(항목명, 계정과목명), ...]

    # 수익/이익 분류 키워드 (항목명에 포함되면 영업외수익으로 분류)
    income_indicators = ['수익', '이익', '차익', '환입', '잡이익']
    # 비용/손실 분류 키워드 (항목명에 포함되면 영업외비용으로 분류)
    expense_indicators = ['비용', '손실', '차손', '잡손실']
    # 손익 항목 (지분법손익 등 - 영업외수익/비용과 별도로 표시)
    mixed_indicators = ['손익']
    # 손익 항목 별도 수집 리스트
    mixed_items = []  # [(항목명, 계정과목명), ...]

    # 제외할 항목 (합계 항목들)
    exclude_keywords = ['영업외수익', '영업외비용', '법인세비용차감전', '세전', '당기순', '계속영업']

    # 영업외 섹션 범위: 영업이익 이후 ~ 법인세비용차감전
    non_op_end_keywords = ['법인세비용차감전', '법인세차감전', '세전이익', '세전순', '당기순이익', '당기순손실', '계속영업이익', '계속영업손실']

    after_op_income = False

    for idx, row in is_df.iterrows():
        acc = normalize(str(row.get(account_col, '')))
        original_acc = str(row.get(account_col, '')).strip()
        clean_acc = re.sub(r'^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.', '', original_acc).strip()
        clean_acc = re.sub(r'\(주석[0-9,]+\)', '', clean_acc).strip()

        # 영업이익/손실 이후부터 수집 시작
        if '영업이익' in acc or '영업손실' in acc or '영업손익' in acc:
            after_op_income = True
            continue

        # 법인세비용차감전 등에서 종료
        if any(end_kw in acc for end_kw in non_op_end_keywords):
            after_op_income = False
            continue

        if not after_op_income:
            continue

        # 제외 항목 건너뛰기
        if any(excl in acc for excl in exclude_keywords):
            continue

        # 빈 항목 건너뛰기
        if not clean_acc:
            continue

        # 항목명에 따라 수익/비용 자동 분류
        is_income = any(ind in acc for ind in income_indicators)
        is_expense = any(ind in acc for ind in expense_indicators)
        is_mixed = any(ind in acc for ind in mixed_indicators)

        # 수익으로 분류
        if is_income and not is_expense and not is_mixed:
            if (clean_acc, acc) not in non_op_income_items:
                non_op_income_items.append((clean_acc, acc))
        # 비용으로 분류
        elif is_expense and not is_income and not is_mixed:
            if (clean_acc, acc) not in non_op_expense_items:
                non_op_expense_items.append((clean_acc, acc))
        # 손익 항목 (지분법손익 등) - 영업외수익/비용과 별도로 수집
        elif is_mixed:
            if (clean_acc, acc) not in mixed_items:
                mixed_items.append((clean_acc, acc))
        # 둘 다 포함된 경우 (수익+비용 동시 포함)
        elif is_income and is_expense:
            if (clean_acc, acc) not in non_op_income_items:
                non_op_income_items.append((clean_acc, acc))

    print(f"[VCM] 영업외수익 하위 항목 추출: {[item[0] for item in non_op_income_items]}")
    print(f"[VCM] 영업외비용 하위 항목 추출: {[item[0] for item in non_op_expense_items]}")
    print(f"[VCM] 손익 항목 추출 (별도 표시): {[item[0] for item in mixed_items]}")

    # 손익계산서 항목 (동적으로 생성)
    is_items = []

    # 매출 항목
    if revenue_sub_items:
        # 매출 합계: 모든 하위 항목의 합
        revenue_keywords = [(item[1], []) for item in revenue_sub_items]
        is_items.append(('매출', revenue_keywords, 'sum'))

        # 매출 하위 항목들
        for item_name, item_keyword in revenue_sub_items:
            is_items.append((f'  {item_name}', [(item_keyword, ['원가'])], 'find'))
    else:
        # 추출 실패 시 기본 항목 사용
        is_items.append(('매출', [('영업수익', []), ('매출액', [])], 'find'))

    # 매출원가 항목
    if cogs_sub_items:
        # 매출원가 합계: 모든 하위 항목의 합
        cogs_keywords = [(item[1], []) for item in cogs_sub_items]
        is_items.append(('매출원가', cogs_keywords, 'sum'))

        # 매출원가 하위 항목들
        for item_name, item_keyword in cogs_sub_items:
            is_items.append((f'  {item_name}', [(item_keyword, [])], 'find'))
    else:
        # 추출 실패 시 기본 항목 사용
        is_items.append(('매출원가', [('영업비용', []), ('매출원가', [])], 'find'))

    # 나머지 손익계산서 항목 (동료 의견 반영: 판관비는 동적으로 금액순 선택)
    is_items.extend([
        ('매출총이익', [], 'calc_gross'),
        ('판매비와관리비', [], 'calc_sga'),
        # 판관비 하위항목은 아래에서 동적으로 추가됨
        ('영업이익', [], 'calc_op'),
    ])

    # 영업외수익 항목 (동적으로 추출된 하위 항목 사용)
    is_items.append(('영업외수익', [], 'find_direct'))  # 직접 찾기
    if non_op_income_items:
        for item_name, item_keyword in non_op_income_items:
            is_items.append((f'  {item_name}', [(item_keyword, [])], 'find'))
    else:
        # 기본 항목 (추출 실패 시)
        is_items.append(('  금융수익', [('이자수익', []), ('기타금융수익', [])], 'sum'))

    # 영업외비용 항목 (동적으로 추출된 하위 항목 사용)
    is_items.append(('영업외비용', [], 'find_direct'))  # 직접 찾기
    if non_op_expense_items:
        for item_name, item_keyword in non_op_expense_items:
            is_items.append((f'  {item_name}', [(item_keyword, [])], 'find'))
    else:
        # 기본 항목 (추출 실패 시)
        is_items.append(('  금융비용', [('이자비용', [])], 'find'))

    # 손익 항목 (지분법손익 등) - 영업외수익/비용과 별도로 표시
    if mixed_items:
        for item_name, item_keyword in mixed_items:
            is_items.append((item_name, [(item_keyword, [])], 'find'))

    is_items.extend([
        ('법인세비용차감전이익', [], 'calc_ebt'),
        ('법인세비용', [('법인세비용', ['차감전']), ('법인세등', ['차감전'])], 'find'),
        ('당기순이익', [], 'calc_net'),
    ])

    # ========== 판관비 상위 6개 항목 미리 결정 (모든 연도 합계 기준) ==========
    def get_sga_items_for_year(is_df, year):
        """특정 연도의 모든 판관비 항목 값을 반환"""
        급여 = find_val(is_df, ['급여', '직원급여', '종업원급여'], year, ['퇴직', '연차', '확정급여', '재측정', '보험수리', '순확정', '경영진', '만기', '지급액']) or 0
        퇴직급여 = find_val(is_df, ['퇴직급여'], year, ['확정급여', '재측정', '보험수리', '순확정', '경영진']) or 0
        복리후생비 = find_val(is_df, ['복리후생비'], year) or 0
        기타장기종업원급여 = find_val(is_df, ['기타장기종업원급여'], year) or 0
        주식보상비용 = find_val(is_df, ['주식보상비용', '주식기준보상비용'], year) or 0
        인건비 = 급여 + 퇴직급여 + 복리후생비 + 기타장기종업원급여 + 주식보상비용
        감가상각비 = find_val(is_df, ['감가상각비'], year, ['무형']) or 0

        return {
            '인건비': 인건비,
            '감가상각비': 감가상각비,
            '수수료비용': find_val(is_df, ['지급수수료', '수수료비용', '판매수수료'], year) or 0,
            '광고선전비': find_val(is_df, ['광고선전비'], year) or 0,
            '수도광열비': find_val(is_df, ['수도광열비'], year) or 0,
            '임차료비용': find_val(is_df, ['지급임차료', '임차료비용', '임차료'], year) or 0,
            '무형자산상각비': find_val(is_df, ['무형자산상각비'], year) or 0,
            '대손상각비': find_val(is_df, ['대손상각비'], year) or 0,
            '연구비': (find_val(is_df, ['경상연구개발비'], year) or 0) + (find_val(is_df, ['경상시험연구비'], year) or 0),
            '보험료': find_val(is_df, ['보험료'], year) or 0,
            '여비교통비': find_val(is_df, ['여비교통비'], year) or 0,
            '접대비': find_val(is_df, ['접대비'], year) or 0,
            '세금과공과': find_val(is_df, ['세금과공과'], year) or 0,
            '차량유지비': find_val(is_df, ['차량유지비'], year) or 0,
            '운반비': find_val(is_df, ['운반비'], year) or 0,
            '교육훈련비': find_val(is_df, ['교육훈련비'], year) or 0,
            '통신비': find_val(is_df, ['통신비'], year) or 0,
            '도서인쇄비': find_val(is_df, ['도서인쇄비'], year) or 0,
            '사무용품비': find_val(is_df, ['사무용품비'], year) or 0,
            '소모품비': find_val(is_df, ['소모품비'], year) or 0,
            '보관료': find_val(is_df, ['보관료'], year) or 0,
            '건물관리비': find_val(is_df, ['건물관리비'], year) or 0,
            '협회비': find_val(is_df, ['협회비'], year) or 0,
            '폐기물처리비': find_val(is_df, ['폐기물처리비'], year) or 0,
        }

    # 모든 연도의 합계 계산
    sga_totals = {}
    for year in fy_cols:
        sga_year = get_sga_items_for_year(is_df, year)
        for item_name, val in sga_year.items():
            sga_totals[item_name] = sga_totals.get(item_name, 0) + (val or 0)

    # 합계 기준 상위 5개 항목명 선택 (순서 유지를 위해 리스트 사용)
    # 인건비, 감가상각비는 별도 계산용이므로 제외하고 선택
    sga_totals_nonzero = {k: v for k, v in sga_totals.items() if v and v > 0 and k not in ['인건비', '감가상각비']}
    sga_sorted_total = sorted(sga_totals_nonzero.items(), key=lambda x: abs(x[1]), reverse=True)
    sga_top5_names = [item[0] for item in sga_sorted_total[:5]]  # 상위 5개 항목명 (순서 유지)

    # ========== 영업외수익 상위 6개 항목 미리 결정 (모든 연도 합계 기준) ==========
    def get_non_op_income_items_for_year(is_df, year):
        """특정 연도의 모든 영업외수익 항목 값을 반환"""
        return {
            '이자수익': find_val(is_df, ['이자수익'], year) or 0,
            '배당금수익': find_val(is_df, ['배당금수익'], year) or 0,
            '외환차익': find_val(is_df, ['외환차익'], year) or 0,
            '외화환산이익': find_val(is_df, ['외화환산이익'], year) or 0,
            '유형자산처분이익': find_val(is_df, ['유형자산처분이익'], year) or 0,
            '무형자산처분이익': find_val(is_df, ['무형자산처분이익'], year) or 0,
            '자산평가이익': find_val(is_df, ['자산평가이익', '당기손익공정가치측정금융자산평가이익'], year) or 0,
            '자산처분이익': find_val(is_df, ['당기손익인식금융자산처분이익', '금융자산처분이익'], year) or 0,
            '파생상품평가이익': find_val(is_df, ['파생상품평가이익'], year) or 0,
            '지분법이익': find_val(is_df, ['지분법이익', '관계기업투자이익'], year) or 0,
            '기타금융수익': find_val(is_df, ['기타금융수익'], year) or 0,
            '기타영업외수익': find_val(is_df, ['기타영업외수익', '기타수익', '잡이익'], year) or 0,
            # 금융수익 (일부 기업에서 합계 항목으로 표시) - 이자수익 등이 없을 때 fallback
            '금융수익': find_val(is_df, ['금융수익'], year, ['기타']) or 0,
        }

    def get_non_op_expense_items_for_year(is_df, year):
        """특정 연도의 모든 영업외비용 항목 값을 반환"""
        return {
            '이자비용': find_val(is_df, ['이자비용'], year) or 0,
            '외환차손': find_val(is_df, ['외환차손'], year) or 0,
            '외화환산손실': find_val(is_df, ['외화환산손실'], year) or 0,
            '유형자산처분손실': find_val(is_df, ['유형자산처분손실', '유무형리스자산처분손실'], year) or 0,
            '무형자산처분손실': find_val(is_df, ['무형자산처분손실'], year) or 0,
            '자산평가손실': find_val(is_df, ['자산평가손실', '당기손익공정가치측정금융자산평가손실'], year) or 0,
            '자산손상차손': find_val(is_df, ['자산손상차손', '유형자산손상차손', '무형자산손상차손'], year) or 0,
            '파생상품평가손실': find_val(is_df, ['파생상품평가손실'], year) or 0,
            '지분법손실': find_val(is_df, ['지분법손실'], year) or 0,
            '기타금융비용': find_val(is_df, ['기타금융비용'], year) or 0,
            '기타영업외비용': find_val(is_df, ['기타영업외비용', '기타비용', '잡손실'], year) or 0,
            # 금융비용 (일부 기업에서 합계 항목으로 표시) - 이자비용 등이 없을 때 fallback
            '금융비용': find_val(is_df, ['금융비용'], year, ['기타']) or 0,
        }

    # 영업외수익 모든 연도 합계 계산
    non_op_income_totals = {}
    for year in fy_cols:
        non_op_income_year = get_non_op_income_items_for_year(is_df, year)
        for item_name, val in non_op_income_year.items():
            non_op_income_totals[item_name] = non_op_income_totals.get(item_name, 0) + (val or 0)

    # 영업외수익 상위 5개 항목 선택 (기타영업외수익 제외)
    non_op_income_totals_nonzero = {k: v for k, v in non_op_income_totals.items() if v and v > 0 and k != '기타영업외수익'}
    non_op_income_sorted = sorted(non_op_income_totals_nonzero.items(), key=lambda x: abs(x[1]), reverse=True)
    non_op_income_top5_names = [item[0] for item in non_op_income_sorted[:5]]
    non_op_income_rest_names = [item[0] for item in non_op_income_sorted[5:]]  # 기타에 포함될 항목들
    print(f"[VCM] 영업외수익 상위 5개 항목: {non_op_income_top5_names}")

    # 영업외비용 모든 연도 합계 계산
    non_op_expense_totals = {}
    for year in fy_cols:
        non_op_expense_year = get_non_op_expense_items_for_year(is_df, year)
        for item_name, val in non_op_expense_year.items():
            non_op_expense_totals[item_name] = non_op_expense_totals.get(item_name, 0) + (val or 0)

    # 영업외비용 상위 5개 항목 선택 (기타영업외비용 제외)
    non_op_expense_totals_nonzero = {k: v for k, v in non_op_expense_totals.items() if v and v > 0 and k != '기타영업외비용'}
    non_op_expense_sorted = sorted(non_op_expense_totals_nonzero.items(), key=lambda x: abs(x[1]), reverse=True)
    non_op_expense_top5_names = [item[0] for item in non_op_expense_sorted[:5]]
    non_op_expense_rest_names = [item[0] for item in non_op_expense_sorted[5:]]  # 기타에 포함될 항목들
    print(f"[VCM] 영업외비용 상위 5개 항목: {non_op_expense_top5_names}")

    # ========== 서비스업 여부를 연도 루프 전에 미리 결정 ==========
    # 어떤 연도라도 판관비가 존재하면 제조업으로 판단 (첫 연도 기준 오류 방지)
    _pre_service_check = False
    for _check_year in fy_cols:
        _check_sga = find_val(is_df, ['판매비와관리비', '판매비및관리비', '판관비'], _check_year) or 0
        if _check_sga > 0:
            _pre_service_check = False
            break
        _check_opex = find_val(is_df, ['영업비용'], _check_year, ['매출원가', '원가']) or 0
        if _check_opex > 0:
            _pre_service_check = True
    print(f"[VCM] 서비스업 사전 판정: {_pre_service_check}")

    # 각 연도별 값 계산
    rows = []
    for year in fy_cols:
        # 표시용 컬럼명 (튜플이면 문자열로 변환)
        year_str = fy_col_map[year]
        row_data = {'항목': '', year_str: None}
        
        # 기본 값들 먼저 계산
        # 서비스업(영업수익/영업비용) vs 제조업(매출액/매출원가) 모두 지원
        # '매출' 키워드 추가: E1 등 '매출' 항목 사용 기업 지원 (기존 키워드 유지)
        영업수익 = find_val(is_df, ['영업수익', '매출액', '매출', '수익'], year, ['원가', '비용', '채권', '총이익']) or 0
        영업비용 = find_val(is_df, ['영업비용', '매출원가'], year) or 0

        # 동적으로 추출된 매출 항목 값 계산
        revenue_values = {}
        for item_name, item_keyword in revenue_sub_items:
            revenue_values[item_name] = find_val(is_df, [item_keyword], year, ['원가']) or 0

        # 동적으로 추출된 매출원가 항목 값 계산
        cogs_values = {}
        for item_name, item_keyword in cogs_sub_items:
            cogs_values[item_name] = find_val(is_df, [item_keyword], year) or 0
        # ========== 판관비 하위항목 (미리 결정된 상위 6개 항목 사용) ==========
        sga_items_all = get_sga_items_for_year(is_df, year)
        인건비 = sga_items_all['인건비']
        감가상각비 = sga_items_all['감가상각비']

        # 미리 결정된 상위 5개 항목만 선택 (합계 기준 순서 유지)
        sga_top5 = [(name, sga_items_all.get(name, 0)) for name in sga_top5_names]
        sga_top5_set = set(sga_top5_names)
        sga_rest = [(k, v) for k, v in sga_items_all.items() if k not in sga_top5_set and k not in ['인건비', '감가상각비'] and v and v > 0]

        # 기타판매비와관리비 계산
        기타판관비 = sum(v for _, v in sga_rest)

        # 표시용 리스트: 인건비 + 상위 5개 (총 6개)
        sga_top6 = [('인건비', 인건비)] + sga_top5

        # 영업외 수익/비용 (직접 찾기)
        영업외수익_direct = find_val(is_df, ['영업외수익'], year) or 0
        영업외비용_direct = find_val(is_df, ['영업외비용'], year) or 0

        # 영업외수익/비용 항목 값 계산 (새 함수 사용)
        non_op_income_year_vals = get_non_op_income_items_for_year(is_df, year)
        non_op_expense_year_vals = get_non_op_expense_items_for_year(is_df, year)

        # 금융수익/비용 (fallback용)
        이자수익 = non_op_income_year_vals.get('이자수익', 0)
        기타금융수익 = non_op_income_year_vals.get('기타금융수익', 0)
        금융수익 = 이자수익 + 기타금융수익
        이자비용 = non_op_expense_year_vals.get('이자비용', 0)
        금융비용 = 이자비용
        법인세 = find_val(is_df, ['법인세비용', '법인세등'], year, ['차감전']) or 0

        # 계산 값
        # 서비스업: 영업수익이 있으면 우선 사용
        매출 = 영업수익 if 영업수익 > 0 else sum(revenue_values.values())
        원가 = 영업비용 if 영업비용 > 0 else sum(cogs_values.values())

        # 매출총이익: 손익계산서에서 직접 찾기 시도, 없으면 계산
        매출총이익_direct = find_val(is_df, ['매출총이익', '총이익'], year) or 0
        매출총이익 = 매출총이익_direct if 매출총이익_direct else (매출 - 원가)

        # 판매비와관리비: 손익계산서에서 직접 찾기 시도
        판관비_direct = find_val(is_df, ['판매비와관리비', '판매비및관리비', '판관비'], year) or 0

        # 서비스업 여부 판단: 사전 판정 결과 사용 (첫 연도 기준 오류 방지)
        # 개별 연도에서 판관비가 0이어도 다른 연도에 판관비가 있으면 제조업
        is_service_business = _pre_service_check

        if is_service_business:
            # 서비스업: 판관비는 0 (영업비용에 이미 포함)
            판관비 = 0
        else:
            # 제조업 등: 판관비가 없으면 주석에서 계산 (인건비 + 상위5개 + 기타)
            판관비 = 판관비_direct if 판관비_direct else (인건비 + sum(v for _, v in sga_top5) + 기타판관비)

        # 영업이익: 손익계산서에서 직접 찾기 시도 (항상 원본값 우선)
        # 영업손실은 음수로 변환 필요
        영업이익_direct = find_val(is_df, ['영업이익', '영업손익'], year) or 0
        if not 영업이익_direct:
            영업손실_direct = find_val(is_df, ['영업손실'], year) or 0
            if 영업손실_direct > 0:
                영업이익_direct = -영업손실_direct  # 영업손실은 음수로 변환
        영업이익 = 영업이익_direct if 영업이익_direct else (매출총이익 - 판관비)

        # 영업외수익/비용: 직접 찾은 값 > 항목 합계 > fallback
        항목_영업외수익 = sum(non_op_income_year_vals.values())
        항목_영업외비용 = sum(non_op_expense_year_vals.values())
        영업외수익 = 영업외수익_direct if 영업외수익_direct else (항목_영업외수익 if 항목_영업외수익 else 금융수익)
        영업외비용 = 영업외비용_direct if 영업외비용_direct else (항목_영업외비용 if 항목_영업외비용 else 금융비용)

        # 법인세비용차감전이익: 손익계산서에서 직접 찾기 시도
        세전이익_direct = find_val(is_df, ['법인세비용차감전순이익', '법인세비용차감전이익', '법인세차감전순이익', '세전이익'], year) or 0
        세전이익 = 세전이익_direct if 세전이익_direct else (영업이익 + 영업외수익 - 영업외비용)

        # 당기순이익: 손익계산서에서 직접 찾기 시도
        # 당기순손실은 음수로 변환 필요
        당기순이익_direct = find_val(is_df, ['당기순이익', '당기순손익', '총당기순이익'], year) or 0
        if not 당기순이익_direct:
            당기순손실_direct = find_val(is_df, ['당기순손실'], year) or 0
            if 당기순손실_direct > 0:
                당기순이익_direct = -당기순손실_direct
        당기순이익 = 당기순이익_direct if 당기순이익_direct else (세전이익 - (법인세 or 0))
        
        values = {
            '매출': 매출,
        }

        # 매출 하위 항목 동적 추가
        for item_name in [item[0] for item in revenue_sub_items]:
            values[f'  {item_name}'] = revenue_values.get(item_name) if revenue_values.get(item_name) else None

        # 매출원가 추가
        values['매출원가'] = 원가

        # 매출원가 하위 항목 동적 추가
        for item_name in [item[0] for item in cogs_sub_items]:
            values[f'  {item_name}'] = cogs_values.get(item_name) if cogs_values.get(item_name) else None

        # 서비스업: 비용 세부항목을 매출원가(영업비용) 하위로 표시
        if is_service_business:
            # 금액순 상위 6개 동적 추가 (인건비, 감가상각비 포함)
            for item_name, item_val in sga_top6:
                values[f'  {item_name}'] = item_val if item_val else None
            # 기타영업비용
            values['  기타영업비용'] = 기타판관비 if 기타판관비 else None

        # EBITDA 계산 (영업이익 + 감가상각비 + 무형자산상각비) - 미리 계산
        무형자산상각비 = sga_items_all.get('무형자산상각비', 0) or 0
        EBITDA = 영업이익 + 감가상각비 + 무형자산상각비

        # % of Sales 계산 (소수점 형태로 저장, 예: 0.127 = 12.7%)
        매출총이익_pct = (매출총이익 / 매출) if 매출 and 매출 != 0 else None
        영업이익_pct = (영업이익 / 매출) if 매출 and 매출 != 0 else None
        당기순이익_pct = (당기순이익 / 매출) if 매출 and 매출 != 0 else None
        EBITDA_pct = (EBITDA / 매출) if 매출 and 매출 != 0 else None

        # 매출총이익 및 % of Sales
        values['매출총이익'] = 매출총이익
        values['% of Sales'] = 매출총이익_pct  # 매출총이익률

        # 판매비와관리비 (서비스업은 생략)
        if is_service_business:
            # 서비스업: 판관비 항목 자체를 생략 (영업비용에 이미 포함)
            pass
        else:
            # 제조업 등: 판관비 및 하위 항목 표시
            values['판매비와관리비'] = 판관비

            # 판관비 하위 항목: 금액순 상위 6개만 표시 (인건비, 감가상각비 포함)
            for item_name, item_val in sga_top6:
                values[f'  {item_name}'] = item_val if item_val else None

            # 기타판매비와관리비
            values['  기타판매비와관리비'] = 기타판관비 if 기타판관비 else None
            # 기타판매비와관리비 세부항목 추가 (툴팁용)
            for item_name, val in sga_rest:
                if val:
                    values[f'    {item_name}'] = val

        # 나머지 손익계산서 항목
        values.update({
            '영업이익': 영업이익,
            '% of Sales (영업이익)': 영업이익_pct,  # 영업이익률
            '영업외수익': 영업외수익,
        })

        # 영업외수익 하위 항목 (상위 5개 + 기타)
        non_op_income_year = get_non_op_income_items_for_year(is_df, year)
        non_op_income_top5_sum = 0
        non_op_income_total_sum = sum(non_op_income_year.values())

        for item_name in non_op_income_top5_names:
            val = non_op_income_year.get(item_name, 0)
            values[f'  {item_name}'] = val if val else None
            non_op_income_top5_sum += val

        # 기타영업외수익 = 전체 합계 - 상위 5개 합계
        기타영업외수익 = non_op_income_total_sum - non_op_income_top5_sum

        # 기타영업외수익 세부항목 수집 (툴팁용)
        기타영업외수익_세부 = []
        for item_name in non_op_income_rest_names:
            val = non_op_income_year.get(item_name, 0)
            if val:
                기타영업외수익_세부.append((item_name, val))

        # 기타영업외수익 (값이 있거나 세부항목이 있으면 추가)
        if 기타영업외수익 > 0 or 기타영업외수익_세부:
            values['  기타영업외수익'] = 기타영업외수익 if 기타영업외수익 > 0 else None
            # 세부항목 추가 (기타영업외수익 바로 다음에)
            for item_name, val in 기타영업외수익_세부:
                values[f'    {item_name}'] = val

        values['영업외비용'] = 영업외비용

        # 영업외비용 하위 항목 (상위 5개 + 기타)
        non_op_expense_year = get_non_op_expense_items_for_year(is_df, year)
        non_op_expense_top5_sum = 0
        non_op_expense_total_sum = sum(non_op_expense_year.values())

        for item_name in non_op_expense_top5_names:
            val = non_op_expense_year.get(item_name, 0)
            values[f'  {item_name}'] = val if val else None
            non_op_expense_top5_sum += val

        # 기타영업외비용 = 전체 합계 - 상위 5개 합계
        기타영업외비용 = non_op_expense_total_sum - non_op_expense_top5_sum

        # 기타영업외비용 세부항목 수집 (툴팁용)
        기타영업외비용_세부 = []
        for item_name in non_op_expense_rest_names:
            val = non_op_expense_year.get(item_name, 0)
            if val:
                기타영업외비용_세부.append((item_name, val))

        # 기타영업외비용 (값이 있거나 세부항목이 있으면 추가)
        if 기타영업외비용 > 0 or 기타영업외비용_세부:
            values['  기타영업외비용'] = 기타영업외비용 if 기타영업외비용 > 0 else None
            # 세부항목 추가 (기타영업외비용 바로 다음에)
            for item_name, val in 기타영업외비용_세부:
                values[f'    {item_name}'] = val

        values.update({
            '법인세비용차감전이익': 세전이익,
            '법인세비용': 법인세 if 법인세 else None,
            '당기순이익': 당기순이익,
            '% of Sales (순이익)': 당기순이익_pct,  # 당기순이익률
            'EBITDA': EBITDA,
            '  [EBITDA]영업이익': 영업이익,  # EBITDA 툴팁용
            '  [EBITDA]감가상각비': 감가상각비,  # EBITDA 툴팁용
            '  [EBITDA]무형자산상각비': 무형자산상각비,  # EBITDA 툴팁용
            '% of Sales (EBITDA)': EBITDA_pct,  # EBITDA 마진
        })

        # 손익계산서 세부항목 - 이미 values에서 추가됨, 여기서는 빈 배열
        is_items_detail = []

        # 첫 번째 연도에서 기존 항목 저장 (부모 설정)
        if year == fy_cols[0]:
            current_parent = ''  # 현재 섹션의 부모 항목
            last_single_indent = ''  # 마지막 싱글 들여쓰기 항목
            for item_name in values:
                # 들여쓰기 없는 항목은 부모 항목 (섹션 헤더)
                if not item_name.startswith('  '):
                    current_parent = item_name
                    last_single_indent = ''
                    rows.append({'항목': item_name, '부모': ''})
                elif item_name.startswith('    '):
                    # 더블 들여쓰기 (4칸) = 기타영업외수익/비용의 세부항목
                    # 부모는 직전 싱글 들여쓰기 항목 (기타영업외수익 또는 기타영업외비용)
                    rows.append({'항목': item_name, '부모': last_single_indent})
                else:
                    # 싱글 들여쓰기 (2칸) = 섹션의 하위 항목
                    last_single_indent = item_name
                    rows.append({'항목': item_name, '부모': current_parent})

            # 세부항목 추가
            for item_name, parent, val in is_items_detail:
                # 이미 추가된 항목은 건너뜀
                if item_name not in [r['항목'] for r in rows]:
                    rows.append({'항목': item_name, '부모': parent})
        else:
            # 이후 연도에서 새로운 항목이 있으면 적절한 위치에 삽입
            existing_items = [r['항목'] for r in rows]
            for item_name in values:
                if item_name not in existing_items:
                    # 새로운 항목 발견 - 적절한 위치에 삽입
                    if item_name.startswith('    '):
                        # 더블 들여쓰기 (4칸) = 기타 항목의 세부항목
                        # 기타판매비와관리비, 기타영업외수익, 기타영업외비용 바로 뒤에 삽입
                        target_parent = ''
                        insert_idx = len(rows)
                        for i in range(len(rows)):
                            row_item = rows[i]['항목']
                            is_기타_parent = ('기타영업외' in row_item or '기타판매비와관리비' in row_item) and row_item.startswith('  ') and not row_item.startswith('    ')
                            if is_기타_parent:
                                target_parent = row_item
                                insert_idx = i + 1
                                # 기존 더블 들여쓰기 뒤에 삽입
                                while insert_idx < len(rows) and rows[insert_idx]['항목'].startswith('    '):
                                    insert_idx += 1
                        rows.insert(insert_idx, {'항목': item_name, '부모': target_parent})
                    elif item_name.startswith('  '):
                        # 싱글 들여쓰기 (2칸) = 섹션의 하위 항목
                        target_parent = ''
                        insert_idx = len(rows)
                        parent_candidates = ['매출원가', '판매비와관리비', '영업외수익', '영업외비용']
                        for i in range(len(rows) - 1, -1, -1):
                            if rows[i]['항목'].startswith('  '):
                                continue
                            elif rows[i]['항목'] in parent_candidates:
                                target_parent = rows[i]['항목']
                                insert_idx = i + 1
                                while insert_idx < len(rows) and rows[insert_idx]['항목'].startswith('  '):
                                    insert_idx += 1
                                break
                        rows.insert(insert_idx, {'항목': item_name, '부모': target_parent})
                    else:
                        rows.append({'항목': item_name, '부모': ''})

        # 각 연도별 값 저장
        for r in rows:
            item_name = r['항목']
            if item_name in values:
                val = values[item_name]
                # % of Sales는 소수점 값이므로 round() 적용하지 않음
                if val is not None and val != 0:
                    if '% of Sales' in item_name:
                        r[year_str] = val  # 소수점 그대로 저장
                    else:
                        r[year_str] = round(val)
                else:
                    r[year_str] = None
            else:
                # 세부항목에서 찾기
                for is_item_name, parent, is_val in is_items_detail:
                    if item_name == is_item_name:
                        r[year_str] = round(is_val) if is_val is not None and is_val != 0 else None
                        break

    # ========== 재무상태표 항목 추가 ==========
    # 재무상태표에서 값 찾기 (원 단위 그대로)
    def find_bs_val(keywords, year, excludes=[]):
        # BS용 컬럼 사용 (IS와 다를 수 있음)
        for _, row in bs_df.iterrows():
            acc_raw = str(row.get(bs_account_col, ''))
            acc = normalize(acc_raw)
            excluded = any(normalize(ex) in acc for ex in excludes)
            if excluded: continue
            for kw in keywords:
                if normalize(kw) in acc:
                    # BS 연도 컬럼 매핑 사용
                    bs_year = year
                    if bs_fy_col_map and fy_col_map:
                        year_str = fy_col_map.get(year, '')
                        for bs_col, bs_year_str in bs_fy_col_map.items():
                            if bs_year_str == year_str:
                                bs_year = bs_col
                                break
                    val = row.get(bs_year)
                    if pd.notna(val):
                        try:
                            val_str = str(val).replace(',', '').strip()
                            # 괄호로 표시된 음수 처리: (1234) → -1234
                            if val_str.startswith('(') and val_str.endswith(')'):
                                val_str = '-' + val_str[1:-1]
                            return float(val_str)
                        except:
                            pass
        return None

    # 재무상태표 행 생성 (부모 컬럼 포함, 세부항목 별도 저장)
    all_bs_items_by_year = {}  # {year_str: bs_items list}

    # ========== 섹션 기반 파싱으로 모든 연도 데이터 수집 ==========
    all_sections = {}  # {year: parsed_sections}

    # BS용 연도 컬럼이 있으면 사용, 없으면 IS 연도 컬럼 사용
    bs_fy_cols = sorted(bs_fy_col_map.keys(), key=lambda c: bs_fy_col_map[c]) if bs_fy_col_map else fy_cols
    bs_col_display = bs_fy_col_map if bs_fy_col_map else fy_col_map

    for bs_year in bs_fy_cols:
        year_str = bs_col_display[bs_year]
        # XBRL 형식이면 분류3 컬럼 기반 파싱, 아니면 행 기반 파싱
        if is_xbrl and '분류3' in bs_df.columns:
            all_sections[year_str] = parse_bs_sections_xbrl(bs_df, bs_year, bs_account_col)
            print(f"[VCM XBRL] {year_str} 섹션 파싱 완료: 유동자산 {len(all_sections[year_str]['유동자산'])}개, 비유동자산 {len(all_sections[year_str]['비유동자산'])}개 항목")
        else:
            all_sections[year_str] = parse_bs_sections(bs_df, bs_year, bs_account_col)
            print(f"[VCM] {year_str} 섹션 파싱 완료: 유동자산 {len(all_sections[year_str]['유동자산'])}개, 비유동자산 {len(all_sections[year_str]['비유동자산'])}개 항목")

    # 항목 정의: (항목명, 부모, 값계산함수)
    # 부모가 있으면 세부항목, 없으면 합계 항목

    for year in fy_cols:
        year_str = fy_col_map[year]
        sections = all_sections.get(year_str, {'유동자산': [], '비유동자산': [], '유동부채': [], '비유동부채': [], '자본': [], '총계': {}})

        # ========== 자산 항목 (섹션 기반) ==========
        # 유동자산 섹션에서 항목 찾기 (복사본 생성 - fallback 항목 추가용)
        유동자산_items = list(sections['유동자산'])
        비유동자산_items = list(sections['비유동자산'])
        유동부채_items = list(sections['유동부채'])
        비유동부채_items = list(sections['비유동부채'])
        자본_items = sections['자본']
        총계 = sections['총계']

        # 섹션 경계 밖의 항목을 find_bs_val로 찾아 섹션에 추가하는 헬퍼 함수
        def add_fallback_to_section(section_items, item_name, keywords, excludes=[], other_sections=[]):
            """섹션에 없는 항목을 find_bs_val로 찾아서 추가 (다른 섹션 중복 방지)"""
            if not find_in_section(section_items, keywords, excludes):
                # 다른 섹션에 이미 있으면 추가하지 않음 (cross-section contamination 방지)
                for other in other_sections:
                    if find_in_section(other, keywords, excludes):
                        return None
                val = find_bs_val(keywords, year, excludes)
                if val and val != 0:
                    section_items.append({'name': item_name, 'value': val})
                    print(f"[VCM FALLBACK] {year_str} {item_name}: {val:,.0f} 추가")
                    return val
            return None

        # 유동자산 fallback 항목 추가 (섹션 경계 밖 데이터 보완, 비유동자산과 cross-check)
        add_fallback_to_section(유동자산_items, '단기금융상품', ['단기금융상품'], [],
                                other_sections=[비유동자산_items])
        add_fallback_to_section(유동자산_items, '매출채권', ['매출채권'], ['장기', '손실', '처분'],
                                other_sections=[비유동자산_items])
        add_fallback_to_section(유동자산_items, '단기대여금', ['단기대여금'], [],
                                other_sections=[비유동자산_items])

        # 유동부채 fallback 항목 추가 (비유동부채와 cross-check)
        add_fallback_to_section(유동부채_items, '단기차입금', ['단기차입금'], [],
                                other_sections=[비유동부채_items])
        add_fallback_to_section(유동부채_items, '매입채무', ['매입채무'], ['장기'],
                                other_sections=[비유동부채_items])
        add_fallback_to_section(유동부채_items, '미지급금', ['미지급금'], ['장기'],
                                other_sections=[비유동부채_items])

        # 비유동자산 fallback 항목 추가 (섹션 경계 밖 데이터 보완, 유동자산과 cross-check)
        add_fallback_to_section(비유동자산_items, '유형자산', ['유형자산'], ['무형', '처분', '사용권'],
                                other_sections=[유동자산_items])
        add_fallback_to_section(비유동자산_items, '장기금융상품', ['장기금융상품', '장기투자자산'], [],
                                other_sections=[유동자산_items])
        add_fallback_to_section(비유동자산_items, '무형자산', ['무형자산'], ['상각', '손상'],
                                other_sections=[유동자산_items])
        add_fallback_to_section(비유동자산_items, '기타의투자자산', ['기타의투자자산', '기타투자자산'], [],
                                other_sections=[유동자산_items])
        add_fallback_to_section(비유동자산_items, '보증금', ['보증금', '임차보증금'], [],
                                other_sections=[유동자산_items])

        # 비유동부채 fallback 항목 추가 (유동성, 유동, 전환 등 유동 항목 제외, 유동부채와 cross-check)
        add_fallback_to_section(비유동부채_items, '장기차입금', ['장기차입금'], ['유동성', '유동'],
                                other_sections=[유동부채_items])
        add_fallback_to_section(비유동부채_items, '사채', ['사채'], ['유동성', '유동', '전환'],
                                other_sections=[유동부채_items])
        add_fallback_to_section(비유동부채_items, '장기미지급금', ['장기미지급금'], [],
                                other_sections=[유동부채_items])
        add_fallback_to_section(비유동부채_items, '임대보증금', ['임대보증금'], [],
                                other_sections=[유동부채_items])

        # 유동자산 총계: 총계 우선, find_bs_val, 마지막으로 items합
        유동자산_from_bs = find_bs_val(['유동자산'], year, ['비유동']) or 0
        유동자산_from_items = sum(item['value'] for item in 유동자산_items) if 유동자산_items else 0
        유동자산 = 총계.get('유동자산') or 유동자산_from_bs or 유동자산_from_items or 0

        # 주요 항목 찾기 (섹션 우선, fallback으로 키워드 검색)
        현금 = find_in_section(유동자산_items, ['현금및현금성자산', '현금']) or find_bs_val(['현금및현금성자산'], year) or 0

        # 단기투자자산: 단기금융상품 + 당기손익-공정가치측정금융자산 + 기타포괄손익-공정가치 등
        단기금융상품_직접 = find_in_section(유동자산_items, ['단기금융상품']) or find_bs_val(['단기금융상품'], year) or 0
        당기손익FVPL_유동 = find_in_section(유동자산_items, ['당기손익-공정가치측정금융자산', '당기손익공정가치측정금융자산', 'FVPL금융자산']) or 0
        기타포괄FVOCI_유동 = find_in_section(유동자산_items, ['기타포괄손익-공정가치측정금융자산', '기타포괄손익공정가치측정금융자산']) or 0
        단기금융상품 = 단기금융상품_직접 + 당기손익FVPL_유동 + 기타포괄FVOCI_유동

        재고자산 = find_in_section(유동자산_items, ['재고자산']) or find_bs_val(['재고자산'], year) or 0

        # 매출채권및기타채권 세부항목 (섹션에서 찾기)
        매출채권 = find_in_section(유동자산_items, ['매출채권'], ['장기', '손실', '처분']) or find_bs_val(['매출채권'], year, ['장기']) or 0
        미수금 = find_in_section(유동자산_items, ['미수금'], ['장기']) or find_bs_val(['미수금'], year) or 0
        미수수익 = find_in_section(유동자산_items, ['미수수익']) or find_bs_val(['미수수익'], year) or 0
        선급금 = find_in_section(유동자산_items, ['선급금']) or find_bs_val(['선급금'], year) or 0
        선급비용 = find_in_section(유동자산_items, ['선급비용']) or find_bs_val(['선급비용'], year) or 0
        계약자산_유동 = find_in_section(유동자산_items, ['계약자산']) or 0
        기타금융자산_유동 = find_in_section(유동자산_items, ['기타금융자산']) or 0
        매출채권및기타채권 = 매출채권 + 미수금 + 미수수익 + 선급금 + 선급비용 + 계약자산_유동 + 기타금융자산_유동

        # 기타비금융자산 (매칭되지 않은 유동자산 항목들)
        matched_유동자산_keywords = [
            ['현금및현금성자산', '현금'], ['단기금융상품'], ['당기손익-공정가치측정금융자산', '당기손익공정가치측정금융자산'],
            ['기타포괄손익-공정가치측정금융자산'], ['재고자산'],
            ['매출채권'], ['미수금'], ['미수수익'], ['선급금'], ['선급비용'], ['계약자산'], ['기타금융자산'],
            ['법인세'], ['매각예정'],
            ['파생상품자산', '파생상품'],  # 파생상품자산이 재고자산으로 오분류되지 않도록
            ['기타자산'],  # 기타자산 누락 방지
        ]
        기타유동자산_items = get_unmatched_items(유동자산_items, matched_유동자산_keywords)
        기타비금융자산 = sum(item['value'] for item in 기타유동자산_items) if 기타유동자산_items else (find_bs_val(['기타유동자산', '기타비금융자산'], year) or 0)

        # ========== 비유동자산 항목 (섹션 기반) ==========
        # 비유동자산 총계: 총계 우선, find_bs_val (매각예정 제외), 마지막으로 items합
        비유동자산_from_bs = find_bs_val(['비유동자산'], year, ['매각예정']) or 0
        비유동자산_from_items = sum(item['value'] for item in 비유동자산_items) if 비유동자산_items else 0
        비유동자산 = 총계.get('비유동자산') or 비유동자산_from_bs or 비유동자산_from_items or 0
        유형자산 = find_in_section(비유동자산_items, ['유형자산'], ['무형', '처분', '사용권']) or find_bs_val(['유형자산'], year, ['무형', '처분']) or 0
        무형자산 = find_in_section(비유동자산_items, ['무형자산'], ['상각', '손상']) or find_bs_val(['무형자산'], year, ['상각']) or 0
        사용권자산 = find_in_section(비유동자산_items, ['사용권자산']) or 0

        # 장기투자자산 세부항목 (섹션 기반)
        장기금융상품 = find_in_section(비유동자산_items, ['장기금융상품', '장기투자자산']) or find_bs_val(['장기금융상품'], year) or 0
        매도가능증권 = find_in_section(비유동자산_items, ['매도가능증권', '매도가능금융자산', '기타포괄손익-공정가치']) or find_bs_val(['매도가능증권', '매도가능금융자산'], year) or 0
        지분법투자 = find_in_section(비유동자산_items, ['지분법적용투자', '관계기업투자', '종속기업투자']) or find_bs_val(['지분법적용투자', '관계기업투자', '종속기업투자'], year) or 0
        장기투자자산 = 장기금융상품 + 매도가능증권 + 지분법투자

        보증금 = find_in_section(비유동자산_items, ['보증금', '임차보증금']) or find_bs_val(['보증금', '임차보증금'], year) or 0

        # 매출채권및기타채권 (비유동) - 장기매출채권 등
        장기매출채권 = find_in_section(비유동자산_items, ['매출채권', '장기매출채권']) or 0
        자산총계 = 총계.get('자산총계') or find_bs_val(['자산총계'], year) or 0

        # ========== 부채 항목 (섹션 기반) ==========
        # 유동부채 총계: 총계 우선, find_bs_val, 마지막으로 items합
        유동부채_from_bs = find_bs_val(['유동부채'], year, ['비유동']) or 0
        유동부채_from_items = sum(item['value'] for item in 유동부채_items) if 유동부채_items else 0
        유동부채 = 총계.get('유동부채') or 유동부채_from_bs or 유동부채_from_items or 0

        # 유동 매입채무및기타채무 세부항목 (섹션 기반)
        매입채무 = find_in_section(유동부채_items, ['매입채무'], ['장기']) or find_bs_val(['매입채무'], year, ['장기']) or 0
        미지급금_유동 = find_in_section(유동부채_items, ['미지급금'], ['장기']) or find_bs_val(['미지급금'], year, ['장기']) or 0
        미지급비용_유동 = find_in_section(유동부채_items, ['미지급비용'], ['장기']) or find_bs_val(['미지급비용'], year, ['장기']) or 0
        선수금_유동 = find_in_section(유동부채_items, ['선수금'], ['장기']) or find_bs_val(['선수금'], year, ['장기']) or 0
        선수수익_유동 = find_in_section(유동부채_items, ['선수수익'], ['장기']) or find_bs_val(['선수수익'], year, ['장기']) or 0
        예수금_유동 = find_in_section(유동부채_items, ['예수금'], ['장기']) or find_bs_val(['예수금'], year, ['장기']) or 0
        계약부채_유동 = find_in_section(유동부채_items, ['계약부채']) or 0
        예수보증금_유동 = find_in_section(유동부채_items, ['예수보증금'], ['장기', '임대']) or find_bs_val(['예수보증금'], year, ['장기', '임대']) or 0
        연차충당부채 = find_in_section(유동부채_items, ['연차충당부채', '연차수당충당부채']) or find_bs_val(['연차충당부채', '연차수당충당부채'], year) or 0
        유동매입채무및기타채무 = 매입채무 + 미지급금_유동 + 미지급비용_유동 + 선수금_유동 + 선수수익_유동 + 예수금_유동 + 예수보증금_유동 + 연차충당부채 + 계약부채_유동

        # 차입금 항목 (섹션 기반) - 전환사채, 전환우선주부채 등 포함
        단기차입금 = find_in_section(유동부채_items, ['단기차입금']) or find_bs_val(['단기차입금'], year) or 0
        유동성장기차입금 = find_in_section(유동부채_items, ['유동성장기부채', '유동성장기차입금']) or find_bs_val(['유동성장기부채', '유동성장기차입금'], year) or 0
        유동성사채 = find_in_section(유동부채_items, ['유동성사채']) or find_bs_val(['유동성사채'], year) or 0
        # 유동성사채가 섹션에 없지만 find_bs_val로 찾았으면 유동부채_items에 추가 (동적 분류용)
        if 유동성사채 and not find_in_section(유동부채_items, ['유동성사채']):
            유동부채_items.append({'name': '유동성사채', 'value': 유동성사채})
        전환사채_유동 = find_in_section(유동부채_items, ['전환사채']) or find_bs_val(['전환사채'], year) or 0
        상환전환우선주부채 = find_in_section(유동부채_items, ['상환전환우선주부채', 'RCPS부채']) or 0
        전환우선주부채 = find_in_section(유동부채_items, ['전환우선주부채', 'CPS부채'], ['상환']) or 0  # 상환전환우선주부채 제외

        # 유동차입부채 합계 (사용자 포맷에 맞춤)
        유동차입부채 = 단기차입금 + 유동성장기차입금 + 유동성사채 + 전환사채_유동 + 상환전환우선주부채 + 전환우선주부채

        # 기타금융부채 (리스부채 등) - 차입금에 포함된 항목 제외
        리스부채_유동 = find_in_section(유동부채_items, ['리스부채', '금융리스부채']) or 0
        파생상품부채_유동 = find_in_section(유동부채_items, ['파생상품부채']) or 0
        기타금융부채_유동 = find_in_section(유동부채_items, ['기타금융부채'], ['리스', '파생']) or 0

        # 기타유동부채 세부항목 (섹션 기반)
        미지급법인세 = find_in_section(유동부채_items, ['미지급법인세', '당기법인세부채']) or find_bs_val(['미지급법인세', '당기법인세부채'], year) or 0
        충당부채_유동 = find_in_section(유동부채_items, ['충당부채', '단기충당부채'], ['연차', '퇴직']) or 0

        # 기타유동부채 (매칭되지 않은 유동부채 항목들)
        matched_유동부채_keywords = [
            ['매입채무'], ['미지급금'], ['미지급비용'], ['선수금'], ['선수수익'], ['예수금'], ['예수보증금'],
            ['연차충당부채'], ['계약부채'], ['단기차입금'], ['유동성장기부채', '유동성장기차입금'],
            ['유동성사채'], ['전환사채'], ['상환전환우선주부채'], ['전환우선주부채'],
            ['리스부채', '금융리스부채'], ['파생상품부채'], ['기타금융부채'],
            ['미지급법인세', '당기법인세부채'], ['충당부채']
        ]
        기타유동부채_items = get_unmatched_items(유동부채_items, matched_유동부채_keywords)
        기타유동부채_기타 = sum(item['value'] for item in 기타유동부채_items) if 기타유동부채_items else (find_bs_val(['기타유동부채'], year) or 0)
        # 기타금융부채 합계 (리스부채, 파생상품부채 등 포함)
        기타금융부채_합계 = 리스부채_유동 + 파생상품부채_유동 + 기타금융부채_유동
        기타유동부채 = 미지급법인세 + 기타유동부채_기타 + 충당부채_유동 + 기타금융부채_합계

        # ========== 비유동부채 항목 (섹션 기반) ==========
        # 비유동부채 총계: 총계 우선, find_bs_val (매각예정 제외), 마지막으로 items합
        비유동부채_from_bs = find_bs_val(['비유동부채'], year, ['매각예정']) or 0
        비유동부채_from_items = sum(item['value'] for item in 비유동부채_items) if 비유동부채_items else 0
        비유동부채 = 총계.get('비유동부채') or 비유동부채_from_bs or 비유동부채_from_items or 0

        # 비유동 매입채무및기타채무 세부항목 (섹션 기반)
        장기매입채무 = find_in_section(비유동부채_items, ['장기매입채무', '매입채무']) or find_bs_val(['장기매입채무'], year) or 0
        장기미지급금 = find_in_section(비유동부채_items, ['장기미지급금', '비유동미지급금']) or find_bs_val(['장기미지급금'], year) or 0
        장기미지급비용 = find_in_section(비유동부채_items, ['장기미지급비용']) or find_bs_val(['장기미지급비용'], year) or 0
        장기선수금 = find_in_section(비유동부채_items, ['장기선수금']) or find_bs_val(['장기선수금'], year) or 0
        장기선수수익 = find_in_section(비유동부채_items, ['장기선수수익']) or find_bs_val(['장기선수수익'], year) or 0
        장기예수금 = find_in_section(비유동부채_items, ['장기예수금']) or find_bs_val(['장기예수금'], year) or 0
        임대보증금 = find_in_section(비유동부채_items, ['임대보증금']) or find_bs_val(['임대보증금'], year) or 0
        예수보증금_비유동 = find_in_section(비유동부채_items, ['예수보증금', '장기예수보증금']) or find_bs_val(['예수보증금', '장기예수보증금'], year) or 0
        계약부채_비유동 = find_in_section(비유동부채_items, ['계약부채']) or 0
        비유동매입채무및기타채무 = 장기매입채무 + 장기미지급금 + 장기미지급비용 + 장기선수금 + 장기선수수익 + 장기예수금 + 임대보증금 + 예수보증금_비유동 + 계약부채_비유동

        # 비유동 사채: '유동성', '유동', '전환' 제외 (유동전환사채 등 유동 항목 제외)
        사채 = find_in_section(비유동부채_items, ['사채'], ['유동성', '유동', '전환']) or 0
        # find_bs_val fallback 시에도 cross-section check — 유동부채에 이미 있으면 비유동에 추가하지 않음
        if not 사채:
            if not find_in_section(유동부채_items, ['사채'], ['유동성', '유동', '전환']):
                사채 = find_bs_val(['사채'], year, ['유동성', '유동', '전환']) or 0
                if 사채:
                    비유동부채_items.append({'name': '사채', 'value': 사채})
        장기차입금 = find_in_section(비유동부채_items, ['장기차입금', '비유동차입부채'], ['유동성']) or find_bs_val(['장기차입금'], year, ['유동성']) or 0
        퇴직급여채무 = find_in_section(비유동부채_items, ['퇴직급여충당부채', '퇴직급여채무', '확정급여채무', '순확정급여부채']) or find_bs_val(['퇴직급여충당부채', '퇴직급여채무', '확정급여채무'], year) or 0
        기타금융부채_비유동 = find_in_section(비유동부채_items, ['기타금융부채', '금융리스부채', '리스부채']) or 0
        충당부채_비유동 = find_in_section(비유동부채_items, ['충당부채', '장기충당부채'], ['퇴직']) or 0
        부채총계 = 총계.get('부채총계') or find_bs_val(['부채총계'], year) or 0

        # ========== 자본 항목 (섹션 기반) ==========
        자본금 = find_in_section(자본_items, ['자본금'], ['잉여금']) or find_bs_val(['자본금'], year, ['잉여금']) or 0
        이익잉여금 = find_in_section(자본_items, ['이익잉여금', '미처분이익잉여금', '결손금']) or find_bs_val(['이익잉여금', '미처분이익잉여금'], year) or 0

        # 기타자본구성요소 세부항목 (섹션 기반)
        자본잉여금 = find_in_section(자본_items, ['자본잉여금']) or find_bs_val(['자본잉여금'], year) or 0
        자본조정 = find_in_section(자본_items, ['자본조정']) or find_bs_val(['자본조정'], year) or 0
        기타포괄손익누계액 = find_in_section(자본_items, ['기타포괄손익누계액']) or find_bs_val(['기타포괄손익누계액'], year) or 0
        기타자본항목 = find_in_section(자본_items, ['기타자본', '기타자본구성요소', '기타자본항목']) or 0
        기타자본 = 자본잉여금 + 자본조정 + 기타포괄손익누계액 + 기타자본항목

        # 자본총계/부채와자본총계 추출 (음수 허용)
        자본총계_from_총계 = 총계.get('자본총계')
        # '자본총계' 검색 시 '부채와자본총계', '부채및자본총계' 제외
        자본총계_from_bs = find_bs_val(['자본총계'], year, excludes=['부채와', '부채및'])
        # 자본총계가 음수인 경우도 허용 (자본잠식)
        if 자본총계_from_총계 is not None:
            자본총계 = 자본총계_from_총계
        elif 자본총계_from_bs is not None:
            자본총계 = 자본총계_from_bs
        else:
            자본총계 = 0

        부채와자본총계 = 총계.get('부채와자본총계') or 총계.get('부채및자본총계') or find_bs_val(['부채와자본총계', '부채및자본총계'], year) or 0

        # ========== 계산 항목 ==========
        nwc = 유동자산 - 유동부채
        # 총차입금: 유동차입부채 + 장기차입금 + 사채 (리스부채는 제외)
        총차입금 = 유동차입부채 + 장기차입금 + 사채
        net_debt = 총차입금 - 현금 - 단기금융상품

        # ========== 동적 VCM 구조 생성 ==========
        # 섹션 원본 항목을 표준 카테고리로 그룹화하고 세부항목도 표시

        # 카테고리 매핑 정의 (키워드 -> 카테고리명)
        def categorize_item(item_name, section_type):
            """원본 항목명을 표준 카테고리로 분류"""
            name_norm = normalize(item_name)

            if section_type == '유동자산':
                if any(k in name_norm for k in ['현금및현금성자산', '현금']):
                    return '현금및현금성자산'
                elif any(k in name_norm for k in ['단기금융상품', '당기손익-공정가치', '당기손익공정가치',
                                                   '기타포괄손익-공정가치', '기타포괄손익공정가치', 'FVPL', 'FVOCI']):
                    return '단기투자자산'
                elif any(k in name_norm for k in ['매출채권', '미수금', '미수수익', '선급금', '선급비용',
                                                   '계약자산', '기타금융자산']):
                    return '매출채권및기타채권'
                elif any(k in name_norm for k in ['재고자산', '상품', '제품', '원재료']) and '파생' not in name_norm:
                    return '재고자산'
                elif any(k in name_norm for k in ['법인세자산', '당기법인세자산']):
                    return '당기법인세자산'
                elif any(k in name_norm for k in ['매각예정', '처분자산집단']):
                    return '매각예정자산'
                else:
                    return '기타유동자산'  # 유동자산의 기타 항목

            elif section_type == '비유동자산':
                if any(k in name_norm for k in ['유형자산']) and '무형' not in name_norm:
                    return '유형자산'
                elif any(k in name_norm for k in ['무형자산']):
                    return '무형자산'
                elif any(k in name_norm for k in ['매출채권', '기타금융자산', '보증금', '임차보증금']):
                    return '매출채권및기타채권'
                elif any(k in name_norm for k in ['장기금융상품', '당기손익-공정가치', '기타포괄손익-공정가치',
                                                   '관계기업투자', '종속기업투자', '지분법', '매도가능']):
                    return '장기투자자산'
                else:
                    return '기타비유동자산'  # 비유동자산의 기타 항목

            elif section_type == '유동부채':
                if any(k in name_norm for k in ['매입채무', '미지급금', '미지급비용', '선수금', '선수수익',
                                                 '예수금', '예수보증금', '계약부채']):
                    return '매입채무및기타채무'
                elif any(k in name_norm for k in ['단기차입금', '유동성장기', '유동성사채', '전환사채',
                                                   '상환전환우선주', '전환우선주']):
                    return '유동차입부채'
                elif any(k in name_norm for k in ['충당부채']) and '리스' not in name_norm:
                    return '단기충당부채'
                elif any(k in name_norm for k in ['매각예정부채', '매각예정비유동부채']):
                    return '매각예정부채'
                elif any(k in name_norm for k in ['법인세부채', '당기법인세부채', '미지급법인세']):
                    return '당기법인세부채'
                elif any(k in name_norm for k in ['리스부채', '파생상품부채', '기타금융부채']):
                    return '기타금융부채'
                else:
                    return '기타비금융부채'

            elif section_type == '비유동부채':
                if any(k in name_norm for k in ['매입채무', '미지급금', '미지급비용', '선수금', '선수수익',
                                                 '예수금', '예수보증금', '임대보증금', '계약부채']):
                    return '매입채무및기타채무'
                # 비유동차입부채: 장기차입금, 사채 등 (단, 유동 항목 제외)
                elif any(k in name_norm for k in ['장기차입금', '사채', '비유동차입부채']) and '유동' not in name_norm:
                    return '비유동차입부채'
                elif any(k in name_norm for k in ['충당부채', '장기충당부채']) and '리스' not in name_norm:
                    return '장기충당부채'
                elif any(k in name_norm for k in ['리스부채', '파생상품부채', '기타금융부채']):
                    return '기타금융부채'
                else:
                    return '기타비금융부채'

            return None

        # 섹션별로 항목을 카테고리로 그룹화
        def group_items_by_category(items, section_type):
            """섹션 항목들을 카테고리로 그룹화하고 합계 계산"""
            groups = {}
            for item in items:
                category = categorize_item(item['name'], section_type)
                if category:
                    if category not in groups:
                        groups[category] = {'total': 0, 'items': []}
                    groups[category]['total'] += item['value']
                    groups[category]['items'].append(item)
            return groups

        # 유동자산 그룹화
        유동자산_groups = group_items_by_category(유동자산_items, '유동자산')
        비유동자산_groups = group_items_by_category(비유동자산_items, '비유동자산')
        유동부채_groups = group_items_by_category(유동부채_items, '유동부채')
        비유동부채_groups = group_items_by_category(비유동부채_items, '비유동부채')

        # ===== 동적 VCM 생성 설정 =====
        MAX_ITEMS = 6  # 각 섹션당 최대 표시 항목 수

        # 필수 항목 정의 (항상 표시해야 하는 카테고리)
        유동자산_필수 = ['현금및현금성자산', '매출채권및기타채권']
        비유동자산_필수 = []  # 금액 기준으로만 정렬
        유동부채_필수 = ['매입채무및기타채무', '유동차입부채']
        비유동부채_필수 = ['비유동차입부채']

        def is_redundant_child(parent_name, child_name):
            """
            세부항목이 부모와 중복되는지 확인
            - 부모가 자식 이름을 포함하거나 그 반대인 경우 (장기충당부채 vs 충당부채)
            - 정확히 같은 이름도 포함 관계이므로 True 반환
            """
            parent_norm = normalize(parent_name).replace('[비유동]', '').replace('[netdebt]', '').replace('[nwc]', '')
            child_norm = normalize(child_name)

            # 이름이 서로 포함 관계인 경우 (장기충당부채 ↔ 충당부채, 같은 이름 포함)
            if parent_norm in child_norm or child_norm in parent_norm:
                return True

            # 둘 다 '기타'를 포함하는 경우, 실제로 유사한 경우만 중복 처리
            # 예: 기타비유동자산 vs 기타장기수취채권 → 다른 항목이므로 중복 아님
            if '기타' in parent_norm and '기타' in child_norm:
                # '기타' 제거 후 나머지 부분이 포함 관계인 경우만 중복
                parent_rest = parent_norm.replace('기타', '')
                child_rest = child_norm.replace('기타', '')
                if parent_rest and child_rest and (parent_rest in child_rest or child_rest in parent_rest):
                    return True

            return False

        def get_display_name(item_name, parent_name):
            """하위항목 표시명 결정 - 부모와 같은 이름이면 '(세부)' 추가"""
            if normalize(item_name) == normalize(parent_name):
                return f"{item_name}(세부)"
            return item_name

        def select_top_items(groups, required_cats, max_items, section_name):
            """
            필수 항목 + 금액 큰 순으로 상위 N개 선택, 나머지는 '기타'로 합산
            """
            # 모든 카테고리를 금액 기준 정렬
            all_cats = []
            for cat, grp in groups.items():
                if grp['total'] and grp['total'] != 0:
                    all_cats.append({
                        'name': cat,
                        'total': grp['total'],
                        'items': grp['items'],
                        'is_required': cat in required_cats
                    })

            # 금액 절댓값 기준 내림차순 정렬
            all_cats.sort(key=lambda x: abs(x['total']), reverse=True)

            # 필수 항목과 옵션 항목 분리
            required = [c for c in all_cats if c['is_required']]
            optional = [c for c in all_cats if not c['is_required']]

            # 선택할 항목 결정
            selected = []
            기타_total = 0
            기타_items = []

            # 1. 필수 항목 먼저 추가
            for cat in required:
                selected.append(cat)

            # 2. 남은 슬롯에 금액 큰 순으로 추가
            remaining_slots = max_items - len(selected)
            for i, cat in enumerate(optional):
                if i < remaining_slots:
                    selected.append(cat)
                else:
                    # 나머지는 기타로 합산
                    기타_total += cat['total']
                    기타_items.extend(cat['items'])

            # 선택된 항목들을 금액 기준으로 다시 정렬
            selected.sort(key=lambda x: abs(x['total']), reverse=True)

            return selected, 기타_total, 기타_items

        # 동적 bs_items 생성
        bs_items = []

        # ===== 유동자산 =====
        bs_items.append(('유동자산', '', 유동자산))

        selected_유동자산, 기타유동자산_total, 기타유동자산_items = select_top_items(
            유동자산_groups, 유동자산_필수, MAX_ITEMS, '유동자산'
        )

        for cat_info in selected_유동자산:
            cat = cat_info['name']
            bs_items.append((cat, '유동자산', cat_info['total']))

            # 매출채권및기타채권의 경우 본체 계산에 사용된 값들을 명시적으로 추가 (누락 방지)
            if cat == '매출채권및기타채권':
                # 명시적으로 하위항목 추가 (본체 계산에 사용된 항목들)
                if 매출채권:
                    bs_items.append(('매출채권', '매출채권및기타채권', 매출채권))
                if 미수금:
                    bs_items.append(('미수금', '매출채권및기타채권', 미수금))
                if 미수수익:
                    bs_items.append(('미수수익', '매출채권및기타채권', 미수수익))
                if 선급금:
                    bs_items.append(('선급금', '매출채권및기타채권', 선급금))
                if 선급비용:
                    bs_items.append(('선급비용', '매출채권및기타채권', 선급비용))
                if 계약자산_유동:
                    bs_items.append(('계약자산', '매출채권및기타채권', 계약자산_유동))
                if 기타금융자산_유동:
                    bs_items.append(('기타금융자산', '매출채권및기타채권', 기타금융자산_유동))
            else:
                # 다른 카테고리 처리
                all_items = cat_info['items']
                # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
                if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize(cat):
                    pass  # 툴팁 안 만듦
                elif len(all_items) >= 1:
                    # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                    valid_items = [i for i in all_items
                                   if not is_redundant_child(cat, i['name'])
                                   or normalize(i['name']) == normalize(cat)]
                    sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                    for item in sorted_items[:5]:  # 최대 5개까지
                        display_name = get_display_name(item['name'], cat)
                        bs_items.append((display_name, cat, item['value']))

        # 기타유동자산 (남은 항목 합계)
        # 중요: 기타유동자산_total은 MAX_ITEMS 초과된 모든 카테고리의 합계이므로,
        # 하위항목도 기타유동자산_items (모든 초과 카테고리의 items)를 사용해야 본체와 일치함
        if 기타유동자산_total and abs(기타유동자산_total) > 0:
            bs_items.append(('기타유동자산', '유동자산', 기타유동자산_total))
            all_items = 기타유동자산_items
            # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
            if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize('기타유동자산'):
                pass  # 툴팁 안 만듦
            elif len(all_items) >= 1:
                # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                valid_items = [i for i in all_items
                               if not is_redundant_child('기타유동자산', i['name'])
                               or normalize(i['name']) == normalize('기타유동자산')]
                sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                for item in sorted_items[:5]:
                    display_name = get_display_name(item['name'], '기타유동자산')
                    bs_items.append((display_name, '기타유동자산', item['value']))

        # 매각예정자산 (유동자산 섹션에 포함)
        매각예정자산 = find_bs_val(['매각예정비유동자산', '매각예정자산', '처분자산집단'], year) or 0
        if 매각예정자산:
            bs_items.append(('매각예정자산', '유동자산', 매각예정자산))

        # ===== 비유동자산 =====
        bs_items.append(('비유동자산', '', 비유동자산))

        # 비유동자산 카테고리명 매핑 (유동과 구분 필요한 항목)
        비유동자산_카테고리_표시명 = {
            '매출채권및기타채권': '매출채권및기타채권[비유동]',
        }

        selected_비유동자산, 기타비유동자산_total, 기타비유동자산_items = select_top_items(
            비유동자산_groups, 비유동자산_필수, MAX_ITEMS, '비유동자산'
        )

        for cat_info in selected_비유동자산:
            cat = cat_info['name']
            cat_display = 비유동자산_카테고리_표시명.get(cat, cat)
            bs_items.append((cat_display, '비유동자산', cat_info['total']))
            all_items = cat_info['items']
            # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
            if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize(cat):
                pass  # 툴팁 안 만듦
            elif len(all_items) >= 1:
                # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                valid_items = [i for i in all_items
                               if not is_redundant_child(cat_display, i['name'])
                               or normalize(i['name']) == normalize(cat)]
                sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                for item in sorted_items[:5]:  # 최대 5개까지
                    display_name = get_display_name(item['name'], cat)
                    bs_items.append((display_name, cat_display, item['value']))

        if 기타비유동자산_total and abs(기타비유동자산_total) > 0:
            bs_items.append(('기타비유동자산', '비유동자산', 기타비유동자산_total))
            all_items = 기타비유동자산_items
            # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
            if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize('기타비유동자산'):
                pass  # 툴팁 안 만듦
            elif len(all_items) >= 1:
                # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                valid_items = [i for i in all_items
                               if not is_redundant_child('기타비유동자산', i['name'])
                               or normalize(i['name']) == normalize('기타비유동자산')]
                sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                for item in sorted_items[:5]:
                    display_name = get_display_name(item['name'], '기타비유동자산')
                    bs_items.append((display_name, '기타비유동자산', item['value']))

        bs_items.append(('자산총계', '', 자산총계))

        # ===== 유동부채 =====
        bs_items.append(('유동부채', '', 유동부채))

        selected_유동부채, 기타유동부채_total, 기타유동부채_items = select_top_items(
            유동부채_groups, 유동부채_필수, MAX_ITEMS, '유동부채'
        )

        for cat_info in selected_유동부채:
            cat = cat_info['name']
            bs_items.append((cat, '유동부채', cat_info['total']))
            all_items = cat_info['items']
            # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
            if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize(cat):
                pass  # 툴팁 안 만듦
            elif len(all_items) >= 1:
                # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                valid_items = [i for i in all_items
                               if not is_redundant_child(cat, i['name'])
                               or normalize(i['name']) == normalize(cat)]
                sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                for item in sorted_items[:5]:  # 최대 5개까지
                    display_name = get_display_name(item['name'], cat)
                    bs_items.append((display_name, cat, item['value']))

        if 기타유동부채_total and abs(기타유동부채_total) > 0:
            bs_items.append(('기타유동부채', '유동부채', 기타유동부채_total))
            all_items = 기타유동부채_items
            # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
            if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize('기타유동부채'):
                pass  # 툴팁 안 만듦
            elif len(all_items) >= 1:
                # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                valid_items = [i for i in all_items
                               if not is_redundant_child('기타유동부채', i['name'])
                               or normalize(i['name']) == normalize('기타유동부채')]
                sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                for item in sorted_items[:5]:
                    display_name = get_display_name(item['name'], '기타유동부채')
                    bs_items.append((display_name, '기타유동부채', item['value']))

        # 매각예정부채 별도 추출 (섹션 외부에 있을 수 있음)
        매각예정부채 = find_bs_val(['매각예정비유동부채', '매각예정부채'], year) or 0
        if 매각예정부채:
            bs_items.append(('매각예정부채', '유동부채', 매각예정부채))

        # ===== 비유동부채 =====
        bs_items.append(('비유동부채', '', 비유동부채))

        # 비유동부채 카테고리명 매핑 (유동과 구분 필요한 항목)
        비유동부채_카테고리_표시명 = {
            '매입채무및기타채무': '매입채무및기타채무[비유동]',
            '기타금융부채': '기타금융부채[비유동]',
            '기타비금융부채': '기타비금융부채[비유동]',
        }

        selected_비유동부채, 기타비유동부채_total, 기타비유동부채_items = select_top_items(
            비유동부채_groups, 비유동부채_필수, MAX_ITEMS, '비유동부채'
        )

        for cat_info in selected_비유동부채:
            cat = cat_info['name']
            cat_display = 비유동부채_카테고리_표시명.get(cat, cat)
            bs_items.append((cat_display, '비유동부채', cat_info['total']))
            all_items = cat_info['items']
            # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
            if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize(cat):
                pass  # 툴팁 안 만듦
            elif len(all_items) >= 1:
                # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                valid_items = [i for i in all_items
                               if not is_redundant_child(cat_display, i['name'])
                               or normalize(i['name']) == normalize(cat)]
                sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                for item in sorted_items[:5]:  # 최대 5개까지
                    display_name = get_display_name(item['name'], cat)
                    bs_items.append((display_name, cat_display, item['value']))

        if 기타비유동부채_total and abs(기타비유동부채_total) > 0:
            bs_items.append(('기타비유동부채', '비유동부채', 기타비유동부채_total))
            all_items = 기타비유동부채_items
            # 세부가 1개뿐이고 카테고리명과 같으면 → 툴팁 불필요
            if len(all_items) == 1 and normalize(all_items[0]['name']) == normalize('기타비유동부채'):
                pass  # 툴팁 안 만듦
            elif len(all_items) >= 1:
                # 포함 관계만 제외, 같은 이름은 display_name으로 구분하여 포함
                valid_items = [i for i in all_items
                               if not is_redundant_child('기타비유동부채', i['name'])
                               or normalize(i['name']) == normalize('기타비유동부채')]
                sorted_items = sorted(valid_items, key=lambda x: abs(x['value']), reverse=True)
                for item in sorted_items[:5]:
                    display_name = get_display_name(item['name'], '기타비유동부채')
                    bs_items.append((display_name, '기타비유동부채', item['value']))

        bs_items.append(('부채총계', '', 부채총계))

        # ===== 자본 =====
        bs_items.append(('자본금', '', 자본금))
        bs_items.append(('이익잉여금', '', 이익잉여금))
        bs_items.append(('기타자본구성요소', '', 기타자본))
        # 자본 세부항목 - 본체 계산에 사용된 값들을 명시적으로 추가 (누락 방지)
        if 자본잉여금:
            bs_items.append(('자본잉여금', '기타자본구성요소', 자본잉여금))
        if 자본조정:
            bs_items.append(('자본조정', '기타자본구성요소', 자본조정))
        if 기타포괄손익누계액:
            bs_items.append(('기타포괄손익누계액', '기타자본구성요소', 기타포괄손익누계액))
        if 기타자본항목:
            bs_items.append(('기타자본항목', '기타자본구성요소', 기타자본항목))

        bs_items.append(('자본총계', '', 자본총계))
        bs_items.append(('부채와자본총계', '', 부채와자본총계))

        # ===== 계산 항목 =====
        bs_items.append(('NWC', '', nwc))
        bs_items.append(('유동자산 [NWC]', 'NWC', 유동자산))
        bs_items.append(('유동부채 [NWC]', 'NWC', 유동부채))

        # Net Debt 계산: 유동차입부채 + 비유동차입부채 - 현금 - 단기투자자산
        유동차입부채_합계 = 유동부채_groups.get('유동차입부채', {}).get('total', 0)
        비유동차입부채_합계 = 비유동부채_groups.get('비유동차입부채', {}).get('total', 0)
        현금_합계 = 유동자산_groups.get('현금및현금성자산', {}).get('total', 0)
        단기투자자산_합계 = 유동자산_groups.get('단기투자자산', {}).get('total', 0)
        net_debt = 유동차입부채_합계 + 비유동차입부채_합계 - 현금_합계 - 단기투자자산_합계

        bs_items.append(('Net Debt', '', net_debt))
        bs_items.append(('유동차입부채 [NetDebt]', 'Net Debt', 유동차입부채_합계))
        bs_items.append(('비유동차입부채 [NetDebt]', 'Net Debt', 비유동차입부채_합계))
        bs_items.append(('현금및현금성자산 [NetDebt]', 'Net Debt', -현금_합계))
        bs_items.append(('단기투자자산 [NetDebt]', 'Net Debt', -단기투자자산_합계))

        # 항목 순서와 값을 저장
        all_bs_items_by_year[year_str] = bs_items

    # ========== 마스터 순서 결정 및 값 병합 ==========
    # 모든 연도의 항목을 올바른 순서로 병합
    master_order = []  # [(항목명, 부모)]
    master_order_set = set()

    # 각 연도의 bs_items를 순회하며 마스터 순서 구축
    # 최신 연도부터 처리하여 최신 데이터의 순서를 우선
    for year_str in sorted(all_bs_items_by_year.keys(), reverse=True):
        bs_items = all_bs_items_by_year[year_str]

        # 첫 번째(최신) 연도는 순서 그대로 사용
        if not master_order:
            for item_name, parent, val in bs_items:
                master_order.append((item_name, parent))
                master_order_set.add(item_name)
        else:
            # 이후 연도: 새 항목만 적절한 위치에 삽입
            insert_idx = 0
            for item_name, parent, val in bs_items:
                if item_name not in master_order_set:
                    # 새 항목: 올바른 위치에 삽입
                    master_order.insert(insert_idx, (item_name, parent))
                    master_order_set.add(item_name)
                else:
                    # 기존 항목: 현재 위치 찾기
                    for i, (name, _) in enumerate(master_order):
                        if name == item_name:
                            insert_idx = i
                            break
                insert_idx += 1

    # 마스터 순서대로 bs_rows 생성
    bs_rows = []
    item_to_row = {}
    for item_name, parent in master_order:
        row = {'항목': item_name, '부모': parent}
        bs_rows.append(row)
        item_to_row[item_name] = row

    # 각 연도의 값 채우기
    for year_str, bs_items in all_bs_items_by_year.items():
        for item_name, parent, val in bs_items:
            if item_name in item_to_row:
                item_to_row[item_name][year_str] = round(val) if val is not None and val != 0 else None

    # 빈 행 필터링 (모든 연도에서 값이 없는 행 제거)
    year_cols = sorted(all_bs_items_by_year.keys())
    filtered_bs_rows = []
    for row in bs_rows:
        has_value = any(row.get(y) is not None and row.get(y) != 0 for y in year_cols)
        if has_value:
            filtered_bs_rows.append(row)
        else:
            print(f"[VCM] 빈 행 제거: {row.get('항목')}")

    # 손익계산서 빈 행 필터링 (모든 연도에서 값이 없는 행 제거)
    filtered_is_rows = []
    for row in rows:
        has_value = any(row.get(y) is not None and row.get(y) != 0 for y in year_cols)
        if has_value:
            filtered_is_rows.append(row)
        else:
            print(f"[VCM] IS 빈 행 제거: {row.get('항목')}")

    # 재무상태표 + 손익계산서 순서로 결합
    all_rows = filtered_bs_rows + filtered_is_rows

    # ========== 타입 컬럼 추가 ==========
    # 하위항목 표시 대상 섹션 (BS + IS)
    sections_with_subitems = [
        '유동자산', '비유동자산', '유동부채', '비유동부채', '자본',
        '매출', '매출원가', '판매비와관리비', '영업외수익', '영업외비용'
    ]

    # 타입 결정 함수
    def get_item_type(item_name, item_parent=''):
        name = item_name.strip()
        parent = (item_parent or '').strip()

        # highlight 타입 (강조 항목)
        highlight_items = ['영업이익', '당기순이익', 'EBITDA']
        if name in highlight_items:
            return 'highlight'

        # total 타입 (합계 항목)
        total_items = ['매출총이익', '자산총계', '부채총계', '자본총계', '부채와자본총계',
                       '법인세비용차감전이익', '법인세비용차감전순이익', 'NWC', 'Net Debt']
        if name in total_items:
            return 'total'

        # percent 타입
        if name == '% of Sales' or name.startswith('%'):
            return 'percent'

        # category 타입 (대분류)
        category_items = ['유동자산', '비유동자산', '유동부채', '비유동부채', '자본',
                          '매출', '매출원가', '판매비와관리비', '영업외수익', '영업외비용',
                          '유동자산 [NWC]', '유동부채 [NWC]',
                          '유동차입부채 [NetDebt]', '비유동차입부채 [NetDebt]',
                          '현금및현금성자산 [NetDebt]', '단기투자자산 [NetDebt]']
        if name in category_items:
            return 'category'

        # subitem 타입 (부모가 있는 하위항목)
        if parent in sections_with_subitems:
            return 'subitem'

        # subitem 타입 (들여쓰기 있는 항목 - IS fallback)
        if item_name.startswith('  '):
            return 'subitem'

        # BS 하위항목 판별: 로마숫자로 시작하지 않고, 대분류/합계/하이라이트가 아닌 항목
        # IS 항목은 로마숫자로 시작하거나 들여쓰기가 있음
        is_section_prefixes = ['Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ', 'Ⅶ', 'Ⅷ', 'Ⅸ', 'Ⅹ']
        is_is_item = any(name.startswith(prefix) for prefix in is_section_prefixes)

        # IS 대분류가 아닌 BS 항목은 subitem으로 처리
        if not is_is_item:
            return 'subitem'

        # 기본값 (IS 대분류)
        return 'item'

    # 각 행에 타입 추가
    for row in all_rows:
        row['타입'] = get_item_type(row.get('항목', ''), row.get('부모', ''))

    # ========== Financials 시트 생성 (단위: 백만원, 포맷팅 완료) ==========
    # 규칙:
    # - 부모 없는 항목 → 표시
    # - BS/IS 섹션의 하위항목 → 표시 (들여쓰기)
    # - NWC/NetDebt 하위항목 → 제외 (툴팁용)
    display_rows = []
    unit_divisor = 1000000  # 원 → 백만원 (1백만 = 1,000,000)

    for row in all_rows:
        item_name = row.get('항목', '')
        item_type = row.get('타입', 'item')
        item_parent = row.get('부모', '')

        # NWC/NetDebt 하위항목 제외 (툴팁용)
        if '[NWC]' in item_name or '[NetDebt]' in item_name:
            continue

        # 부모 없는 항목 → 표시
        if not item_parent or not str(item_parent).strip():
            pass  # 표시
        # BS/IS 섹션의 하위항목 → 표시
        elif str(item_parent).strip() in sections_with_subitems:
            pass  # 표시
        # 그 외 (기타XXX 하위 등) → 제외
        else:
            continue

        # 빈 행 필터링: 모든 연도에 값이 없으면 스킵
        has_any_value = any(row.get(col) is not None and row.get(col) != 0 for col in year_cols)
        if not has_any_value:
            continue

        display_row = {'항목': item_name}

        for col in year_cols:
            val = row.get(col)
            if val is not None and val != 0:
                if item_type == 'percent':
                    # 퍼센트는 소수점 형태를 % 형식으로 변환 (0.127 → "12.7%", 1.5 → "150.0%")
                    if isinstance(val, (int, float)):
                        # 소수점 값이면 % 형식으로 변환 (절대값이 10 이하면 소수점 형태로 간주)
                        if abs(val) <= 10:
                            display_row[col] = f"{val * 100:.1f}%"
                        else:
                            # 이미 100을 곱한 값이면 그대로 % 추가
                            display_row[col] = f"{val:.1f}%"
                    elif isinstance(val, str) and '%' in val:
                        # 이미 % 형식 문자열이면 그대로 사용
                        display_row[col] = val
                    else:
                        display_row[col] = val
                else:
                    # 숫자는 백만원 단위로 변환하고 포맷팅 (정수로 표시)
                    converted = val / unit_divisor
                    # 정수로 반올림, 천 단위 콤마 (0이면 빈값 처리)
                    rounded = round(converted)
                    if rounded == 0:
                        display_row[col] = ''  # 0이면 빈값
                    else:
                        display_row[col] = f"{rounded:,}"  # 정수로 표시
            else:
                display_row[col] = ''

        # 2차 필터: 백만원 변환 후 모든 연도가 빈 칸이면 제거
        has_display_value = any(display_row.get(col) not in (None, '', 0) for col in year_cols)
        if has_display_value:
            display_rows.append(display_row)
        else:
            print(f"[VCM] Financials 빈 행 제거 (변환 후): {item_name}")

    # 컬럼 순서 정리: 항목, 타입, 부모, FY연도들...
    vcm_df = pd.DataFrame(all_rows)
    cols_order = ['항목', '타입', '부모'] + [c for c in vcm_df.columns if c not in ['항목', '타입', '부모']]
    vcm_df = vcm_df[[c for c in cols_order if c in vcm_df.columns]]

    # Financials(표시용) DataFrame
    display_df = pd.DataFrame(display_rows)

    return vcm_df, display_df


def normalize_xbrl_columns(df):
    """XBRL 형식의 복잡한 튜플 컬럼을 간단한 문자열로 변환"""
    if df is None or df.empty:
        return df
    
    # 튜플 컬럼이 있는지 확인
    has_tuple_cols = any(isinstance(c, tuple) for c in df.columns)
    if not has_tuple_cols:
        return df  # 이미 정규화된 경우
    
    print(f"[XBRL] 컬럼 정규화 시작: {len(df.columns)}개 컬럼")
    print(f"[XBRL] 원본 컬럼 타입들: {[(col, type(col).__name__) for col in df.columns if 'FY' in str(col) or '2025' in str(col)]}")
    
    new_columns = []
    for col in df.columns:
        if isinstance(col, tuple):
            # 튜플 컬럼 처리
            first_part = col[0] if len(col) > 0 else ''
            second_part = col[1] if len(col) > 1 else ''
            
            # 날짜 형식 컬럼 ('20240101-20241231', ...) 또는 ('20241231', ...)
            if isinstance(first_part, str):
                # FY로 시작하는 컬럼 (병합된 분기 데이터)
                if first_part.startswith('FY'):
                    new_columns.append(first_part)
                # '20240101-20241231' 형식 (17자)
                elif '-' in first_part and len(first_part) == 17:
                    year = first_part[:4]
                    new_columns.append(f'FY{year}')
                # '20241231' 형식 (8자, 숫자만)
                elif len(first_part) == 8 and first_part.isdigit():
                    year = first_part[:4]
                    new_columns.append(f'FY{year}')
                # label_ko, concept_id 등의 메타데이터 컬럼
                elif isinstance(second_part, str):
                    if 'label_ko' in second_part:
                        new_columns.append('계정과목')
                    elif 'concept_id' in second_part:
                        new_columns.append('개념ID')
                    elif 'label_en' in second_part:
                        new_columns.append('계정과목(영문)')
                    elif 'class0' in second_part:
                        new_columns.append('분류1')
                    elif 'class1' in second_part:
                        new_columns.append('분류2')
                    elif 'class2' in second_part:
                        new_columns.append('분류3')
                    elif 'class3' in second_part:
                        new_columns.append('분류4')
                    else:
                        new_columns.append(str(second_part))
                else:
                    new_columns.append(str(first_part)[:50])
            else:
                new_columns.append(str(first_part)[:50])
        else:
            new_columns.append(col)
    
    df_copy = df.copy()
    df_copy.columns = new_columns
    
    # FY 컬럼을 과거→현재 순서로 정렬 (FY2017, FY2018, ..., FY2024, FY2025 3Q)
    non_fy_cols = [c for c in df_copy.columns if not str(c).startswith('FY')]
    fy_cols = [c for c in df_copy.columns if str(c).startswith('FY')]
    
    # FY 컬럼 정렬 (연도 기준 오름차순, 분기는 뒤에)
    def fy_sort_key(col):
        col_str = str(col)
        # FY2025 3Q -> (2025, 3), FY2024 -> (2024, 0)
        if ' ' in col_str:
            parts = col_str.replace('FY', '').split(' ')
            year = int(parts[0]) if parts[0].isdigit() else 9999
            quarter = int(parts[1].replace('Q', '')) if len(parts) > 1 else 0
            return (year, quarter)
        else:
            year = int(col_str.replace('FY', '')) if col_str.replace('FY', '').isdigit() else 9999
            return (year, 0)
    
    fy_cols_sorted = sorted(fy_cols, key=fy_sort_key)
    
    print(f"[XBRL] 정규화 후 FY 컬럼: {fy_cols_sorted}")
    
    # 메타데이터 컬럼 + 정렬된 FY 컬럼 순서로 재배열
    new_order = non_fy_cols + fy_cols_sorted
    df_copy = df_copy[new_order]
    
    return df_copy


def cleanup_old_excel_files(output_dir='output', days=7):
    """일주일 이상 지난 엑셀 파일 삭제"""
    import glob
    import time

    try:
        # output 디렉토리의 모든 .xlsx 파일 검색
        pattern = os.path.join(output_dir, '*.xlsx')
        excel_files = glob.glob(pattern)

        # 현재 시간
        current_time = time.time()
        # 7일을 초 단위로 변환
        cutoff_time = current_time - (days * 24 * 60 * 60)

        deleted_count = 0
        for file_path in excel_files:
            try:
                # 파일 생성 시간 확인
                file_mtime = os.path.getmtime(file_path)

                # 7일 이상 지난 파일 삭제
                if file_mtime < cutoff_time:
                    os.remove(file_path)
                    deleted_count += 1
                    print(f"[Cleanup] 오래된 파일 삭제: {os.path.basename(file_path)}")
            except Exception as e:
                print(f"[Cleanup] 파일 삭제 실패: {file_path}, 오류: {e}")

        if deleted_count > 0:
            print(f"[Cleanup] 총 {deleted_count}개 파일 삭제 완료")
        else:
            print(f"[Cleanup] 삭제할 오래된 파일 없음 (총 {len(excel_files)}개 파일 확인)")

    except Exception as e:
        print(f"[Cleanup] 파일 정리 실패: {e}")


def save_to_excel(fs_data, filepath: str, company_info: Optional[Dict[str, Any]] = None):
    """재무제표를 엑셀 파일로 저장 (주석 테이블 포함)"""
    # 엑셀 저장 전 오래된 파일 삭제
    cleanup_old_excel_files()

    sheet_names = {
        'bs': '재무상태표',
        'is': '손익계산서',
        'cis': '포괄손익계산서',
        'cf': '현금흐름표'
    }

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # 기업개황정보 시트 (가장 먼저)
        if company_info:
            try:
                # 업종명 조회 (KSIC 코드 → 업종명)
                induty_code = company_info.get('induty_code', '')
                induty_name = KSIC_CODES.get(induty_code, '') if induty_code and KSIC_CODES else ''

                # 웹 프론트와 동일한 순서/항목
                info_rows = [
                    ['항목', '내용'],
                    ['회사명', company_info.get('corp_name', '')],
                    ['영문명', company_info.get('corp_name_eng', '')],
                    ['대표자', company_info.get('ceo_nm', '')],
                    ['시장구분', company_info.get('market_name', '')],
                    ['종목코드', company_info.get('stock_code', '')],
                    ['법인번호', company_info.get('jurir_no', '')],
                    ['사업자번호', company_info.get('bizr_no', '')],
                    ['업종코드', induty_code],
                    ['업종명', induty_name],
                    ['설립일', company_info.get('est_dt_formatted', '')],
                    ['결산월', company_info.get('acc_mt_formatted', '')],
                    ['주소', company_info.get('adres', '')],
                    ['홈페이지', company_info.get('hm_url', '')],
                ]
                info_df = pd.DataFrame(info_rows[1:], columns=info_rows[0])
                info_df.to_excel(writer, sheet_name='기업개황', index=False)

                # 기업개황 시트 정렬 스타일 적용
                from openpyxl.styles import Alignment, PatternFill, Font
                info_ws = writer.book['기업개황']

                # 헤더 행 스타일 적용 (#131313 배경, 흰색 글자)
                header_fill = PatternFill(start_color='131313', end_color='131313', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                for col in range(1, info_ws.max_column + 1):
                    cell = info_ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font

                # 법인번호, 사업자번호, 업종코드는 텍스트로 포맷 (콤마 방지)
                text_format_rows = ['법인번호', '사업자번호', '업종코드']

                for row in range(2, info_ws.max_row + 1):  # 헤더 제외
                    cell_a = info_ws.cell(row=row, column=1)
                    cell_b = info_ws.cell(row=row, column=2)
                    cell_a.alignment = Alignment(horizontal='left', vertical='center')  # A열 좌측
                    cell_b.alignment = Alignment(horizontal='right', vertical='center')  # B열 우측

                    # 텍스트 포맷 적용 (숫자 콤마 방지)
                    if cell_a.value in text_format_rows:
                        cell_b.number_format = '@'  # 텍스트 포맷

                print(f"[Excel] 기업개황 시트 저장 완료")
            except Exception as e:
                print(f"[Excel] 기업개황 시트 저장 실패: {e}")
        # 헤더 스타일 정의 (공통)
        from openpyxl.styles import PatternFill, Font, Alignment
        common_header_fill = PatternFill(start_color='131313', end_color='131313', fill_type='solid')
        common_header_font = Font(color='FFFFFF', bold=True)

        def apply_header_style(ws):
            """시트의 헤더 행(1행)에 #131313 배경, 흰색 글자 적용"""
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = common_header_fill
                cell.font = common_header_font

        for key, name in sheet_names.items():
            try:
                df = fs_data[key]
                if df is not None and not df.empty:
                    # XBRL 형식인 경우 컬럼 정규화
                    df_normalized = normalize_xbrl_columns(df)
                    df_normalized.to_excel(writer, sheet_name=name, index=False)
                    # 헤더 스타일 적용
                    apply_header_style(writer.book[name])
            except Exception:
                pass
        
        # 주석 테이블들 저장 (각 재무제표별)
        notes = fs_data.get('notes', {})
        if notes:
            # 주석 시트 헤더 스타일 적용 함수
            def apply_note_header_style(ws):
                """주석 시트의 헤더 행(1행)에 #131313 배경, 흰색 글자, B열 이후 중앙정렬"""
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = common_header_fill
                    cell.font = common_header_font
                    if col >= 2:  # B열 이후 중앙정렬
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # 손익계산서 관련 주석들
            is_notes = notes.get('is_notes', [])
            for idx, note in enumerate(is_notes[:5]):  # 최대 5개
                try:
                    note_df = note['df']
                    if note_df is not None and not note_df.empty:
                        note_df_normalized = normalize_xbrl_columns(note_df)
                        sheet_name = f'손익주석{idx+1}'
                        note_df_normalized.to_excel(writer, sheet_name=sheet_name, index=False)
                        apply_note_header_style(writer.book[sheet_name])
                        print(f"[Excel] 주석 저장: {sheet_name}")
                except Exception as e:
                    print(f"[Excel] 손익주석 저장 실패: {e}")
            
            # 재무상태표 관련 주석들
            bs_notes = notes.get('bs_notes', [])
            for idx, note in enumerate(bs_notes[:5]):  # 최대 5개
                try:
                    note_df = note['df']
                    if note_df is not None and not note_df.empty:
                        note_df_normalized = normalize_xbrl_columns(note_df)
                        sheet_name = f'재무주석{idx+1}'
                        note_df_normalized.to_excel(writer, sheet_name=sheet_name, index=False)
                        apply_note_header_style(writer.book[sheet_name])
                        print(f"[Excel] 주석 저장: {sheet_name}")
                except Exception as e:
                    print(f"[Excel] 재무주석 저장 실패: {e}")

            # 현금흐름표 관련 주석들
            cf_notes = notes.get('cf_notes', [])
            for idx, note in enumerate(cf_notes[:3]):  # 최대 3개
                try:
                    note_df = note['df']
                    if note_df is not None and not note_df.empty:
                        note_df_normalized = normalize_xbrl_columns(note_df)
                        sheet_name = f'현금주석{idx+1}'
                        note_df_normalized.to_excel(writer, sheet_name=sheet_name, index=False)
                        apply_note_header_style(writer.book[sheet_name])
                        print(f"[Excel] 주석 저장: {sheet_name}")
                except Exception as e:
                    print(f"[Excel] 현금주석 저장 실패: {e}")
        
        # VCM 전용 포맷 시트 추가
        try:
            print(f"[VCM] create_vcm_format 호출 시작...")
            vcm_result = create_vcm_format(fs_data, filepath)

            # 튜플 반환 (vcm_df, display_df) 처리
            if isinstance(vcm_result, tuple):
                vcm_df, display_df = vcm_result
            else:
                vcm_df = vcm_result
                display_df = None

            print(f"[VCM] create_vcm_format 완료: {len(vcm_df) if vcm_df is not None else 0}개 항목")

            # preview_data 직접 저장 (엑셀 형식과 무관하게 원본 데이터 유지)
            vcm_preview_data = None
            display_preview_data = None
            if vcm_df is not None and not vcm_df.empty:
                vcm_preview_data = safe_dataframe_to_json(vcm_df)
                print(f"[VCM] vcm_preview_data 생성: {len(vcm_preview_data)}개 행")
            if display_df is not None and not display_df.empty:
                display_preview_data = safe_dataframe_to_json(display_df)
                print(f"[VCM] display_preview_data 생성: {len(display_preview_data)}개 행")

            if vcm_df is not None and not vcm_df.empty:
                print(f"[VCM] Frontdata 저장 중... ({len(vcm_df)}행)")
                vcm_df.to_excel(writer, sheet_name='Frontdata', index=False)

                # Frontdata 헤더 스타일 적용
                from openpyxl.styles import PatternFill, Font, Alignment
                fd_ws = writer.book['Frontdata']
                fd_header_fill = PatternFill(start_color='131313', end_color='131313', fill_type='solid')
                fd_header_font = Font(bold=True, color='FFFFFF')
                fd_header_align = Alignment(horizontal='center', vertical='center')
                for col in range(1, fd_ws.max_column + 1):
                    cell = fd_ws.cell(row=1, column=col)
                    cell.fill = fd_header_fill
                    cell.font = fd_header_font
                    cell.alignment = fd_header_align
                fd_ws.row_dimensions[1].height = 30  # 헤더 행 높이 2배

                # Frontdata 하위항목 들여쓰기 적용 (타입이 subitem인 경우)
                # 컬럼 인덱스 찾기
                col_indices = {}
                for col in range(1, fd_ws.max_column + 1):
                    header_val = fd_ws.cell(row=1, column=col).value
                    if header_val:
                        col_indices[header_val] = col

                type_col = col_indices.get('타입')
                item_col = col_indices.get('항목')

                if type_col and item_col:
                    for row in range(2, fd_ws.max_row + 1):
                        item_type = fd_ws.cell(row=row, column=type_col).value
                        if item_type in ['subitem', 'percent']:
                            # 항목 셀에 들여쓰기 적용
                            item_cell = fd_ws.cell(row=row, column=item_col)
                            item_cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)

                print(f"[VCM] Frontdata 시트 저장 완료")

            if display_df is not None and not display_df.empty:
                print(f"[VCM] Financials 저장 중... ({len(display_df)}행)")

                # 타입 정보는 vcm_df에서 가져옴
                item_names = display_df['항목'].tolist() if '항목' in display_df.columns else []

                # vcm_df에서 타입 정보 매핑
                type_map = {}
                if vcm_df is not None and '항목' in vcm_df.columns and '타입' in vcm_df.columns:
                    for _, row in vcm_df.iterrows():
                        type_map[row['항목']] = row.get('타입', 'item')
                type_info = [type_map.get(name, 'item') for name in item_names]

                display_df_clean = display_df.copy()

                # 재무상태표/손익계산서 분리 (매출 기준)
                is_start_idx = None
                for idx, name in enumerate(item_names):
                    if name and name.strip() == '매출':
                        is_start_idx = idx
                        break

                if is_start_idx is None:
                    is_start_idx = len(item_names)  # 손익계산서 없으면 전체가 재무상태표

                bs_rows = display_df_clean.iloc[:is_start_idx].reset_index(drop=True)
                is_rows = display_df_clean.iloc[is_start_idx:].reset_index(drop=True)
                bs_types = type_info[:is_start_idx]
                is_types = type_info[is_start_idx:]

                print(f"[VCM] 재무상태표: {len(bs_rows)}행, 손익계산서: {len(is_rows)}행")

                # 빈 시트 생성 후 좌우 배치
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

                # 시트 생성
                if 'Financials' in writer.book.sheetnames:
                    del writer.book['Financials']
                ws = writer.book.create_sheet('Financials')

                # 스타일 정의
                header_fill = PatternFill(start_color='131313', end_color='131313', fill_type='solid')  # 섹션 헤더 (진한 검정)
                th_fill = PatternFill(start_color='6b7280', end_color='6b7280', fill_type='solid')  # 컬럼 헤더 (연한 회색)
                category_fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
                total_fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
                highlight_fill = PatternFill(start_color='fef2f2', end_color='fef2f2', fill_type='solid')

                white_font = Font(color='FFFFFF', bold=True)
                bold_font = Font(bold=True)
                th_font = Font(color='FFFFFF', bold=True)  # 컬럼 헤더용 흰색 글자
                total_font = Font(bold=True, color='92400e')
                highlight_font = Font(bold=True, color='991b1b')
                gray_font = Font(color='6b7280')
                negative_font = Font(color='dc2626')  # 음수용 빨간색 폰트
                negative_bold_font = Font(bold=True, color='dc2626')  # 음수용 굵은 빨간색 폰트
                thin_border = Border(bottom=Side(style='thin', color='e5e7eb'))

                # 컬럼 정보
                data_cols = list(display_df_clean.columns)  # (단위: 백만원), FY2020, FY2021, ...
                num_cols = len(data_cols)
                is_start_col = num_cols + 2  # 재무상태표 컬럼 + 빈 컬럼 1개 후 시작

                def apply_cell_style(cell, item_type, is_first_col=False, is_header=False, is_section_header=False):
                    """셀 스타일 적용"""
                    if is_section_header:
                        cell.fill = header_fill
                        cell.font = white_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif is_header:
                        cell.fill = th_fill
                        cell.font = th_font  # 흰색 글자
                        cell.alignment = Alignment(horizontal='left' if is_first_col else 'right', vertical='center')
                    else:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left' if is_first_col else 'right', vertical='center')

                        if item_type == 'category':
                            cell.fill = category_fill
                            cell.font = bold_font
                        elif item_type == 'total':
                            cell.fill = total_fill
                            cell.font = total_font
                        elif item_type == 'highlight':
                            cell.fill = highlight_fill
                            cell.font = highlight_font
                        elif item_type in ['subitem', 'percent']:
                            cell.font = gray_font
                            if is_first_col:
                                cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)

                def write_table(start_col, section_name, df_rows, types_list):
                    """테이블 작성 (섹션 헤더 + 컬럼 헤더 + 데이터)"""
                    # 행1: 섹션 헤더
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + num_cols - 1)
                    cell = ws.cell(row=1, column=start_col, value=section_name)
                    apply_cell_style(cell, None, is_section_header=True)

                    # 행2: 컬럼 헤더
                    for col_idx, col_name in enumerate(data_cols):
                        # 첫 번째 컬럼(항목)은 "(단위: 백만원)"으로 표시
                        display_col_name = '(단위: 백만원)' if col_idx == 0 else col_name
                        cell = ws.cell(row=2, column=start_col + col_idx, value=display_col_name)
                        apply_cell_style(cell, None, is_first_col=(col_idx == 0), is_header=True)

                    # 행3~: 데이터
                    for row_idx, (_, row_data) in enumerate(df_rows.iterrows()):
                        item_type = types_list[row_idx] if row_idx < len(types_list) else 'item'
                        for col_idx, col_name in enumerate(data_cols):
                            cell_value = row_data[col_name]
                            cell = ws.cell(row=3 + row_idx, column=start_col + col_idx, value=cell_value)
                            apply_cell_style(cell, item_type, is_first_col=(col_idx == 0))

                            # 숫자 컬럼에서 음수인 경우 빨간색 글자 적용
                            if col_idx > 0 and cell_value is not None:
                                try:
                                    # 문자열에서 숫자 추출 (콤마 제거)
                                    num_val = float(str(cell_value).replace(',', '')) if isinstance(cell_value, str) else float(cell_value)
                                    if num_val < 0:
                                        # 기존 스타일 유지하면서 글자색만 빨간색으로
                                        if item_type in ['total', 'highlight', 'category']:
                                            cell.font = negative_bold_font
                                        else:
                                            cell.font = negative_font
                                except (ValueError, TypeError):
                                    pass

                # 재무상태표 작성 (A열부터)
                write_table(1, '재무상태표', bs_rows, bs_types)

                # 손익계산서 작성 (빈 컬럼 1개 후)
                if len(is_rows) > 0:
                    write_table(is_start_col, '손익계산서', is_rows, is_types)

                # 헤더 행 높이 설정 (섹션 헤더만 2배)
                ws.row_dimensions[1].height = 30  # 섹션 헤더만 높게
                # 행2 (컬럼 헤더)는 기본 높이 유지

                print(f"[VCM] Financials 시트 저장 완료 (좌우 배치)")

        except Exception as e:
            import traceback
            print(f"[VCM] VCM 포맷 생성 실패: {e}")
            print(f"[VCM] 상세 에러:\n{traceback.format_exc()}")

        # 모든 시트의 컬럼 너비 자동 조절
        try:
            from openpyxl.utils import get_column_letter
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Financials 시트는 고정 너비 사용 (좌우 배치)
                if sheet_name == 'Financials':
                    # 재무상태표: A~F열 (항목명 20, FY컬럼 10)
                    worksheet.column_dimensions['A'].width = 20
                    worksheet.column_dimensions['B'].width = 10
                    worksheet.column_dimensions['C'].width = 10
                    worksheet.column_dimensions['D'].width = 10
                    worksheet.column_dimensions['E'].width = 10
                    worksheet.column_dimensions['F'].width = 10
                    # 빈 컬럼: G열
                    worksheet.column_dimensions['G'].width = 3
                    # 손익계산서: H~M열 (항목명 20, FY컬럼 10)
                    worksheet.column_dimensions['H'].width = 20
                    worksheet.column_dimensions['I'].width = 10
                    worksheet.column_dimensions['J'].width = 10
                    worksheet.column_dimensions['K'].width = 10
                    worksheet.column_dimensions['L'].width = 10
                    worksheet.column_dimensions['M'].width = 10
                    continue

                for column_cells in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column_cells[0].column)

                    for cell in column_cells:
                        try:
                            if cell.value:
                                # 한글은 약 2배 너비 차지
                                cell_length = 0
                                cell_str = str(cell.value)
                                for char in cell_str:
                                    if ord(char) > 127:  # 한글/특수문자
                                        cell_length += 2
                                    else:
                                        cell_length += 1
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    # 최소 너비 8, 최대 너비 60
                    adjusted_width = min(max(max_length + 2, 8), 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            print(f"[Excel] 컬럼 너비 자동 조절 완료")

            # 모든 시트의 헤더 행 높이 설정 (2배 = 30pt)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                worksheet.row_dimensions[1].height = 30
            print(f"[Excel] 헤더 행 높이 설정 완료")
        except Exception as e:
            print(f"[Excel] 컬럼 너비 조절 실패: {e}")

        # 모든 시트의 숫자 셀 포맷팅 (우측 정렬 + 천 단위 구분자)
        # 텍스트로 저장된 숫자도 실제 숫자로 변환
        try:
            from openpyxl.styles import Alignment
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        # 이미 숫자인 경우
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            if isinstance(cell.value, int) or (isinstance(cell.value, float) and cell.value == int(cell.value)):
                                cell.number_format = '#,##0'
                            else:
                                cell.number_format = '#,##0.0'  # 소수점 한 자리
                        # 문자열인데 숫자로 변환 가능한 경우
                        elif isinstance(cell.value, str) and cell.value.strip():
                            try:
                                # 콤마 제거 후 숫자 변환 시도
                                cleaned = cell.value.replace(',', '').strip()
                                # 음수 처리: -123 또는 (123) 형식
                                is_negative = False
                                if cleaned.startswith('(') and cleaned.endswith(')'):
                                    cleaned = cleaned[1:-1]
                                    is_negative = True
                                elif cleaned.startswith('-'):
                                    cleaned = cleaned[1:]
                                    is_negative = True

                                # 퍼센트 처리: 6.8% -> 0.068
                                is_percent = cleaned.endswith('%')
                                if is_percent:
                                    cleaned = cleaned[:-1].strip()

                                # 숫자인지 확인 (소수점 하나만 허용)
                                if cleaned.replace('.', '', 1).isdigit():
                                    num_value = float(cleaned)
                                    if is_negative:
                                        num_value = -num_value
                                    # 퍼센트면 100으로 나누고 퍼센트 포맷 적용
                                    if is_percent:
                                        num_value = num_value / 100
                                        cell.value = num_value
                                        cell.number_format = '0.0%'
                                    # 정수면 int로 변환
                                    elif num_value == int(num_value):
                                        cell.value = int(num_value)
                                        cell.number_format = '#,##0'
                                    else:
                                        cell.value = num_value
                                        cell.number_format = '#,##0.0'  # 소수점 한 자리
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                            except (ValueError, TypeError):
                                pass  # 숫자로 변환 불가능한 문자열은 무시
            print(f"[Excel] 숫자 포맷팅 완료 (우측 정렬 + 천 단위 구분자, 텍스트→숫자 변환 포함)")
        except Exception as e:
            print(f"[Excel] 숫자 포맷팅 실패: {e}")

        # 시트 순서 재배치: 기업개황 → Financials → ... → Frontdata(맨 끝)
        try:
            workbook = writer.book

            # Financials를 기업개황 바로 다음으로 이동
            if 'Financials' in workbook.sheetnames:
                current_idx = workbook.sheetnames.index('Financials')
                if current_idx != 1:
                    workbook.move_sheet('Financials', offset=1 - current_idx)

            # Frontdata를 맨 끝으로 이동
            if 'Frontdata' in workbook.sheetnames:
                current_idx = workbook.sheetnames.index('Frontdata')
                target_idx = len(workbook.sheetnames) - 1
                if current_idx != target_idx:
                    workbook.move_sheet('Frontdata', offset=target_idx - current_idx)

            print(f"[Excel] 시트 순서 재배치 완료: {workbook.sheetnames}")
        except Exception as e:
            print(f"[Excel] 시트 순서 재배치 실패: {e}")


@app.post("/api/heartbeat/{task_id}")
async def heartbeat(task_id: str):
    """작업 유지 heartbeat - 브라우저가 열려있는 동안 작업 삭제 방지"""
    task = TASKS.get(task_id)
    if not task:
        return {"success": False}

    task['last_accessed'] = time.time()
    return {"success": True}


@app.get("/api/status/{task_id}")
async def get_task_status(task_id: str):
    """작업 상태 조회 API"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    # 마지막 접근 시간 업데이트 (유휴 감지용)
    task['last_accessed'] = time.time()

    result = {
        "status": task['status'],
        "progress": task['progress'],
        "message": task['message'],
        "filename": task.get('filename')
    }

    # 완료 시 미리보기 데이터 포함
    if task['status'] == 'completed' and task.get('preview_data'):
        result['preview_data'] = task['preview_data']

    # 현재주소 (사업보고서 기준) 포함
    if task.get('current_address'):
        result['current_address'] = task['current_address']

    # 현재 대표자 (사업보고서 기준) 포함
    if task.get('current_ceo'):
        result['current_ceo'] = task['current_ceo']

    return result


@app.post("/api/cancel/{task_id}")
async def cancel_task(task_id: str):
    """작업 취소 API"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")
    
    # 완료된 작업은 파일 삭제하지 않고 task 정보만 제거
    if task.get('status') == 'completed':
        # 완료된 작업: 파일은 유지, task 정보만 제거
        del TASKS[task_id]
        return {"success": True, "message": "작업 정보가 정리되었습니다."}
    
    # 진행 중인 작업: 취소 처리
    task['cancelled'] = True
    task['status'] = 'cancelled'
    task['message'] = '취소됨'
    
    # 정리 (진행 중인 작업만 파일 삭제)
    cleanup_task(task_id)
    
    return {"success": True, "message": "작업이 취소되었습니다."}


# ============================================================
# 재무분석 AI 분석 API
# ============================================================
@app.post("/api/analyze/{task_id}")
async def analyze_financial_data(task_id: str):
    """
    재무제표 AI 분석 시작 API

    추출 완료된 재무 데이터를 AI로 분석하여 인사이트를 생성합니다.
    """
    import threading
    import asyncio

    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    if task['status'] != 'completed':
        raise HTTPException(status_code=400, detail="재무제표 추출이 완료되지 않았습니다.")

    # 이미 분석 중이면 중복 실행 방지
    if task.get('analysis_status') == 'running':
        print(f"[ANALYSIS] 이미 분석 중인 작업입니다: task_id={task_id}")
        return {"success": True, "message": "이미 분석이 진행 중입니다.", "already_running": True}

    # 분석 상태 초기화
    task['analysis_status'] = 'running'
    task['analysis_progress'] = 0
    task['analysis_message'] = '분석 준비 중'
    task['analysis_result'] = None

    # 별도 스레드에서 async 함수 실행 (요청 종료 후에도 계속 실행)
    def run_in_thread():
        asyncio.run(run_financial_analysis(task_id))

    thread = threading.Thread(target=run_in_thread, daemon=True)
    thread.start()

    print(f"[ANALYSIS] 분석 스레드 시작됨: task_id={task_id}")

    return {"success": True, "message": "AI 분석이 시작되었습니다."}


async def run_financial_analysis(task_id: str):
    """AI 분석 백그라운드 작업"""
    import traceback

    print(f"[ANALYSIS] 분석 백그라운드 작업 시작: task_id={task_id}")

    task = TASKS.get(task_id)
    if not task:
        print(f"[ANALYSIS ERROR] Task not found: {task_id}")
        return

    # 취소 체크 함수
    def is_cancelled():
        return task.get('analysis_status') == 'cancelled'

    # 진행 상태 업데이트 콜백 (취소 체크 포함)
    def update_progress(progress: int, message: str):
        # 취소 확인
        if is_cancelled():
            raise Exception("사용자 이탈로 분석 취소됨")
        task['analysis_progress'] = progress
        task['analysis_message'] = message
        print(f"[ANALYSIS] Progress: {progress}% - {message}")

    try:
        # 취소 확인
        if is_cancelled():
            print(f"[ANALYSIS] 시작 전 취소됨: {task_id}")
            return

        print("[ANALYSIS] FinancialInsightAnalyzer 임포트 중...")
        from financial_insight_analyzer import FinancialInsightAnalyzer

        update_progress(5, '[1/5] 분석기 초기화 중')

        analyzer = FinancialInsightAnalyzer()

        # 재무 데이터와 기업 정보 가져오기
        preview_data = task.get('preview_data', {})
        company_info = task.get('company_info', {})

        print(f"[ANALYSIS] preview_data 키: {list(preview_data.keys()) if preview_data else '없음'}")
        print(f"[ANALYSIS] file_path: {task.get('file_path')}")

        # ★ preview_data가 비어있으면 복구 시도
        if not preview_data or not preview_data.get('vcm_display'):
            # 1차: Excel 파일에서 재로드
            excel_filepath = task.get('file_path')
            print(f"[ANALYSIS] preview_data 비어있음, Excel에서 재로드 시도: {excel_filepath}")
            if excel_filepath and os.path.exists(excel_filepath):
                try:
                    # vcm_display (Financials 시트) 로드
                    fin_df = pd.read_excel(excel_filepath, sheet_name='Financials', header=1, engine='openpyxl')
                    # 좌우 병렬 구조를 상하 구조로 변환
                    if '항목' in fin_df.columns and '항목.1' in fin_df.columns:
                        left_cols = [col for col in fin_df.columns if not col.endswith('.1')]
                        right_cols = [col for col in fin_df.columns if col.endswith('.1')]
                        left_df = fin_df[left_cols].copy()
                        right_df = fin_df[right_cols].copy()
                        right_df.columns = [col.replace('.1', '') for col in right_df.columns]
                        fin_df = pd.concat([left_df, right_df], ignore_index=True)
                        fin_df = fin_df.dropna(subset=['항목'])
                        fin_df = fin_df[fin_df['항목'].astype(str).str.strip() != '']
                    preview_data['vcm_display'] = safe_dataframe_to_json(fin_df)
                    print(f"[ANALYSIS] vcm_display Excel에서 로드 완료: {len(preview_data['vcm_display'])}개 행")

                    # vcm (Frontdata 시트) 로드
                    vcm_df = pd.read_excel(excel_filepath, sheet_name='Frontdata', engine='openpyxl')
                    preview_data['vcm'] = safe_dataframe_to_json(vcm_df)
                    print(f"[ANALYSIS] vcm Excel에서 로드 완료: {len(preview_data['vcm'])}개 행")

                    # 현금흐름표 (cf) 로드
                    try:
                        cf_df = pd.read_excel(excel_filepath, sheet_name='현금흐름표', engine='openpyxl')
                        preview_data['cf'] = safe_dataframe_to_json(cf_df)
                        print(f"[ANALYSIS] cf Excel에서 로드 완료: {len(preview_data['cf'])}개 행")
                    except Exception as cf_err:
                        print(f"[ANALYSIS] 현금흐름표 로드 실패 (무시): {cf_err}")

                except Exception as excel_err:
                    print(f"[ANALYSIS ERROR] Excel에서 데이터 로드 실패: {excel_err}")
                    raise Exception(f"재무 데이터를 불러올 수 없습니다. 재무제표를 다시 추출해주세요.")
            else:
                # 2차: Excel 파일 없음 - 추출 데이터에서 재생성 시도
                fs_data = task.get('fs_data')
                if fs_data:
                    print(f"[ANALYSIS] Excel 없음, fs_data에서 VCM 재생성 시도...")
                    try:
                        vcm_result = create_vcm_format(fs_data, None)
                        if isinstance(vcm_result, tuple):
                            vcm_df, display_df = vcm_result
                        else:
                            vcm_df = vcm_result
                            display_df = None

                        if vcm_df is not None and not vcm_df.empty:
                            preview_data['vcm'] = safe_dataframe_to_json(vcm_df)
                        if display_df is not None and not display_df.empty:
                            preview_data['vcm_display'] = safe_dataframe_to_json(display_df)
                        print(f"[ANALYSIS] VCM 재생성 완료: vcm_display={len(preview_data.get('vcm_display', []))}행")
                    except Exception as vcm_err:
                        print(f"[ANALYSIS ERROR] VCM 재생성 실패: {vcm_err}")
                        raise Exception(f"재무 데이터를 불러올 수 없습니다. 재무제표를 다시 추출해주세요.")
                else:
                    print(f"[ANALYSIS ERROR] Excel 파일 없고 fs_data도 없음")
                    raise Exception(f"재무제표 파일을 찾을 수 없습니다. 재무제표를 다시 추출해주세요.")

        # company_info가 없으면 기본값 설정
        if not company_info:
            company_info = {
                'corp_name': task.get('corp_name', '알 수 없음'),
                'corp_code': task.get('corp_code', ''),
                'induty_code': ''
            }
        # company_info에 corp_code가 없으면 task에서 보충
        if not company_info.get('corp_code'):
            company_info['corp_code'] = task.get('corp_code', '')

        # 사용자 설정 기간을 company_info에 추가 (AI 분석 시 DART 주석 추출에 사용)
        analysis_start_year = task.get('start_year')
        analysis_end_year = task.get('end_year')
        if analysis_start_year:
            company_info['analysis_start_year'] = analysis_start_year
        if analysis_end_year:
            company_info['analysis_end_year'] = analysis_end_year
        print(f"[ANALYSIS] 분석 기간: {analysis_start_year} ~ {analysis_end_year}")

        # 분석 실행 (콜백 전달)
        result = await analyzer.analyze(preview_data, company_info, update_progress)

        # 사업보고서에서 최신 대표자/주소 추출
        update_progress(95, '[5/5] 최신 기업정보 확인 중')
        try:
            from dart_company_info import DartCompanyInfo
            dart_info = DartCompanyInfo()
            corp_code = task.get('corp_code', '')
            if corp_code:
                latest_info = dart_info.get_company_info_from_report(corp_code)
                if latest_info:
                    # 정기보고서가 없는 경우, 기업개황정보를 최신으로 사용
                    if latest_info.get('no_report'):
                        if company_info and company_info.get('ceo_nm'):
                            result['current_ceo'] = company_info['ceo_nm']
                        if company_info and company_info.get('adres'):
                            result['current_address'] = company_info['adres']
                        result['no_report'] = True  # 프론트엔드에 알림
                    else:
                        if latest_info.get('ceo'):
                            result['current_ceo'] = latest_info['ceo']
                        if latest_info.get('address'):
                            result['current_address'] = latest_info['address']
        except Exception as info_err:
            print(f"[ANALYSIS] 최신 기업정보 추출 실패: {info_err}")
            import traceback
            traceback.print_exc()

        task['analysis_status'] = 'completed'
        task['analysis_progress'] = 100
        task['analysis_message'] = '분석 완료'
        task['analysis_result'] = result

        # 메모리 절약: AI 분석 완료 후 preview_data 정리
        # 챗봇이 아직 초기화되지 않은 경우 경량 컨텍스트를 보존
        if 'preview_data' in task:
            if 'chatbot' not in task:
                # 챗봇 미초기화 → 재무 컨텍스트 보존 (엑셀에 포함되는 모든 데이터)
                pd = task['preview_data']
                task['chatbot_context'] = {
                    'vcm_display': pd.get('vcm_display'),
                    'vcm': pd.get('vcm'),
                    'is': pd.get('is'),
                    'bs': pd.get('bs'),
                    'cis': pd.get('cis'),
                    'cf': pd.get('cf'),
                    'notes': pd.get('notes'),
                }
                print(f"[ANALYSIS] 챗봇용 재무 컨텍스트 보존: {task_id}")
            del task['preview_data']
            print(f"[ANALYSIS] preview_data 메모리 해제: {task_id}")

        # LLM 사용량 로깅 (로그인한 사용자인 경우)
        if task.get('user_id'):
            try:
                # 토큰 정보 추출 (result에 포함되어 있으면 사용, 없으면 기본값)
                token_info = result.get('token_usage', {}) if result else {}
                db.log_llm_usage(
                    user_id=task['user_id'],
                    corp_code=task.get('corp_code', ''),
                    corp_name=task.get('corp_name', ''),
                    model_name=token_info.get('model', 'gemini-2.5-pro'),
                    input_tokens=token_info.get('input_tokens', 0),
                    output_tokens=token_info.get('output_tokens', 0),
                    cost=token_info.get('cost', 0)
                )
                print(f"[ANALYSIS] LLM 사용량 기록 완료: user_id={task['user_id']}")
            except Exception as log_err:
                print(f"[ANALYSIS] LLM 사용량 기록 실패: {log_err}")

    except Exception as e:
        print(f"[ANALYSIS ERROR] {e}")
        print(traceback.format_exc())
        task['analysis_status'] = 'failed'
        task['analysis_message'] = f'분석 실패: {str(e)}'
        task['analysis_result'] = None


@app.get("/api/analyze-status/{task_id}")
async def get_analysis_status(task_id: str):
    """AI 분석 상태 조회 API"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    # 마지막 접근 시간 업데이트 (유휴 감지용)
    task['last_accessed'] = time.time()

    return {
        "success": True,
        "status": task.get('analysis_status', 'not_started'),
        "progress": task.get('analysis_progress', 0),
        "message": task.get('analysis_message', ''),
        "result": task.get('analysis_result'),
        "filename": task.get('filename')  # 다운로드용 파일명
    }


# ============================================================
# 기업 리서치 API
# ============================================================
@app.post("/api/super-research/{task_id}")
async def start_super_research(task_id: str):
    """
    기업 리서치 시작 API

    기업 정보를 기반으로 종합 리서치를 수행합니다.
    """
    import threading
    import asyncio

    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    if task['status'] != 'completed':
        raise HTTPException(status_code=400, detail="재무제표 추출이 완료되지 않았습니다.")

    # 이미 리서치 중이면 중복 실행 방지
    if task.get('super_research_status') == 'running':
        print(f"[SUPER_RESEARCH] 이미 리서치 중인 작업입니다: task_id={task_id}")
        return {"success": True, "message": "이미 리서치가 진행 중입니다.", "already_running": True}

    # 리서치 상태 초기화
    task['super_research_status'] = 'running'
    task['super_research_progress'] = 0
    task['super_research_message'] = '리서치 준비 중'
    task['super_research_result'] = None

    # 별도 스레드에서 async 함수 실행
    def run_in_thread():
        asyncio.run(run_super_research(task_id))

    thread = threading.Thread(target=run_in_thread, daemon=True)
    thread.start()

    print(f"[SUPER_RESEARCH] 리서치 스레드 시작됨: task_id={task_id}")

    return {"success": True, "message": "기업 리서치가 시작되었습니다."}


async def run_super_research(task_id: str):
    """기업 리서치 백그라운드 작업"""
    import traceback

    print(f"[SUPER_RESEARCH] 리서치 백그라운드 작업 시작: task_id={task_id}")

    task = TASKS.get(task_id)
    if not task:
        print(f"[SUPER_RESEARCH ERROR] Task not found: {task_id}")
        return

    # 취소 체크 함수
    def is_cancelled():
        return task.get('super_research_status') == 'cancelled'

    # 진행 상태 업데이트 콜백 (취소 체크 포함)
    def update_progress(progress: int, message: str, step_result=None):
        # 취소 확인
        if is_cancelled():
            raise Exception("사용자 이탈로 리서치 취소됨")
        task['super_research_progress'] = progress
        task['super_research_message'] = message
        print(f"[SUPER_RESEARCH] Progress: {progress}% - {message}")

    try:
        # 취소 확인
        if is_cancelled():
            print(f"[SUPER_RESEARCH] 시작 전 취소됨: {task_id}")
            return

        print("[SUPER_RESEARCH] SuperResearchPipeline 임포트 중...")
        from super_research_pipeline import SuperResearchPipeline

        update_progress(5, '[Step 1/5] 기업 정보 준비 중')

        pipeline = SuperResearchPipeline()

        # 기업 정보 가져오기
        company_info = task.get('company_info', {})
        if not company_info:
            company_info = {
                'corp_code': task.get('corp_code', ''),
                'corp_name': task.get('corp_name', '알 수 없음')
            }

        # 파이프라인 실행
        result = await pipeline.run(company_info, update_progress)

        # 결과 저장
        task['super_research_status'] = 'completed'
        task['super_research_progress'] = 100
        task['super_research_message'] = '리서치 완료'
        task['super_research_result'] = {
            'success': result.success,
            'report': result.report,
            'keywords': result.keywords,
            'competitors_domestic': [
                {
                    'name': c.name,
                    'ticker': c.ticker,
                    'business': c.business,
                    'reason': c.reason,
                    'detailed_business': c.detailed_business
                } for c in result.competitors_domestic
            ] if result.competitors_domestic else [],
            'competitors_international': [
                {
                    'name': c.name,
                    'ticker': c.ticker,
                    'business': c.business,
                    'reason': c.reason,
                    'detailed_business': c.detailed_business
                } for c in result.competitors_international
            ] if result.competitors_international else [],
            'partners_domestic': [
                {
                    'name': p.name,
                    'ticker': p.ticker,
                    'business': p.business,
                    'reason': p.reason,
                    'relationship_type': p.relationship_type,
                    'detailed_business': p.detailed_business
                } for p in result.partners_domestic
            ] if result.partners_domestic else [],
            'partners_international': [
                {
                    'name': p.name,
                    'ticker': p.ticker,
                    'business': p.business,
                    'reason': p.reason,
                    'relationship_type': p.relationship_type,
                    'detailed_business': p.detailed_business
                } for p in result.partners_international
            ] if result.partners_international else [],
            'mna_domestic': [
                {
                    'acquirer': m.acquirer,
                    'target': m.target,
                    'date': m.date,
                    'price': m.price,
                    'source_url': getattr(m, 'source_url', '')
                } for m in result.mna_domestic
            ] if result.mna_domestic else [],
            'mna_international': [
                {
                    'acquirer': m.acquirer,
                    'target': m.target,
                    'date': m.date,
                    'price': m.price,
                    'source_url': getattr(m, 'source_url', '')
                } for m in result.mna_international
            ] if result.mna_international else [],
            'business_summary': result.business.raw_business[:2000] if result.business and result.business.raw_business else '',
            'major_news': result.major_news.raw_news if result.major_news else ''
        }

        print(f"[SUPER_RESEARCH] 리서치 완료: task_id={task_id}")

        # 메모리 절약: 리서치 완료 후 preview_data 정리 (챗봇 활성화 시 유지)
        if 'preview_data' in task and 'chatbot' not in task:
            del task['preview_data']
            print(f"[SUPER_RESEARCH] preview_data 메모리 해제: {task_id}")

    except Exception as e:
        print(f"[SUPER_RESEARCH ERROR] 리서치 실패: {e}")
        traceback.print_exc()
        task['super_research_status'] = 'error'
        task['super_research_message'] = f'리서치 실패: {str(e)}'
        task['super_research_result'] = None


@app.get("/api/super-research-status/{task_id}")
async def get_super_research_status(task_id: str):
    """기업 리서치 상태 조회 API"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    # 마지막 접근 시간 업데이트
    task['last_accessed'] = time.time()

    return {
        "success": True,
        "status": task.get('super_research_status', 'not_started'),
        "progress": task.get('super_research_progress', 0),
        "message": task.get('super_research_message', ''),
        "result": task.get('super_research_result')
    }


# ============================================================
# PE 챗봇 API
# ============================================================

class ChatMessageRequest(BaseModel):
    message: str


@app.post("/api/chat/init/{task_id}")
async def init_chatbot(task_id: str):
    """챗봇 세션 초기화"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    task['last_accessed'] = time.time()

    try:
        from pe_chatbot import PEChatbot

        company_info = task.get('company_info', {})
        preview_data = task.get('preview_data') or task.get('chatbot_context') or {}
        analysis_result = task.get('analysis_result')
        research_result = task.get('super_research_result')

        chatbot = PEChatbot(
            company_info=company_info,
            preview_data=preview_data,
            analysis_result=analysis_result,
            research_result=research_result
        )

        task['chatbot'] = chatbot

        return {
            "success": True,
            "suggestions": chatbot.get_suggestions(),
            "company_name": chatbot.company_name
        }

    except Exception as e:
        print(f"[CHAT] 챗봇 초기화 실패: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"챗봇 초기화 실패: {str(e)}")


@app.post("/api/chat/message/{task_id}")
async def chat_message(task_id: str, request: ChatMessageRequest):
    """채팅 메시지 전송 (SSE 스트리밍)"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    task['last_accessed'] = time.time()

    chatbot = task.get('chatbot')
    if not chatbot:
        raise HTTPException(status_code=400, detail="챗봇이 초기화되지 않았습니다. /api/chat/init을 먼저 호출하세요.")

    message = request.message.strip()
    if not message:
        raise HTTPException(status_code=400, detail="메시지가 비어있습니다.")

    # AI 분석/리서치 결과 업데이트 (완료된 경우)
    analysis_result = task.get('analysis_result')
    research_result = task.get('super_research_result')
    if analysis_result or research_result:
        chatbot.update_context(
            analysis_result=analysis_result,
            research_result=research_result
        )

    async def event_generator():
        queue = asyncio.Queue()

        async def producer():
            try:
                async for event in chatbot.chat(message):
                    await queue.put(event)
            except Exception as e:
                print(f"[CHAT] 스트리밍 오류: {e}")
                import traceback
                traceback.print_exc()
                await queue.put({"type": "error", "content": str(e)})
            finally:
                await queue.put(None)  # 종료 신호

        asyncio.create_task(producer())

        while True:
            try:
                event = await asyncio.wait_for(queue.get(), timeout=15)
                if event is None:
                    break
                if isinstance(event, dict):
                    yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                else:
                    yield f"data: {json.dumps({'type': 'token', 'content': event}, ensure_ascii=False)}\n\n"
            except asyncio.TimeoutError:
                # SSE keepalive — 연결 유지용 (클라이언트에서 무시됨)
                yield ": keepalive\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/chat/history/{task_id}")
async def get_chat_history(task_id: str):
    """대화 내역 조회"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    task['last_accessed'] = time.time()

    chatbot = task.get('chatbot')
    if not chatbot:
        return {"success": True, "messages": [], "suggestions": []}

    return {
        "success": True,
        "messages": chatbot.get_history(),
        "suggestions": chatbot.get_suggestions()
    }


@app.post("/api/chat/update-context/{task_id}")
async def update_chat_context(task_id: str):
    """챗봇 컨텍스트 업데이트 (AI 분석/리서치 완료 후 호출)"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    chatbot = task.get('chatbot')
    if not chatbot:
        return {"success": False, "message": "챗봇이 초기화되지 않았습니다."}

    chatbot.update_context(
        analysis_result=task.get('analysis_result'),
        research_result=task.get('super_research_result')
    )

    return {
        "success": True,
        "suggestions": chatbot.get_suggestions()
    }


def apply_excel_formatting(filepath: str):
    """기존 엑셀 파일에 포맷팅 재적용 (문자열→숫자 변환 포함)"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, PatternFill, Font
        from openpyxl.utils import get_column_letter

        wb = load_workbook(filepath)
        total_converted = 0

        # 헤더 스타일 정의 (#131313 배경, 흰색 글자)
        header_fill = PatternFill(start_color='131313', end_color='131313', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # Financials 시트는 별도 처리 (이미 스타일 적용됨, 문자열→숫자 변환도 건너뜀)
        skip_sheets = ['Financials']

        for sheet_name in wb.sheetnames:
            # Financials 시트는 완전히 건너뜀 (퍼센트 문자열 유지)
            if sheet_name in skip_sheets:
                continue
            ws = wb[sheet_name]
            sheet_converted = 0

            # 헤더 행(1행) 스타일 적용 (Financials 제외)
            if sheet_name not in skip_sheets:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    # 모든 헤더: 가로 중앙정렬 + 세로 중앙정렬
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Frontdata 시트: 하위항목 들여쓰기 적용
            if sheet_name == 'Frontdata':
                # 컬럼 인덱스 찾기
                col_indices = {}
                for col in range(1, ws.max_column + 1):
                    header_val = ws.cell(row=1, column=col).value
                    if header_val:
                        col_indices[header_val] = col

                type_col = col_indices.get('타입')
                item_col = col_indices.get('항목')

                if type_col and item_col:
                    for row in range(2, ws.max_row + 1):
                        item_type = ws.cell(row=row, column=type_col).value
                        if item_type in ['subitem', 'percent']:
                            item_cell = ws.cell(row=row, column=item_col)
                            item_cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)

            # 컬럼 너비 자동 조절
            for column_cells in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_length = 0
                            cell_str = str(cell.value)
                            for char in cell_str:
                                if ord(char) > 127:
                                    cell_length += 2
                                else:
                                    cell_length += 1
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 8), 60)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 숫자 셀 포맷팅 (우측 정렬 + 천 단위 구분자 + 문자열→숫자 변환)
            for row in ws.iter_rows():
                for cell in row:
                    # 이미 숫자인 경우
                    if isinstance(cell.value, (int, float)) and cell.value is not None:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if isinstance(cell.value, int) or (isinstance(cell.value, float) and cell.value == int(cell.value)):
                            cell.number_format = '#,##0'
                        else:
                            cell.number_format = '#,##0.0'  # 소수점 한 자리
                    # 문자열인데 숫자로 변환 가능한 경우
                    elif isinstance(cell.value, str) and cell.value.strip():
                        try:
                            cleaned = cell.value.replace(',', '').strip()

                            # '-'만 있는 경우 스킵 (대시 기호)
                            if cleaned == '-' or cleaned == '':
                                continue

                            is_negative = False
                            if cleaned.startswith('(') and cleaned.endswith(')'):
                                cleaned = cleaned[1:-1]
                                is_negative = True
                            elif cleaned.startswith('-') and len(cleaned) > 1:
                                cleaned = cleaned[1:]
                                is_negative = True

                            # 퍼센트 처리
                            is_percent = cleaned.endswith('%')
                            if is_percent:
                                cleaned = cleaned[:-1].strip()

                            # 빈 문자열 체크
                            if not cleaned:
                                continue

                            if cleaned.replace('.', '', 1).isdigit():
                                num_value = float(cleaned)
                                if is_negative:
                                    num_value = -num_value
                                if is_percent:
                                    num_value = num_value / 100
                                    cell.value = num_value
                                    cell.number_format = '0.0%'
                                elif num_value == int(num_value):
                                    cell.value = int(num_value)
                                    cell.number_format = '#,##0'
                                else:
                                    cell.value = num_value
                                    cell.number_format = '#,##0.0'  # 소수점 한 자리
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                                sheet_converted += 1
                        except (ValueError, TypeError, OverflowError):
                            pass

            if sheet_converted > 0:
                print(f"  - {sheet_name}: {sheet_converted}개 셀 변환")
            total_converted += sheet_converted

        wb.save(filepath)
        print(f"[Excel] 포맷팅 재적용 완료: {filepath} (총 {total_converted}개 셀 변환)")
        return True
    except Exception as e:
        import traceback
        print(f"[Excel] 포맷팅 재적용 실패: {e}")
        traceback.print_exc()
        return False


def add_super_research_sheet(filepath: str, task: dict):
    """기업 리서치 결과를 엑셀 시트로 추가 (프론트엔드와 동일한 형태로 통합)"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

        wb = load_workbook(filepath)

        # 기존 기업 리서치 관련 시트들 모두 삭제
        sheets_to_delete = ['기업 리서치', '기업 리서치 - 국내 경쟁사', '기업 리서치 - 해외 경쟁사', '기업 리서치 - M&A']
        for sheet_name in sheets_to_delete:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        # 새 시트 생성
        ws = wb.create_sheet('기업 리서치')

        # 스타일 정의
        header_fill = PatternFill(start_color='131313', end_color='131313', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        section_font = Font(bold=True, size=11, color='1a1a1a')
        subsection_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin', color='cccccc'),
            right=Side(style='thin', color='cccccc'),
            top=Side(style='thin', color='cccccc'),
            bottom=Side(style='thin', color='cccccc')
        )

        # 기업 리서치 상태 확인
        super_research_status = task.get('super_research_status', 'not_started')
        super_research_result = task.get('super_research_result')

        if super_research_status == 'completed' and super_research_result:
            report = super_research_result.get('report', '')
            row_idx = 1

            if isinstance(report, dict):
                # 1. 제목
                if report.get('title'):
                    ws.cell(row=row_idx, column=1, value=report['title'])
                    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=16)
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                    row_idx += 2

                # 2. 기업 개요 (테이블 형식)
                if report.get('overview'):
                    ws.cell(row=row_idx, column=1, value='기업 개요')
                    ws.cell(row=row_idx, column=1).font = section_font
                    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                    row_idx += 1
                    for key, val in report['overview'].items():
                        if val:
                            ws.cell(row=row_idx, column=1, value=key)
                            ws.cell(row=row_idx, column=1).font = Font(bold=True)
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=val)
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
                            row_idx += 1
                    row_idx += 1

                # 3. 사업 영역
                if report.get('business'):
                    ws.cell(row=row_idx, column=1, value='사업 영역')
                    ws.cell(row=row_idx, column=1).font = section_font
                    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                    row_idx += 1
                    if report['business'].get('summary'):
                        ws.cell(row=row_idx, column=1, value=report['business']['summary'])
                        ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                        row_idx += 1
                    for area in report['business'].get('areas', []):
                        ws.cell(row=row_idx, column=1, value=f"• {area.get('name', '')}")
                        ws.cell(row=row_idx, column=1).font = Font(bold=True)
                        ws.cell(row=row_idx, column=2, value=area.get('description', ''))
                        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
                        row_idx += 1
                    row_idx += 1

                # 4. 경쟁사 분석
                has_domestic = report.get('competitors', {}).get('domestic')
                has_international = report.get('competitors', {}).get('international')
                if has_domestic or has_international:
                    ws.cell(row=row_idx, column=1, value='경쟁사 분석')
                    ws.cell(row=row_idx, column=1).font = section_font
                    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                    row_idx += 1

                    # 국내 경쟁사
                    if has_domestic:
                        ws.cell(row=row_idx, column=1, value='▶ 국내 경쟁사')
                        ws.cell(row=row_idx, column=1).font = subsection_font
                        row_idx += 1
                        # 헤더
                        headers = ['경쟁사', '티커', '경쟁 분야', '경쟁 이유']
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col, value=h)
                            ws.cell(row=row_idx, column=col).fill = header_fill
                            ws.cell(row=row_idx, column=col).font = header_font
                            ws.cell(row=row_idx, column=col).border = thin_border
                        row_idx += 1
                        for comp in report['competitors']['domestic']:
                            ws.cell(row=row_idx, column=1, value=comp.get('name', ''))
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=comp.get('ticker', ''))
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=3, value=comp.get('field', ''))
                            ws.cell(row=row_idx, column=3).border = thin_border
                            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)
                            ws.cell(row=row_idx, column=4, value=comp.get('reason', ''))
                            ws.cell(row=row_idx, column=4).border = thin_border
                            ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)
                            row_idx += 1
                        row_idx += 1

                    # 해외 경쟁사
                    if has_international:
                        ws.cell(row=row_idx, column=1, value='▶ 해외 경쟁사')
                        ws.cell(row=row_idx, column=1).font = subsection_font
                        row_idx += 1
                        # 헤더
                        headers = ['경쟁사', '티커', '경쟁 분야', '경쟁 이유']
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col, value=h)
                            ws.cell(row=row_idx, column=col).fill = header_fill
                            ws.cell(row=row_idx, column=col).font = header_font
                            ws.cell(row=row_idx, column=col).border = thin_border
                        row_idx += 1
                        for comp in report['competitors']['international']:
                            ws.cell(row=row_idx, column=1, value=comp.get('name', ''))
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=comp.get('ticker', ''))
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=3, value=comp.get('field', ''))
                            ws.cell(row=row_idx, column=3).border = thin_border
                            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)
                            ws.cell(row=row_idx, column=4, value=comp.get('reason', ''))
                            ws.cell(row=row_idx, column=4).border = thin_border
                            ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)
                            row_idx += 1
                        row_idx += 1

                # 4-1. 협력사 분석
                has_partners_domestic = report.get('partners', {}).get('domestic')
                has_partners_international = report.get('partners', {}).get('international')
                if has_partners_domestic or has_partners_international:
                    ws.cell(row=row_idx, column=1, value='협력사 분석')
                    ws.cell(row=row_idx, column=1).font = section_font
                    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
                    row_idx += 1

                    # 국내 협력사
                    if has_partners_domestic:
                        ws.cell(row=row_idx, column=1, value='▶ 국내 협력사')
                        ws.cell(row=row_idx, column=1).font = subsection_font
                        row_idx += 1
                        headers = ['기업명', '티커', '관계 유형', '사업 분야', '관계 설명']
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col, value=h)
                            ws.cell(row=row_idx, column=col).fill = header_fill
                            ws.cell(row=row_idx, column=col).font = header_font
                            ws.cell(row=row_idx, column=col).border = thin_border
                        row_idx += 1
                        rel_labels = {'partner': '제휴', 'supplier': '공급업체', 'customer': '고객사'}
                        for p in report['partners']['domestic']:
                            ws.cell(row=row_idx, column=1, value=p.get('name', ''))
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=p.get('ticker', ''))
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=3, value=rel_labels.get(p.get('relationship_type', ''), p.get('relationship_type', '')))
                            ws.cell(row=row_idx, column=3).border = thin_border
                            ws.cell(row=row_idx, column=4, value=p.get('field', ''))
                            ws.cell(row=row_idx, column=4).border = thin_border
                            ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)
                            ws.cell(row=row_idx, column=5, value=p.get('reason', ''))
                            ws.cell(row=row_idx, column=5).border = thin_border
                            ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True)
                            row_idx += 1
                        row_idx += 1

                    # 해외 협력사
                    if has_partners_international:
                        ws.cell(row=row_idx, column=1, value='▶ 해외 협력사')
                        ws.cell(row=row_idx, column=1).font = subsection_font
                        row_idx += 1
                        headers = ['기업명', '티커', '관계 유형', '사업 분야', '관계 설명']
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col, value=h)
                            ws.cell(row=row_idx, column=col).fill = header_fill
                            ws.cell(row=row_idx, column=col).font = header_font
                            ws.cell(row=row_idx, column=col).border = thin_border
                        row_idx += 1
                        rel_labels = {'partner': '제휴', 'supplier': '공급업체', 'customer': '고객사'}
                        for p in report['partners']['international']:
                            ws.cell(row=row_idx, column=1, value=p.get('name', ''))
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=p.get('ticker', ''))
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=3, value=rel_labels.get(p.get('relationship_type', ''), p.get('relationship_type', '')))
                            ws.cell(row=row_idx, column=3).border = thin_border
                            ws.cell(row=row_idx, column=4, value=p.get('field', ''))
                            ws.cell(row=row_idx, column=4).border = thin_border
                            ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)
                            ws.cell(row=row_idx, column=5, value=p.get('reason', ''))
                            ws.cell(row=row_idx, column=5).border = thin_border
                            ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True)
                            row_idx += 1
                        row_idx += 1

                # 5. M&A 사례
                has_mna_domestic = report.get('mna', {}).get('domestic')
                has_mna_international = report.get('mna', {}).get('international')
                if has_mna_domestic or has_mna_international:
                    ws.cell(row=row_idx, column=1, value='M&A 사례')
                    ws.cell(row=row_idx, column=1).font = section_font
                    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                    row_idx += 1

                    # 국내 M&A
                    if has_mna_domestic:
                        ws.cell(row=row_idx, column=1, value='▶ 국내 M&A')
                        ws.cell(row=row_idx, column=1).font = subsection_font
                        row_idx += 1
                        headers = ['인수기업', '피인수기업', '시기', '금액', '출처']
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col, value=h)
                            ws.cell(row=row_idx, column=col).fill = header_fill
                            ws.cell(row=row_idx, column=col).font = header_font
                            ws.cell(row=row_idx, column=col).border = thin_border
                        row_idx += 1
                        for m in report['mna']['domestic']:
                            ws.cell(row=row_idx, column=1, value=m.get('acquirer', '-'))
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=m.get('target', '-'))
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=3, value=m.get('date', '-'))
                            ws.cell(row=row_idx, column=3).border = thin_border
                            ws.cell(row=row_idx, column=4, value=m.get('price', '-'))
                            ws.cell(row=row_idx, column=4).border = thin_border
                            source_url = m.get('source_url', '')
                            ws.cell(row=row_idx, column=5, value=source_url if source_url else '[미검증]')
                            ws.cell(row=row_idx, column=5).border = thin_border
                            row_idx += 1
                        row_idx += 1

                    # 해외 M&A
                    if has_mna_international:
                        ws.cell(row=row_idx, column=1, value='▶ 해외 M&A')
                        ws.cell(row=row_idx, column=1).font = subsection_font
                        row_idx += 1
                        headers = ['인수기업', '피인수기업', '시기', '금액', '출처']
                        for col, h in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col, value=h)
                            ws.cell(row=row_idx, column=col).fill = header_fill
                            ws.cell(row=row_idx, column=col).font = header_font
                            ws.cell(row=row_idx, column=col).border = thin_border
                        row_idx += 1
                        for m in report['mna']['international']:
                            ws.cell(row=row_idx, column=1, value=m.get('acquirer', '-'))
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=m.get('target', '-'))
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=3, value=m.get('date', '-'))
                            ws.cell(row=row_idx, column=3).border = thin_border
                            ws.cell(row=row_idx, column=4, value=m.get('price', '-'))
                            ws.cell(row=row_idx, column=4).border = thin_border
                            source_url = m.get('source_url', '')
                            ws.cell(row=row_idx, column=5, value=source_url if source_url else '[미검증]')
                            ws.cell(row=row_idx, column=5).border = thin_border
                            row_idx += 1
                        row_idx += 1

                # 6. 주요 뉴스
                if report.get('news'):
                    ws.cell(row=row_idx, column=1, value='주요 뉴스')
                    ws.cell(row=row_idx, column=1).font = section_font
                    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                    row_idx += 1
                    news_data = report['news']
                    if isinstance(news_data, list):
                        for news_item in news_data:
                            type_str = news_item.get('type', '') if isinstance(news_item, dict) else ''
                            date_str = news_item.get('date', '') if isinstance(news_item, dict) else ''
                            title_str = news_item.get('title', '') if isinstance(news_item, dict) else str(news_item)
                            desc_str = news_item.get('description', '') if isinstance(news_item, dict) else ''
                            url_str = news_item.get('url', '') if isinstance(news_item, dict) else ''
                            label = f"[{type_str}] {date_str}" if type_str else date_str
                            # 제목 + 설명 합쳐서 표시
                            full_title = f"{title_str}\n{desc_str}" if desc_str else title_str
                            ws.cell(row=row_idx, column=1, value=label)
                            ws.cell(row=row_idx, column=1).border = thin_border
                            ws.cell(row=row_idx, column=2, value=full_title)
                            ws.cell(row=row_idx, column=2).border = thin_border
                            ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
                            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
                            ws.cell(row=row_idx, column=4, value=url_str)
                            ws.cell(row=row_idx, column=4).border = thin_border
                            row_idx += 1
                    else:
                        ws.cell(row=row_idx, column=1, value=str(news_data))
                        ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)

            elif report:
                # 문자열인 경우 (fallback)
                lines = report.split('\n')
                for i, line in enumerate(lines, start=1):
                    ws.cell(row=i, column=1, value=line)
                    ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True)

            # 컬럼 너비 설정
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 50

            print(f"[Excel] 기업 리서치 시트 추가 완료 (통합)")

        else:
            # 진행 중 또는 미시작인 경우: 안내 메시지 표시
            ws.cell(row=1, column=1, value='기업 리서치')
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).font = header_font

            if super_research_status == 'running':
                progress = task.get('super_research_progress', 0)
                message = f"기업 리서치가 진행 중입니다. ({progress}%)\n완료 후 다시 엑셀 다운로드를 시도하세요."
            else:
                message = "아직 기업 리서치 완료 전입니다.\n완료 이후 다시 엑셀 다운로드를 시도하세요."

            ws.cell(row=2, column=1, value=message)
            ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
            ws.column_dimensions['A'].width = 60
            ws.row_dimensions[2].height = 40

            print(f"[Excel] 기업 리서치 시트 추가 (대기 메시지)")

        # 시트 순서 정리: 재무분석AI(원본) 바로 다음에 기업 리서치 배치
        base_idx = None
        if '재무분석AI(원본)' in wb.sheetnames:
            base_idx = wb.sheetnames.index('재무분석AI(원본)')
        elif '재무분석AI(요약)' in wb.sheetnames:
            base_idx = wb.sheetnames.index('재무분석AI(요약)')
        elif 'Financials' in wb.sheetnames:
            base_idx = wb.sheetnames.index('Financials')

        if base_idx is not None and '기업 리서치' in wb.sheetnames:
            target_idx = base_idx + 1
            current_idx = wb.sheetnames.index('기업 리서치')
            if current_idx != target_idx:
                wb.move_sheet('기업 리서치', offset=target_idx - current_idx)

        wb.save(filepath)

    except Exception as e:
        import traceback
        print(f"[Excel] 기업 리서치 시트 추가 실패: {e}")
        traceback.print_exc()


@app.get("/api/download/{task_id}")
async def download_file(task_id: str):
    """파일 다운로드 API"""
    task = TASKS.get(task_id)

    # 동시 사용자 환경에서 다른 사람 파일 다운로드 방지
    if not task:
        raise HTTPException(status_code=404, detail="작업이 만료되었습니다. 다시 추출해주세요.")

    if task['status'] != 'completed':
        raise HTTPException(status_code=400, detail="작업이 완료되지 않았습니다.")

    filepath = task.get('file_path')
    if not filepath or not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")

    # 기업 리서치 시트 추가
    add_super_research_sheet(filepath, task)

    # 다운로드 전 포맷팅 재적용 (서버 코드 변경사항 반영)
    apply_excel_formatting(filepath)

    filename = task.get('filename', 'financial_statement.xlsx')

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.get("/api/download-file/{filename}")
async def download_file_by_name(filename: str):
    """파일명으로 직접 다운로드 (서버 재시작 후에도 사용 가능)"""
    # 보안: 파일명에 경로 조작 문자 포함 시 거부
    if '..' in filename or '/' in filename or '\\' in filename:
        raise HTTPException(status_code=400, detail="잘못된 파일명입니다.")

    output_dir = os.path.join(os.path.dirname(__file__), 'output')
    filepath = os.path.join(output_dir, filename)

    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")

    # 다운로드 전 포맷팅 재적용
    apply_excel_formatting(filepath)

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.post("/api/add-insight/{task_id}")
async def add_insight_to_excel(task_id: str, request: Request):
    """재무분석 AI를 엑셀 파일에 추가하는 API - AI 분석 완료 시 엑셀 파일 생성"""
    task = TASKS.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다.")

    try:
        data = await request.json()
        report = data.get('report', '')
        summary_report = data.get('summary_report', '')

        if not report:
            raise HTTPException(status_code=400, detail="보고서 내용이 없습니다.")

        filepath = task.get('file_path')

        # 엑셀 파일이 없으면 새로 생성
        if not filepath or not os.path.exists(filepath):
            print(f"[ADD-INSIGHT] 엑셀 파일 없음, 새로 생성 시작")

            # task에서 데이터 가져오기
            fs_data = task.get('fs_data')
            company_info = task.get('company_info')
            corp_name = task.get('corp_name', '기업')

            if not fs_data:
                raise HTTPException(status_code=400, detail="재무 데이터가 없습니다. 추출을 다시 진행해주세요.")

            # 파일명 생성
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{corp_name}_재무제표_{timestamp}.xlsx"
            output_dir = os.path.join(os.path.dirname(__file__), 'output')
            filepath = os.path.join(output_dir, filename)

            # 엑셀 파일 생성 (동기 함수를 스레드에서 실행)
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(
                None,
                lambda: save_to_excel(fs_data, filepath, company_info)
            )

            # task에 파일 경로 저장
            task['file_path'] = filepath
            task['filename'] = filename
            print(f"[ADD-INSIGHT] 엑셀 파일 생성 완료: {filepath}")

        # 엑셀 파일에 재무분석 AI 시트 추가 (요약 + 원본)
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = load_workbook(filepath)

        # 스타일 정의
        title_font = Font(name='맑은 고딕', size=14, bold=True, color='1F4E79')
        header_font = Font(name='맑은 고딕', size=12, bold=True, color='2E75B6')
        subheader_font = Font(name='맑은 고딕', size=11, bold=True, color='404040')
        normal_font = Font(name='맑은 고딕', size=10)

        # 마크다운 및 HTML 포맷팅 제거 함수
        def strip_markdown(text):
            import re
            text = re.sub(r'<strong>(.+?)</strong>', r'\1', text)
            text = re.sub(r'<em>(.+?)</em>', r'\1', text)
            text = re.sub(r'<li>', '', text)
            text = re.sub(r'</li>', '', text)
            text = re.sub(r'<ul>|</ul>|<ol>|</ol>', '', text)
            text = re.sub(r'<p>|</p>', '', text)
            text = re.sub(r'<br\s*/?>', '', text)
            text = re.sub(r'<h[1-6]>|</h[1-6]>', '', text)
            text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
            text = re.sub(r'__(.+?)__', r'\1', text)
            text = re.sub(r'(?<!\*)\*([^*\n]+)\*(?!\*)', r'\1', text)
            text = re.sub(r'(?<!_)_([^_\n]+)_(?!_)', r'\1', text)
            text = re.sub(r'^[-*]\s+', '', text)
            text = re.sub(r'^\d+\.\s+', '', text)
            return text

        def write_report_to_sheet(wb, sheet_name, report_text):
            """보고서를 엑셀 시트에 작성하는 헬퍼 함수"""
            # 기존 시트가 있으면 삭제
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

            sheet = wb.create_sheet(sheet_name)

            lines = report_text.split('\n')
            row = 1
            for line in lines:
                clean_line = strip_markdown(line.strip())
                cell = sheet.cell(row=row, column=1, value=clean_line)

                if line.startswith('# '):
                    cell.value = strip_markdown(line[2:].strip())
                    cell.font = title_font
                elif line.startswith('## '):
                    cell.value = strip_markdown(line[3:].strip())
                    cell.font = header_font
                elif line.startswith('### '):
                    cell.value = strip_markdown(line[4:].strip())
                    cell.font = subheader_font
                else:
                    cell.font = normal_font

                cell.alignment = Alignment(wrap_text=True, vertical='top')
                row += 1

            sheet.column_dimensions['A'].width = 100
            return sheet

        # 기존 '재무분석 AI' 시트도 삭제 (이전 버전 호환)
        if '재무분석 AI' in wb.sheetnames:
            del wb['재무분석 AI']

        # 요약본 시트 작성
        if summary_report:
            write_report_to_sheet(wb, '재무분석AI(요약)', summary_report)
        # 원본 시트 작성
        write_report_to_sheet(wb, '재무분석AI(원본)', report)

        # 시트 순서 정리: Financials → 재무분석AI(요약) → 재무분석AI(원본) → ...
        if 'Financials' in wb.sheetnames:
            financials_idx = wb.sheetnames.index('Financials')

            if summary_report and '재무분석AI(요약)' in wb.sheetnames:
                summary_idx = wb.sheetnames.index('재무분석AI(요약)')
                target = financials_idx + 1
                if summary_idx != target:
                    wb.move_sheet('재무분석AI(요약)', offset=target - summary_idx)

            if '재무분석AI(원본)' in wb.sheetnames:
                original_idx = wb.sheetnames.index('재무분석AI(원본)')
                target = wb.sheetnames.index('재무분석AI(요약)') + 1 if summary_report and '재무분석AI(요약)' in wb.sheetnames else financials_idx + 1
                if original_idx != target:
                    wb.move_sheet('재무분석AI(원본)', offset=target - original_idx)

        wb.save(filepath)
        wb.close()

        # 완료 시간 갱신 (TTL 리셋) - AI 분석 완료 후 다운로드 시간 확보
        task['completed_at'] = time.time()

        return {"success": True, "message": "재무분석 AI(원본+요약)가 엑셀에 추가되었습니다."}

    except Exception as e:
        print(f"[오류] 재무분석 AI 엑셀 추가 실패: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/task/{task_id}")
async def delete_task(task_id: str):
    """작업 및 파일 삭제 API"""
    task = TASKS.get(task_id)
    if task:
        # 파일 삭제
        if task.get('file_path') and os.path.exists(task['file_path']):
            try:
                os.remove(task['file_path'])
            except Exception:
                pass
        # 작업 삭제
        del TASKS[task_id]

    return {"success": True}


# ============================================================
# 메인 실행
# ============================================================
if __name__ == "__main__":
    import uvicorn
    import os

    port = int(os.environ.get("PORT", 8000))

    print("=" * 50)
    print("DART 재무제표 추출기 서버")
    print("=" * 50)
    print(f"서버 시작: http://localhost:{port}")
    print("종료: Ctrl+C")
    print("=" * 50)

    uvicorn.run(app, host="0.0.0.0", port=port)
