#!/usr/bin/env python3
"""
PERS 운용 종합 테스트
- 다양한 FY 연도 수 (2년, 3년, 5년)
- 엣지케이스 기업 (자본잠식, 별도재무제표, 비상장사)
- 엑셀 컬럼 너비 검증
- 세션/Heartbeat/다운로드 시나리오
- 챗봇 초기화/메시지
- 동시성 테스트
"""

import requests
import time
import json
import sys
import os
import asyncio
import concurrent.futures
from datetime import datetime

BASE_URL = "http://localhost:8001"
RESULTS = []
TASK_IDS = []  # cleanup용


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def record(name, passed, detail=""):
    status = "PASS" if passed else "FAIL"
    RESULTS.append({"name": name, "passed": passed, "detail": detail})
    icon = "\u2713" if passed else "\u2717"
    log(f"{icon} {name}: {detail}")
    return passed


def extract_company(corp_code, corp_name, start_year=2022, end_year=2024, timeout=360):
    """기업 추출 후 task_id 반환"""
    resp = requests.post(f"{BASE_URL}/api/extract", json={
        "corp_code": corp_code,
        "corp_name": corp_name,
        "start_year": start_year,
        "end_year": end_year
    })
    if resp.status_code != 200:
        return None, f"extract API failed: {resp.status_code}"

    task_id = resp.json().get("task_id")
    TASK_IDS.append(task_id)

    # 완료 대기
    start = time.time()
    while time.time() - start < timeout:
        time.sleep(3)
        r = requests.get(f"{BASE_URL}/api/status/{task_id}")
        data = r.json()
        status = data.get("status")
        if status == "completed":
            return task_id, f"{time.time()-start:.0f}s"
        elif status == "error":
            return None, f"error: {data.get('error', 'unknown')}"
    return None, "timeout"


def create_excel(task_id):
    """add-insight로 엑셀 생성 (빈 보고서)"""
    resp = requests.post(f"{BASE_URL}/api/add-insight/{task_id}", json={
        "report": "(테스트용 보고서)",
        "summary_report": "(테스트용 요약)"
    })
    return resp.status_code == 200, resp.text[:200]


def download_excel(task_id):
    """엑셀 다운로드 후 파일 경로 반환"""
    r = requests.get(f"{BASE_URL}/api/download/{task_id}")
    if r.status_code != 200:
        return None, f"download failed: {r.status_code}"

    path = f"/tmp/pers_optest_{task_id[:8]}.xlsx"
    with open(path, "wb") as f:
        f.write(r.content)
    return path, f"{len(r.content)} bytes"


def verify_column_widths(excel_path, company_name):
    """Financials 시트 컬럼 너비 검증"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(excel_path)
    if "Financials" not in wb.sheetnames:
        return False, "Financials 시트 없음"

    ws = wb["Financials"]
    issues = []

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width or 8

        # 셀 내용 확인 (Excel 표시 포맷 기준)
        has_content = False
        max_content_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                # 숫자 타입은 Excel 표시 포맷 기준으로 길이 계산
                if isinstance(v, float):
                    display = f"{v:,.1f}" if abs(v) >= 100 else f"{v:.2f}"
                elif isinstance(v, int):
                    display = f"{v:,}"
                else:
                    display = str(v)
                if display.strip():
                    has_content = True
                    cl = sum(2 if ord(c) > 127 else 1 for c in display)
                    max_content_len = max(max_content_len, cl)

        if has_content:
            # 내용이 있는 컬럼이 너무 좁으면 문제
            if width < 10:
                issues.append(f"{letter}열: 내용있는데 width={width:.0f} (max_len={max_content_len})")
            # 항목명 컬럼(한글 많음)이 너무 좁으면
            if max_content_len > 15 and width < max_content_len * 0.7:
                issues.append(f"{letter}열: 내용({max_content_len}) 대비 width={width:.0f} 부족")
        else:
            # 빈 컬럼이 너무 넓으면
            if width > 5:
                issues.append(f"{letter}열: 빈 컬럼인데 width={width:.0f}")

    wb.close()

    if issues:
        return False, "; ".join(issues)
    return True, f"{ws.max_column}열 모두 정상"


def verify_excel_sheets(excel_path):
    """엑셀 시트 구조 검증"""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    sheets = wb.sheetnames

    required = ["Financials", "Frontdata"]
    missing = [s for s in required if s not in sheets]

    issues = []
    if missing:
        issues.append(f"필수 시트 누락: {missing}")

    # Financials 데이터 행 수
    if "Financials" in sheets:
        ws = wb["Financials"]
        if ws.max_row < 5:
            issues.append(f"Financials 데이터 부족: {ws.max_row}행")

    # Frontdata 검증
    if "Frontdata" in sheets:
        ws = wb["Frontdata"]
        if ws.max_row < 5:
            issues.append(f"Frontdata 데이터 부족: {ws.max_row}행")

    wb.close()

    if issues:
        return False, "; ".join(issues)
    return True, f"시트 {len(sheets)}개: {', '.join(sheets[:5])}..."


def verify_vcm_data(task_id):
    """VCM 데이터 검증 (자산등식, 대차균형)"""
    r = requests.get(f"{BASE_URL}/api/status/{task_id}")
    data = r.json()
    preview = data.get("preview_data", {})
    vcm = preview.get("vcm", [])
    vcm_display = preview.get("vcm_display", [])

    if not vcm or not vcm_display:
        return False, "VCM 데이터 없음"

    issues = []

    # FY 컬럼 찾기
    if not vcm_display:
        return False, "vcm_display 비어있음"

    first_row = vcm_display[0]
    fy_cols = [k for k in first_row.keys() if k.startswith("FY")]
    if not fy_cols:
        return False, "FY 컬럼 없음"

    latest_fy = sorted(fy_cols)[-1]

    # 항목 찾기 헬퍼
    def find_val(rows, item_name):
        for row in rows:
            name = row.get("항목", row.get("(단위: 백만원)", ""))
            if name == item_name:
                v = row.get(latest_fy)
                if v is not None:
                    try:
                        return int(float(str(v).replace(",", "")))
                    except:
                        pass
        return None

    # 자산 등식: 자산총계 = 유동자산 + 비유동자산
    total_assets = find_val(vcm_display, "자산총계")
    current_assets = find_val(vcm_display, "유동자산")
    noncurrent_assets = find_val(vcm_display, "비유동자산")

    if total_assets is not None and current_assets is not None and noncurrent_assets is not None:
        diff = abs(total_assets - (current_assets + noncurrent_assets))
        if diff > 2:
            issues.append(f"자산등식 불일치: {total_assets} != {current_assets}+{noncurrent_assets} (차이={diff})")

    # 대차 균형
    total_le = find_val(vcm_display, "부채와자본총계")
    total_liab = find_val(vcm_display, "부채총계")
    total_equity = find_val(vcm_display, "자본총계")

    if total_le is not None and total_liab is not None and total_equity is not None:
        diff = abs(total_le - (total_liab + total_equity))
        if diff > 2:
            issues.append(f"대차균형 불일치: {total_le} != {total_liab}+{total_equity} (차이={diff})")

    if issues:
        return False, "; ".join(issues)
    return True, f"{latest_fy}: 자산={total_assets}, 부채자본={total_le}"


def cleanup_task(task_id):
    """작업 취소/정리"""
    try:
        requests.post(f"{BASE_URL}/api/cancel/{task_id}")
    except:
        pass


# ================================================================
# 테스트 섹션들
# ================================================================

def test_1_server_health():
    """서버 상태 확인"""
    log("=" * 60)
    log("테스트 1: 서버 상태 확인")
    log("=" * 60)

    try:
        r = requests.get(f"{BASE_URL}/", timeout=5)
        record("서버_응답", r.status_code == 200, f"HTTP {r.status_code}")
    except Exception as e:
        record("서버_응답", False, str(e))
        return False

    # API 엔드포인트 접근성
    endpoints = [
        ("GET", "/api/auth/me"),
    ]
    for method, path in endpoints:
        try:
            if method == "GET":
                r = requests.get(f"{BASE_URL}{path}", timeout=5)
            record(f"엔드포인트_{path}", r.status_code in [200, 401, 403], f"HTTP {r.status_code}")
        except Exception as e:
            record(f"엔드포인트_{path}", False, str(e))

    return True


def test_2_fy_variations():
    """FY 연도 수별 테스트 (2년, 3년, 5년)"""
    log("=" * 60)
    log("테스트 2: FY 연도 수별 추출 + 컬럼 너비 검증")
    log("=" * 60)

    cases = [
        {"name": "E1", "corp_code": "00356361", "start": 2023, "end": 2024, "fy_count": 2},
        {"name": "E1", "corp_code": "00356361", "start": 2022, "end": 2024, "fy_count": 3},
        {"name": "E1", "corp_code": "00356361", "start": 2020, "end": 2024, "fy_count": 5},
    ]

    for case in cases:
        label = f"FY{case['fy_count']}년_{case['name']}"
        log(f"  추출 중: {case['name']} ({case['start']}~{case['end']})...")

        task_id, msg = extract_company(
            case["corp_code"], case["name"],
            start_year=case["start"], end_year=case["end"]
        )

        if not task_id:
            record(f"{label}_추출", False, msg)
            continue

        record(f"{label}_추출", True, msg)

        # VCM 검증
        ok, detail = verify_vcm_data(task_id)
        record(f"{label}_VCM", ok, detail)

        # 엑셀 생성 + 다운로드
        ok, detail = create_excel(task_id)
        if not ok:
            record(f"{label}_엑셀생성", False, detail)
            continue

        path, detail = download_excel(task_id)
        if not path:
            record(f"{label}_다운로드", False, detail)
            continue
        record(f"{label}_다운로드", True, detail)

        # 컬럼 너비 검증
        ok, detail = verify_column_widths(path, case["name"])
        record(f"{label}_컬럼너비", ok, detail)

        # 시트 구조 검증
        ok, detail = verify_excel_sheets(path)
        record(f"{label}_시트구조", ok, detail)

        # 파일 정리
        try:
            os.remove(path)
        except:
            pass


def test_3_edge_cases():
    """엣지케이스 기업 테스트"""
    log("=" * 60)
    log("테스트 3: 엣지케이스 기업 추출 + 검증")
    log("=" * 60)

    cases = [
        # 자본잠식
        {"name": "이노벡스", "corp_code": "01353848", "note": "자본잠식"},
        # 별도재무제표
        {"name": "이노메트리", "corp_code": "01011888", "note": "IS구조특성"},
        # 소형주
        {"name": "로지스몬", "corp_code": "01060735", "note": "소형KOSDAQ별도"},
    ]

    for case in cases:
        label = f"엣지_{case['note']}"
        log(f"  추출 중: {case['name']} ({case['note']})...")

        task_id, msg = extract_company(case["corp_code"], case["name"])

        if not task_id:
            record(f"{label}_추출", False, msg)
            continue

        record(f"{label}_추출", True, f"{case['name']} {msg}")

        # VCM 검증
        ok, detail = verify_vcm_data(task_id)
        record(f"{label}_VCM", ok, detail)

        # 엑셀 + 컬럼 너비
        ok, _ = create_excel(task_id)
        if ok:
            path, _ = download_excel(task_id)
            if path:
                ok, detail = verify_column_widths(path, case["name"])
                record(f"{label}_컬럼너비", ok, detail)
                try:
                    os.remove(path)
                except:
                    pass


def test_4_unlisted():
    """비상장사 테스트"""
    log("=" * 60)
    log("테스트 4: 비상장사 추출 테스트")
    log("=" * 60)

    # 비상장사 - 데이터 있는 기업
    cases = [
        {"name": "교보생명보험", "corp_code": "00120182", "note": "대형비상장"},
    ]

    for case in cases:
        label = f"비상장_{case['note']}"
        log(f"  추출 중: {case['name']} ({case['note']})...")

        task_id, msg = extract_company(case["corp_code"], case["name"])

        if not task_id:
            record(f"{label}_추출", False, msg)
            continue

        record(f"{label}_추출", True, f"{case['name']} {msg}")

        # VCM 검증
        ok, detail = verify_vcm_data(task_id)
        record(f"{label}_VCM", ok, detail)


def test_5_heartbeat_session():
    """Heartbeat 및 세션 시나리오 테스트"""
    log("=" * 60)
    log("테스트 5: Heartbeat / 세션 만료 시나리오")
    log("=" * 60)

    # 존재하지 않는 task heartbeat
    fake_id = "00000000-0000-0000-0000-000000000000"
    r = requests.post(f"{BASE_URL}/api/heartbeat/{fake_id}")
    data = r.json()
    record("hb_미존재_task",
           data.get("expired") == True and data.get("success") == False,
           f"expired={data.get('expired')}, success={data.get('success')}")

    # 실제 task로 heartbeat
    task_id, msg = extract_company("00356361", "E1", 2023, 2024)
    if task_id:
        r = requests.post(f"{BASE_URL}/api/heartbeat/{task_id}")
        data = r.json()
        record("hb_정상_task", data.get("success") == True, f"success={data.get('success')}")

        # 엑셀 생성
        create_excel(task_id)

        # 다운로드 후 재다운로드 (같은 파일 두 번 요청)
        r1 = requests.get(f"{BASE_URL}/api/download/{task_id}")
        r2 = requests.get(f"{BASE_URL}/api/download/{task_id}")
        record("다운로드_재요청",
               r1.status_code == 200 and r2.status_code == 200 and len(r1.content) == len(r2.content),
               f"1차={r1.status_code}({len(r1.content)}B), 2차={r2.status_code}({len(r2.content)}B)")

        # 다운로드-파일 엔드포인트 (filename 기반)
        status_r = requests.get(f"{BASE_URL}/api/status/{task_id}")
        filename = status_r.json().get("filename")
        if filename:
            r = requests.get(f"{BASE_URL}/api/download-file/{filename}")
            record("다운로드_파일명기반", r.status_code == 200, f"{filename} → {r.status_code}")

        # 취소 후 다운로드 시도
        requests.post(f"{BASE_URL}/api/cancel/{task_id}")
        time.sleep(1)
        r = requests.get(f"{BASE_URL}/api/download/{task_id}")
        # 취소된 작업은 COMPLETED_FILES에 있을 수 있음
        record("취소후_다운로드", r.status_code in [200, 404], f"HTTP {r.status_code}")
    else:
        record("hb_정상_task", False, f"추출 실패: {msg}")


def test_6_chatbot():
    """챗봇 기능 테스트"""
    log("=" * 60)
    log("테스트 6: PE 챗봇 기능")
    log("=" * 60)

    # 추출 먼저
    task_id, msg = extract_company("00356361", "E1", 2023, 2024)
    if not task_id:
        record("챗봇_추출", False, msg)
        return

    # 챗봇 초기화
    r = requests.post(f"{BASE_URL}/api/chat/init/{task_id}")
    if r.status_code == 200:
        record("챗봇_초기화", True, f"HTTP {r.status_code}")
    else:
        record("챗봇_초기화", False, f"HTTP {r.status_code}: {r.text[:100]}")
        return

    # 메시지 전송 (SSE 스트리밍)
    try:
        r = requests.post(
            f"{BASE_URL}/api/chat/message/{task_id}",
            json={"message": "이 회사의 매출 추이를 알려줘"},
            stream=True,
            timeout=60
        )

        response_text = ""
        for line in r.iter_lines(decode_unicode=True):
            if line and line.startswith("data: "):
                try:
                    chunk = json.loads(line[6:])
                    ctype = chunk.get("type", "")
                    if ctype in ("content", "token"):
                        response_text += chunk.get("text", "") or chunk.get("content", "")
                    elif ctype == "done":
                        break
                    elif ctype == "error":
                        response_text += f"[ERROR: {chunk.get('content', '')}]"
                        break
                except:
                    pass

        has_response = len(response_text) > 20
        record("챗봇_응답", has_response, f"응답 {len(response_text)}자")

        # 응답에 숫자/데이터가 포함되어 있는지 (할루시네이션 방지 확인)
        has_numbers = any(c.isdigit() for c in response_text)
        record("챗봇_데이터포함", has_numbers, f"숫자 포함={has_numbers}")
    except Exception as e:
        record("챗봇_응답", False, str(e))

    # 대화 히스토리
    r = requests.get(f"{BASE_URL}/api/chat/history/{task_id}")
    if r.status_code == 200:
        history = r.json()
        record("챗봇_히스토리", len(history) >= 2, f"대화 {len(history)}턴")
    else:
        record("챗봇_히스토리", False, f"HTTP {r.status_code}")

    # 잘못된 task_id로 챗봇 초기화
    r = requests.post(f"{BASE_URL}/api/chat/init/invalid-task-id-12345")
    record("챗봇_잘못된ID", r.status_code in [404, 400], f"HTTP {r.status_code}")


def test_7_concurrent():
    """동시성 테스트"""
    log("=" * 60)
    log("테스트 7: 동시성 테스트")
    log("=" * 60)

    # 동시 heartbeat 10개
    import threading

    results = []

    def send_hb(task_id_or_fake):
        try:
            r = requests.post(f"{BASE_URL}/api/heartbeat/{task_id_or_fake}", timeout=10)
            results.append(r.status_code)
        except Exception as e:
            results.append(str(e))

    fake_ids = [f"fake-{i:04d}-0000-0000-000000000000" for i in range(10)]
    threads = [threading.Thread(target=send_hb, args=(fid,)) for fid in fake_ids]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=15)

    all_200 = all(r == 200 for r in results)
    record("동시_heartbeat_10", all_200, f"결과: {results}")

    # 동시 상태 조회
    results2 = []

    def check_status(task_id):
        try:
            r = requests.get(f"{BASE_URL}/api/status/{task_id}", timeout=10)
            results2.append(r.status_code)
        except Exception as e:
            results2.append(str(e))

    threads2 = [threading.Thread(target=check_status, args=(fid,)) for fid in fake_ids]
    for t in threads2:
        t.start()
    for t in threads2:
        t.join(timeout=15)

    # 존재하지 않는 task는 404
    all_404 = all(r == 404 for r in results2)
    record("동시_상태조회_10", all_404, f"결과: {results2}")

    # 동시 추출 요청 2개 (서버 부하 테스트)
    log("  동시 추출 2개 시작...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        futures = [
            executor.submit(extract_company, "00356361", "E1", 2023, 2024),
            executor.submit(extract_company, "00126380", "삼성전자", 2023, 2024),
        ]

        results3 = []
        for f in concurrent.futures.as_completed(futures, timeout=400):
            tid, msg = f.result()
            results3.append((tid, msg))

    both_ok = all(tid is not None for tid, _ in results3)
    record("동시_추출_2개", both_ok,
           f"결과: {[f'{t[:8]}..({m})' if t else f'FAIL({m})' for t,m in results3]}")


def test_8_error_handling():
    """에러 처리 테스트"""
    log("=" * 60)
    log("테스트 8: 에러 처리 / 비정상 입력")
    log("=" * 60)

    # 잘못된 corp_code
    r = requests.post(f"{BASE_URL}/api/extract", json={
        "corp_code": "99999999",
        "corp_name": "존재하지않는회사",
        "start_year": 2023,
        "end_year": 2024
    })
    if r.status_code == 200:
        task_id = r.json().get("task_id")
        TASK_IDS.append(task_id)
        # 추출이 시작되지만 에러로 끝날 수 있음
        for _ in range(60):
            time.sleep(3)
            sr = requests.get(f"{BASE_URL}/api/status/{task_id}")
            sd = sr.json()
            if sd.get("status") in ["completed", "error"]:
                break
        final_status = sd.get("status")
        record("잘못된_corp_code", final_status in ["error", "completed"], f"status={final_status}")
    else:
        record("잘못된_corp_code", True, f"HTTP {r.status_code} (즉시 거부)")

    # 빈 요청
    r = requests.post(f"{BASE_URL}/api/extract", json={})
    record("빈_요청", r.status_code in [400, 422, 500], f"HTTP {r.status_code}")

    # 잘못된 연도 범위
    r = requests.post(f"{BASE_URL}/api/extract", json={
        "corp_code": "00356361",
        "corp_name": "E1",
        "start_year": 2030,
        "end_year": 2024
    })
    record("잘못된_연도", r.status_code in [200, 400, 422], f"HTTP {r.status_code}")
    if r.status_code == 200:
        tid = r.json().get("task_id")
        TASK_IDS.append(tid)

    # 존재하지 않는 task 상태 조회
    r = requests.get(f"{BASE_URL}/api/status/nonexistent-task")
    record("미존재_task_상태", r.status_code == 404, f"HTTP {r.status_code}")

    # 존재하지 않는 task 다운로드
    r = requests.get(f"{BASE_URL}/api/download/nonexistent-task")
    record("미존재_task_다운로드", r.status_code == 404, f"HTTP {r.status_code}")

    # 경로 조작 시도 (보안)
    r = requests.get(f"{BASE_URL}/api/download-file/../../../etc/passwd")
    record("경로조작_차단", r.status_code in [400, 403, 404], f"HTTP {r.status_code}")


def test_9_production():
    """프로덕션 서버 기본 테스트"""
    log("=" * 60)
    log("테스트 9: 프로덕션 서버 상태")
    log("=" * 60)

    prod_url = "http://localhost:8000"

    try:
        r = requests.get(f"{prod_url}/", timeout=5)
        record("프로덕션_서버", r.status_code == 200, f"HTTP {r.status_code}")
    except Exception as e:
        record("프로덕션_서버", False, str(e))
        return

    # 프로덕션에서 간단한 추출
    log("  프로덕션 추출 테스트 (E1)...")
    try:
        resp = requests.post(f"{prod_url}/api/extract", json={
            "corp_code": "00356361",
            "corp_name": "E1",
            "start_year": 2023,
            "end_year": 2024
        })
        if resp.status_code == 200:
            task_id = resp.json().get("task_id")
            start = time.time()
            while time.time() - start < 360:
                time.sleep(3)
                r = requests.get(f"{prod_url}/api/status/{task_id}")
                data = r.json()
                if data.get("status") == "completed":
                    record("프로덕션_추출", True, f"{time.time()-start:.0f}s")

                    # 프로덕션 엑셀 생성 + 다운로드
                    requests.post(f"{prod_url}/api/add-insight/{task_id}", json={
                        "report": "(프로덕션 테스트)",
                        "summary_report": "(테스트)"
                    })
                    r = requests.get(f"{prod_url}/api/download/{task_id}")
                    if r.status_code == 200:
                        path = f"/tmp/pers_prod_test.xlsx"
                        with open(path, "wb") as f:
                            f.write(r.content)
                        ok, detail = verify_column_widths(path, "E1")
                        record("프로덕션_컬럼너비", ok, detail)
                        try:
                            os.remove(path)
                        except:
                            pass
                    else:
                        record("프로덕션_다운로드", False, f"HTTP {r.status_code}")

                    # cleanup
                    requests.post(f"{prod_url}/api/cancel/{task_id}")
                    break
                elif data.get("status") == "error":
                    record("프로덕션_추출", False, data.get("error", "unknown"))
                    break
            else:
                record("프로덕션_추출", False, "timeout")
        else:
            record("프로덕션_추출", False, f"HTTP {resp.status_code}")
    except Exception as e:
        record("프로덕션_추출", False, str(e))


# ================================================================
# 메인
# ================================================================

def main():
    global BASE_URL

    if "--url" in sys.argv:
        idx = sys.argv.index("--url")
        BASE_URL = sys.argv[idx + 1]

    log("=" * 70)
    log("PERS 운용 종합 테스트 시작")
    log(f"  개발서버: {BASE_URL}")
    log(f"  시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log("=" * 70)

    start_time = time.time()

    # 순서대로 실행
    test_1_server_health()
    test_2_fy_variations()
    test_3_edge_cases()
    test_4_unlisted()
    test_5_heartbeat_session()
    test_6_chatbot()
    test_7_concurrent()
    test_8_error_handling()
    test_9_production()

    elapsed = time.time() - start_time

    # 정리
    for tid in TASK_IDS:
        cleanup_task(tid)

    # 결과 요약
    log("")
    log("=" * 70)
    log("테스트 결과 요약")
    log("=" * 70)

    passed = sum(1 for r in RESULTS if r["passed"])
    failed = sum(1 for r in RESULTS if not r["passed"])

    for r in RESULTS:
        status = "[PASS]" if r["passed"] else "[FAIL]"
        log(f"  {status} {r['name']}: {r['detail']}")

    log("")
    log(f"  총 {len(RESULTS)}개 테스트: {passed} PASS, {failed} FAIL ({elapsed:.0f}초 소요)")

    if failed > 0:
        log(f"\u2717 실패 항목 {failed}개!")
        log("")
        log("  실패 상세:")
        for r in RESULTS:
            if not r["passed"]:
                log(f"    - {r['name']}: {r['detail']}")
    else:
        log(f"\u2713 전체 통과!")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
