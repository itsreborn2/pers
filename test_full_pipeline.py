#!/usr/bin/env python3
"""
PERS 통합 파이프라인 테스트 스크립트
====================================
DART 추출 → 데이터 검증 → AI 분석 → 기업 리서치 → 엑셀 다운로드까지 전체 파이프라인 테스트.

사용법:
    python3 test_full_pipeline.py                     # 기본 (삼성전자, 빠른 테스트)
    python3 test_full_pipeline.py --full              # 전체 테스트 (AI분석 + 리서치 포함)
    python3 test_full_pipeline.py --company 농업회사법인조인 --corp-code 00474205
    python3 test_full_pipeline.py --company E1 --corp-code 00356361 --full
"""

import asyncio
import aiohttp
import json
import os
import sys
import time
import argparse
from datetime import datetime

# ============================================================
# 설정
# ============================================================
BASE_URL = os.environ.get("PERS_URL", "http://localhost:8001")
DEFAULT_CORP_CODE = "00126380"  # 삼성전자
DEFAULT_CORP_NAME = "삼성전자"
DEFAULT_START_YEAR = 2022
DEFAULT_END_YEAR = 2024
EXTRACT_TIMEOUT = 360  # 추출 타임아웃 (초)
ANALYSIS_TIMEOUT = 300  # AI 분석 타임아웃 (초)
RESEARCH_TIMEOUT = 600  # 기업 리서치 타임아웃 (초)

# 테스트 결과 저장
results = {
    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "base_url": BASE_URL,
    "tests": {}
}


def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = {"INFO": "  ", "PASS": "✓ ", "FAIL": "✗ ", "WARN": "⚠ ", "STEP": "→ "}
    print(f"[{ts}] {prefix.get(level, '  ')}{msg}")


def record(test_name, passed, detail=""):
    results["tests"][test_name] = {"passed": passed, "detail": detail}
    log(f"{test_name}: {detail}", "PASS" if passed else "FAIL")


# ============================================================
# 1. 서버 상태 확인
# ============================================================
async def test_server_health(session):
    log("=" * 60)
    log("1단계: 서버 상태 확인", "STEP")
    log("=" * 60)
    try:
        async with session.get(f"{BASE_URL}/") as resp:
            if resp.status == 200:
                record("서버_상태", True, f"HTTP {resp.status} OK")
                return True
            else:
                record("서버_상태", False, f"HTTP {resp.status}")
                return False
    except Exception as e:
        record("서버_상태", False, str(e))
        return False


# ============================================================
# 2. DART 추출 테스트
# ============================================================
async def test_extraction(session, corp_code, corp_name, start_year, end_year):
    log("=" * 60)
    log(f"2단계: DART 재무제표 추출 ({corp_name})", "STEP")
    log("=" * 60)

    # 2-1. 추출 시작
    payload = {
        "corp_code": corp_code,
        "corp_name": corp_name,
        "start_year": start_year,
        "end_year": end_year
    }
    try:
        async with session.post(f"{BASE_URL}/api/extract", json=payload) as resp:
            data = await resp.json()
            if resp.status != 200 or not data.get("task_id"):
                record("추출_시작", False, f"HTTP {resp.status}: {data}")
                return None
            task_id = data["task_id"]
            record("추출_시작", True, f"task_id={task_id[:12]}...")
    except Exception as e:
        record("추출_시작", False, str(e))
        return None

    # 2-2. 상태 폴링
    start_time = time.time()
    poll_count = 0
    last_progress = -1

    while time.time() - start_time < EXTRACT_TIMEOUT:
        await asyncio.sleep(4)
        poll_count += 1

        try:
            # heartbeat 매 3번째 폴링마다
            if poll_count % 3 == 0:
                async with session.post(f"{BASE_URL}/api/heartbeat/{task_id}") as hb_resp:
                    hb_data = await hb_resp.json()
                    if hb_data.get("expired"):
                        record("heartbeat", False, "작업이 조기 만료됨!")
                        return None

            async with session.get(f"{BASE_URL}/api/status/{task_id}") as resp:
                status_data = await resp.json()
                progress = status_data.get("progress", 0)
                status = status_data.get("status", "")
                msg = status_data.get("message", "")

                if progress != last_progress:
                    log(f"  진행률: {progress}% - {msg}")
                    last_progress = progress

                if status == "completed":
                    elapsed = time.time() - start_time
                    record("추출_완료", True, f"{elapsed:.1f}초 소요")
                    return {"task_id": task_id, "status_data": status_data}
                elif status in ("error", "cancelled", "failed"):
                    record("추출_완료", False, f"상태={status}: {msg}")
                    return None
        except Exception as e:
            log(f"  폴링 오류: {e}", "WARN")

    record("추출_완료", False, f"타임아웃 ({EXTRACT_TIMEOUT}초)")
    return None


# ============================================================
# 3. 데이터 검증 테스트
# ============================================================
async def test_data_validation(session, task_id, status_data):
    log("=" * 60)
    log("3단계: 재무 데이터 검증", "STEP")
    log("=" * 60)

    preview = status_data.get("preview_data", {})

    # 3-1. VCM 데이터 존재 확인
    vcm = preview.get("vcm", [])
    vcm_display = preview.get("vcm_display", [])

    if not vcm:
        record("VCM_데이터", False, "vcm 데이터 없음")
    else:
        record("VCM_데이터", True, f"vcm={len(vcm)}행, vcm_display={len(vcm_display)}행")

    # 3-2. 재무상태표 데이터 확인
    bs_data = preview.get("bs", [])
    if not bs_data:
        record("재무상태표_데이터", False, "BS 데이터 없음")
    else:
        record("재무상태표_데이터", True, f"{len(bs_data)}행")

    # 3-3. 손익계산서 데이터 확인
    is_data = preview.get("is", [])
    if not is_data:
        record("손익계산서_데이터", False, "IS 데이터 없음")
    else:
        record("손익계산서_데이터", True, f"{len(is_data)}행")

    # 3-4. 현금흐름표 데이터 확인
    cf_data = preview.get("cf", [])
    if not cf_data:
        record("현금흐름표_데이터", False, "CF 데이터 없음")
    else:
        record("현금흐름표_데이터", True, f"{len(cf_data)}행")

    # 3-5. VCM 수치 검증 (자산=유동+비유동, 자산=부채+자본)
    if vcm_display:
        _validate_vcm_equations(vcm_display)
    else:
        record("자산_등식_검증", False, "VCM display 데이터 없음")
        record("대차_균형_검증", False, "VCM display 데이터 없음")


def _validate_vcm_equations(vcm_display):
    """VCM 숫자 검증: 자산 등식, 대차 균형"""
    # FY 컬럼 찾기
    if not vcm_display:
        return

    fy_cols = [k for k in vcm_display[0].keys() if k.startswith("FY")]
    if not fy_cols:
        record("자산_등식_검증", False, "FY 컬럼 없음")
        return

    latest_fy = sorted(fy_cols)[-1]

    def find_val(name_pattern):
        for row in vcm_display:
            item = str(row.get("항목", "")).strip()
            if item == name_pattern:
                v = row.get(latest_fy)
                if v is not None and v != "" and v != "-":
                    try:
                        return int(float(str(v).replace(",", "").replace(" ", "")))
                    except (ValueError, TypeError):
                        return None
        return None

    # 자산 등식: 자산총계 = 유동자산 + 비유동자산 (+ 매각예정자산)
    total_assets = find_val("자산총계")
    current_assets = find_val("유동자산")
    noncurrent_assets = find_val("비유동자산")

    if total_assets is not None and current_assets is not None and noncurrent_assets is not None:
        calc = current_assets + noncurrent_assets
        diff = abs(total_assets - calc)
        if diff <= 2:
            record("자산_등식_검증", True, f"{latest_fy}: {total_assets:,} = {current_assets:,} + {noncurrent_assets:,}")
        else:
            # 매각예정자산 확인
            disposal = find_val("매각예정자산") or 0
            calc2 = current_assets + noncurrent_assets + disposal
            diff2 = abs(total_assets - calc2)
            if diff2 <= 2:
                record("자산_등식_검증", True, f"{latest_fy}: {total_assets:,} = {current_assets:,} + {noncurrent_assets:,} + 매각예정 {disposal:,}")
            else:
                record("자산_등식_검증", False, f"{latest_fy}: 자산총계 {total_assets:,} ≠ 유동 {current_assets:,} + 비유동 {noncurrent_assets:,} (차이: {diff})")
    else:
        record("자산_등식_검증", False, f"값 누락: 자산총계={total_assets}, 유동={current_assets}, 비유동={noncurrent_assets}")

    # 대차 균형: 자산총계 = 부채총계 + 자본총계
    total_liabilities = find_val("부채총계")
    total_equity = find_val("자본총계")
    balance_total = find_val("부채와자본총계") or total_assets

    if total_liabilities is not None and total_equity is not None and balance_total is not None:
        calc = total_liabilities + total_equity
        diff = abs(balance_total - calc)
        if diff <= 2:
            record("대차_균형_검증", True, f"{latest_fy}: {balance_total:,} = 부채 {total_liabilities:,} + 자본 {total_equity:,}")
        else:
            record("대차_균형_검증", False, f"{latest_fy}: {balance_total:,} ≠ 부채 {total_liabilities:,} + 자본 {total_equity:,} (차이: {diff})")
    else:
        record("대차_균형_검증", False, f"값 누락: 부채총계={total_liabilities}, 자본총계={total_equity}")


# ============================================================
# 4. AI 분석 테스트
# ============================================================
async def test_ai_analysis(session, task_id):
    log("=" * 60)
    log("4단계: 재무분석 AI 테스트", "STEP")
    log("=" * 60)

    # 4-1. 분석 시작
    try:
        async with session.post(f"{BASE_URL}/api/analyze/{task_id}") as resp:
            data = await resp.json()
            if resp.status != 200:
                record("AI분석_시작", False, f"HTTP {resp.status}: {data}")
                return None
            record("AI분석_시작", True, data.get("message", ""))
    except Exception as e:
        record("AI분석_시작", False, str(e))
        return None

    # 4-2. 분석 상태 폴링
    start_time = time.time()
    last_progress = -1

    while time.time() - start_time < ANALYSIS_TIMEOUT:
        await asyncio.sleep(5)
        try:
            async with session.get(f"{BASE_URL}/api/analyze-status/{task_id}") as resp:
                data = await resp.json()
                progress = data.get("progress", 0)
                status = data.get("status", "")
                msg = data.get("message", "")

                if progress != last_progress:
                    log(f"  AI분석 진행률: {progress}% - {msg}")
                    last_progress = progress

                if status == "completed":
                    elapsed = time.time() - start_time
                    result = data.get("result", {})
                    report = result.get("report", "")
                    summary = result.get("summary_report", "")
                    record("AI분석_완료", True, f"{elapsed:.1f}초 소요, 보고서 {len(report)}자, 요약 {len(summary)}자")

                    # 4-3. add-insight 호출 (엑셀에 AI 시트 추가)
                    if report:
                        await _add_insight(session, task_id, report, summary)
                    return result

                elif status in ("error", "cancelled", "failed"):
                    record("AI분석_완료", False, f"상태={status}: {msg}")
                    return None
        except Exception as e:
            log(f"  폴링 오류: {e}", "WARN")

    record("AI분석_완료", False, f"타임아웃 ({ANALYSIS_TIMEOUT}초)")
    return None


async def _add_insight(session, task_id, report, summary_report):
    """AI 분석 결과를 엑셀에 추가"""
    try:
        payload = {"report": report, "summary_report": summary_report}
        async with session.post(f"{BASE_URL}/api/add-insight/{task_id}", json=payload) as resp:
            data = await resp.json()
            if resp.status == 200:
                record("엑셀_AI시트_추가", True, data.get("message", ""))
            else:
                record("엑셀_AI시트_추가", False, f"HTTP {resp.status}: {data}")
    except Exception as e:
        record("엑셀_AI시트_추가", False, str(e))


# ============================================================
# 5. 기업 리서치 테스트
# ============================================================
async def test_super_research(session, task_id):
    log("=" * 60)
    log("5단계: 기업 리서치 테스트", "STEP")
    log("=" * 60)

    # 5-1. 리서치 시작
    try:
        async with session.post(f"{BASE_URL}/api/super-research/{task_id}") as resp:
            data = await resp.json()
            if resp.status != 200:
                record("기업리서치_시작", False, f"HTTP {resp.status}: {data}")
                return None
            record("기업리서치_시작", True, data.get("message", ""))
    except Exception as e:
        record("기업리서치_시작", False, str(e))
        return None

    # 5-2. 리서치 상태 폴링
    start_time = time.time()
    last_progress = -1

    while time.time() - start_time < RESEARCH_TIMEOUT:
        await asyncio.sleep(5)
        try:
            async with session.get(f"{BASE_URL}/api/super-research-status/{task_id}") as resp:
                data = await resp.json()
                progress = data.get("progress", 0)
                status = data.get("status", "")
                msg = data.get("message", "")

                if progress != last_progress:
                    log(f"  리서치 진행률: {progress}% - {msg}")
                    last_progress = progress

                if status == "completed":
                    elapsed = time.time() - start_time
                    result_data = data.get("result", {})
                    record("기업리서치_완료", True, f"{elapsed:.1f}초 소요")
                    return result_data
                elif status in ("error", "cancelled", "failed"):
                    record("기업리서치_완료", False, f"상태={status}: {msg}")
                    return None
        except Exception as e:
            log(f"  폴링 오류: {e}", "WARN")

    record("기업리서치_완료", False, f"타임아웃 ({RESEARCH_TIMEOUT}초)")
    return None


# ============================================================
# 6. 엑셀 다운로드 테스트
# ============================================================
async def test_excel_download(session, task_id, expect_ai_sheets=False):
    stage = "최종" if expect_ai_sheets else "기본"
    log("=" * 60)
    log(f"6단계: 엑셀 다운로드 테스트 ({stage})", "STEP")
    log("=" * 60)

    # 6-1. 다운로드
    try:
        async with session.get(f"{BASE_URL}/api/download/{task_id}") as resp:
            if resp.status != 200:
                body = await resp.text()
                record(f"엑셀다운로드_{stage}", False, f"HTTP {resp.status}: {body[:200]}")
                return False

            content = await resp.read()
            if len(content) < 1000:
                record(f"엑셀다운로드_{stage}", False, f"파일 크기 비정상: {len(content)} bytes")
                return False

            # 임시 파일에 저장
            tmp_path = f"/tmp/pers_test_{task_id[:8]}_{stage}.xlsx"
            with open(tmp_path, "wb") as f:
                f.write(content)

            record(f"엑셀다운로드_{stage}", True, f"{len(content):,} bytes → {tmp_path}")
    except Exception as e:
        record(f"엑셀다운로드_{stage}", False, str(e))
        return False

    # 6-2. 엑셀 시트 검증
    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        log(f"  시트 목록: {sheets}")

        # 필수 시트 확인 (기업개황은 company_info 전달 시에만 생성)
        required_sheets = ["Financials", "Frontdata"]
        optional_sheets = ["기업개황"]
        missing = [s for s in required_sheets if s not in sheets]
        if missing:
            record(f"엑셀시트_{stage}", False, f"필수 시트 누락: {missing}")
        else:
            opt_present = [s for s in optional_sheets if s in sheets]
            record(f"엑셀시트_{stage}", True, f"총 {len(sheets)}개 시트 (필수 OK, 기업개황={'있음' if opt_present else '없음(정상)'})")

        # AI 시트 확인 (full 테스트)
        if expect_ai_sheets:
            ai_sheets = [s for s in sheets if "재무분석" in s]
            if ai_sheets:
                record("엑셀_AI시트_존재", True, f"AI 시트: {ai_sheets}")
            else:
                record("엑셀_AI시트_존재", False, "AI 분석 시트 없음")

        # Financials 시트 데이터 확인
        fin_ws = wb["Financials"]
        row_count = 0
        for row in fin_ws.iter_rows(min_row=1, max_row=200):
            if any(cell.value for cell in row):
                row_count += 1
        if row_count > 5:
            record(f"Financials시트_{stage}", True, f"{row_count}행 데이터")
        else:
            record(f"Financials시트_{stage}", False, f"데이터 부족: {row_count}행")

        wb.close()
        # 임시 파일 정리
        os.remove(tmp_path)
        return True

    except ImportError:
        record(f"엑셀시트_{stage}", False, "openpyxl 미설치")
        return False
    except Exception as e:
        record(f"엑셀시트_{stage}", False, str(e))
        return False


# ============================================================
# 7. Heartbeat 만료 복구 테스트
# ============================================================
async def test_heartbeat_expiry(session, task_id):
    log("=" * 60)
    log("7단계: Heartbeat 만료 응답 테스트", "STEP")
    log("=" * 60)

    # 존재하는 task에 heartbeat
    try:
        async with session.post(f"{BASE_URL}/api/heartbeat/{task_id}") as resp:
            data = await resp.json()
            if data.get("success") is True:
                record("heartbeat_정상", True, "success=true")
            elif data.get("expired"):
                record("heartbeat_정상", False, "이미 만료됨 (예상치 못한 상태)")
            else:
                record("heartbeat_정상", False, f"예상치 못한 응답: {data}")
    except Exception as e:
        record("heartbeat_정상", False, str(e))

    # 존재하지 않는 task에 heartbeat (만료 응답 확인)
    try:
        fake_id = "00000000-0000-0000-0000-000000000000"
        async with session.post(f"{BASE_URL}/api/heartbeat/{fake_id}") as resp:
            data = await resp.json()
            if data.get("expired") is True and data.get("success") is False:
                record("heartbeat_만료감지", True, "expired=true, success=false")
            else:
                record("heartbeat_만료감지", False, f"예상치 못한 응답: {data}")
    except Exception as e:
        record("heartbeat_만료감지", False, str(e))


# ============================================================
# 메인 실행
# ============================================================
async def main():
    parser = argparse.ArgumentParser(description="PERS 통합 파이프라인 테스트")
    parser.add_argument("--company", default=DEFAULT_CORP_NAME, help="회사명")
    parser.add_argument("--corp-code", default=DEFAULT_CORP_CODE, help="DART corp_code")
    parser.add_argument("--start-year", type=int, default=DEFAULT_START_YEAR, help="시작 연도")
    parser.add_argument("--end-year", type=int, default=DEFAULT_END_YEAR, help="종료 연도")
    parser.add_argument("--full", action="store_true", help="AI분석 + 기업리서치 포함 전체 테스트")
    parser.add_argument("--url", default=BASE_URL, help="서버 URL")
    args = parser.parse_args()

    base_url = args.url
    # 모듈 레벨 변수 업데이트를 위해 globals 사용
    globals()["BASE_URL"] = base_url
    results["base_url"] = base_url

    log("=" * 60)
    log(f"PERS 통합 파이프라인 테스트 시작")
    log(f"  서버: {BASE_URL}")
    log(f"  기업: {args.company} ({args.corp_code})")
    log(f"  기간: {args.start_year}~{args.end_year}")
    log(f"  모드: {'전체 (AI + 리서치)' if args.full else '기본 (추출 + 검증 + 다운로드)'}")
    log("=" * 60)

    timeout = aiohttp.ClientTimeout(total=60)
    async with aiohttp.ClientSession(timeout=timeout) as session:

        # 1. 서버 상태 확인
        if not await test_server_health(session):
            log("서버 접속 불가 — 테스트 중단", "FAIL")
            return

        # 2. DART 추출 테스트
        result = await test_extraction(session, args.corp_code, args.company, args.start_year, args.end_year)
        if not result:
            log("추출 실패 — 테스트 중단", "FAIL")
            _print_summary()
            return

        task_id = result["task_id"]
        status_data = result["status_data"]

        # 3. 데이터 검증 테스트
        await test_data_validation(session, task_id, status_data)

        # 7. Heartbeat 테스트 (추출 직후)
        await test_heartbeat_expiry(session, task_id)

        if args.full:
            # 4. AI 분석 테스트
            analysis_result = await test_ai_analysis(session, task_id)

            # 5. 기업 리서치 테스트
            research_result = await test_super_research(session, task_id)

            # 6. 최종 엑셀 다운로드 (AI + 리서치 포함)
            await test_excel_download(session, task_id, expect_ai_sheets=True)
        else:
            # 6. 기본 엑셀 다운로드 (추출만)
            # AI분석 없이 다운로드하면 file_path=None이므로 add-insight 먼저 필요
            # 빠른 테스트를 위해 빈 보고서로 엑셀 생성
            log("  (빠른 테스트: 빈 AI 보고서로 엑셀 생성)")
            await _add_insight(session, task_id, "(테스트용 보고서)", "(테스트용 요약)")
            await test_excel_download(session, task_id, expect_ai_sheets=False)

        # 작업 정리
        try:
            async with session.post(f"{BASE_URL}/api/cancel/{task_id}") as resp:
                log(f"  작업 정리: {resp.status}")
        except:
            pass

    _print_summary()


def _print_summary():
    log("")
    log("=" * 60)
    log("테스트 결과 요약")
    log("=" * 60)

    total = len(results["tests"])
    passed = sum(1 for t in results["tests"].values() if t["passed"])
    failed = total - passed

    for name, info in results["tests"].items():
        status = "PASS" if info["passed"] else "FAIL"
        log(f"  [{status}] {name}: {info['detail']}")

    log("")
    log(f"  총 {total}개 테스트: {passed} PASS, {failed} FAIL")

    if failed > 0:
        log(f"  ❌ {failed}개 실패!", "FAIL")
        sys.exit(1)
    else:
        log(f"  전체 통과!", "PASS")
        sys.exit(0)


if __name__ == "__main__":
    asyncio.run(main())
